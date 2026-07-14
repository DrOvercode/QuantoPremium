"""
core.py — Quanto Financial Intelligence Platform v1.0
=======================================================
Run:  uvicorn core:app --reload --port 8000
Then: http://localhost:8000

Requirements:
    python -m pip install fastapi uvicorn python-multipart pdfplumber pytesseract pillow openpyxl ollama numpy openai google-generativeai anthropic

OCR PROVIDERS (configure via environment variables):
    QUANTO_OCR_PROVIDER  = "claude" | "gemini" | "openai" | "local"  (default: claude)
    ANTHROPIC_API_KEY    = your Anthropic API key (recommended — most accurate, and PDFs are
                            sent natively without needing pdf2image/poppler installed)
    GEMINI_API_KEY       = your Google Gemini API key
    OPENAI_API_KEY       = your OpenAI API key
    (if no key is set, falls back to local pdfplumber + pytesseract)

NOTE: This build contains only the Statement Generator and Forecasting Engine.
The multi-entity Financial Consolidation Engine has been removed.
"""
from typing import Optional
import io, re, json, math, tempfile, statistics, base64, os, uuid
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse


app = FastAPI(title="Quanto", version="1.0.0")
OLLAMA_MODEL = "gemini-3-flash-preview:latest"
OUTPUT_DIR   = Path(tempfile.gettempdir()) / "quanto_outputs"
OUTPUT_DIR.mkdir(exist_ok=True)
from google import genai

OCR_PROVIDER = "gemini"

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
GEMINI_API_KEY = os.environ.get(
    "GEMINI_API_KEY",
    "AQ.Ab8RN6LnFOwbXFwSvH0K1A4X13d85DuEM4PtZ3bcQU6wiqhzrQ"
)

print("OCR_PROVIDER:", OCR_PROVIDER)
print("GEMINI KEY EXISTS:", bool(GEMINI_API_KEY))





print("OCR_PROVIDER:", OCR_PROVIDER)
print("GEMINI KEY EXISTS:", bool(GEMINI_API_KEY))
# =====================================================================================
# FREE-PLAN USAGE LIMITS
# =====================================================================================
# QUANTO_PLAN controls whether usage limits are enforced. "paid" plans are unrestricted.
QUANTO_PLAN = os.environ.get("QUANTO_PLAN", "free").lower()

FREE_STATEMENT_LIMIT = 10
FREE_FORECAST_LIMIT = 10

# Usage counters are persisted to disk (next to this script) so they survive app restarts.
USAGE_FILE = Path(__file__).resolve().parent / "quanto_usage.json"

STATEMENT_LIMIT_MESSAGE = "You have reached the free plan limit of 2 financial statements. Please upgrade your plan to continue."
FORECAST_LIMIT_MESSAGE = "You have reached the free plan limit of 2 forecasts. Please upgrade your plan to continue."


def _load_usage() -> Dict[str, int]:
    """Load persisted usage counters from disk. Returns defaults if the file is missing or unreadable."""
    default = {"financial_statements_generated": 0, "forecasts_generated": 0}
    if USAGE_FILE.exists():
        try:
            with open(USAGE_FILE, "r") as f:
                data = json.load(f)
            default.update({k: data.get(k, default[k]) for k in default})
        except Exception:
            pass
    return default


def _save_usage(usage: Dict[str, int]) -> None:
    """Persist usage counters to disk."""
    try:
        with open(USAGE_FILE, "w") as f:
            json.dump(usage, f)
    except Exception:
        pass


def _increment_usage(key: str) -> None:
    """Increment a single usage counter and persist it. Only called after successful generation."""
    usage = _load_usage()
    usage[key] = usage.get(key, 0) + 1
    _save_usage(usage)


def _enforce_free_plan_limit(usage_key: str, limit: int, error_message: str) -> None:
    """Raises an HTTPException (stopping execution) if a free-plan user has hit the given limit.
    Paid plans are never restricted."""
    if QUANTO_PLAN == "paid":
        return
    usage = _load_usage()
    if usage.get(usage_key, 0) >= limit:
        raise HTTPException(status_code=403, detail=error_message)

STATEMENTS = {
    "income_statement":      {"label":"Income Statement",                    "aliases":"P&L · Statement of Earnings · Statement of Operations","icon":"📈","category":"statements","sources":["trial_balance"]},
    "balance_sheet":         {"label":"Balance Sheet",                       "aliases":"Statement of Financial Position",                       "icon":"⚖️","category":"statements","sources":["trial_balance"]},
    "retained_earnings":     {"label":"Statement of Retained Earnings",      "aliases":"RE Rollforward",                                        "icon":"🔄","category":"statements","sources":["trial_balance"]},
    "equity_statement":      {"label":"Statement of Shareholders' Equity",   "aliases":"Changes in Equity · Owners' Equity",                    "icon":"🏛️","category":"statements","sources":["trial_balance"]},
    "trial_balance":         {"label":"Trial Balance",                       "aliases":"Adjusted / Unadjusted TB",                              "icon":"📋","category":"statements","sources":["trial_balance"]},
    "ratio_analysis":        {"label":"Financial Ratio Analysis",            "aliases":"Liquidity · Solvency · Profitability · Efficiency",      "icon":"📊","category":"statements","sources":["trial_balance"]},
    "liquidity_report":      {"label":"Liquidity Report",                    "aliases":"Current Ratio · Quick Ratio · Cash Ratio",               "icon":"💧","category":"statements","sources":["trial_balance"]},
    "solvency_report":       {"label":"Solvency Report",                     "aliases":"Debt-to-Equity · Leverage Analysis",                     "icon":"🏗️","category":"statements","sources":["trial_balance"]},
    "profitability_report":  {"label":"Profitability Report",                "aliases":"Margins · ROA · ROE",                                    "icon":"💰","category":"statements","sources":["trial_balance"]},
    "working_capital":       {"label":"Working Capital Report",              "aliases":"Cash Conversion Cycle · DIO · DPO",                      "icon":"⚙️","category":"statements","sources":["trial_balance"]},
    "cash_flow_statement":   {"label":"Cash Flow Statement",                 "aliases":"Operating · Investing · Financing Activities",           "icon":"💵","category":"statements","sources":["prior_balance_sheet","current_balance_sheet","income_statement","transaction_details"]},
    "ar_aging":              {"label":"Accounts Receivable Aging",           "aliases":"Customer Aging · Receivables Schedule",                  "icon":"📥","category":"statements","sources":["customer_invoices","due_dates","customer_balances"]},
    "ap_aging":              {"label":"Accounts Payable Aging",              "aliases":"Supplier Aging · Payables Schedule",                     "icon":"📤","category":"statements","sources":["supplier_invoices","due_dates","supplier_balances"]},
    "fixed_asset_schedule":  {"label":"Fixed Asset Schedule",                "aliases":"PPE Schedule · Depreciation Schedule",                   "icon":"🏭","category":"statements","sources":["fixed_asset_register","purchase_dates","depreciation_rates","useful_lives"]},
    "inventory_schedule":    {"label":"Inventory Schedule",                  "aliases":"Stock Schedule · FIFO / Weighted Average",               "icon":"📦","category":"statements","sources":["inventory_records","quantities","costing_method"]},
    "inventory_rollforward": {"label":"Inventory Rollforward",               "aliases":"Inventory Movement · Opening/Closing Stock",             "icon":"🔁","category":"statements","sources":["opening_inventory","purchases","sales","adjustments"]},
    "equity_rollforward":    {"label":"Equity Rollforward",                  "aliases":"Capital Account Movement",                               "icon":"📈","category":"statements","sources":["share_issues","dividends","owner_contributions","retained_earnings_movements"]},
    "debt_schedule":         {"label":"Debt Schedule",                       "aliases":"Loan Schedule · Debt Repayment Plan",                    "icon":"🏦","category":"statements","sources":["loan_agreements","repayment_schedules","interest_rates"]},
    "lease_schedule":        {"label":"Lease Schedule",                      "aliases":"IFRS 16 · Right-of-Use Assets",                          "icon":"🏢","category":"statements","sources":["lease_contracts","payment_schedules","lease_terms"]},
    "bank_reconciliation":   {"label":"Bank Reconciliation",                 "aliases":"Cash Reconciliation · Bank Recon",                       "icon":"🏧","category":"statements","sources":["bank_statements","cash_ledger"]},
    "account_reconciliation":{"label":"Account Reconciliations",            "aliases":"Subledger Reconciliation · Control Account",             "icon":"🔍","category":"statements","sources":["external_statements","subledgers","supporting_documents"]},
    "audit_working_papers":  {"label":"Audit Working Papers",                "aliases":"Lead Schedules · Audit File",                            "icon":"📝","category":"statements","sources":["general_ledger","lead_schedules","supporting_documentation"]},
    "notes_financial_stmts": {"label":"Notes to Financial Statements",       "aliases":"Disclosures · Accounting Policies",                      "icon":"📄","category":"statements","sources":["management_disclosures","accounting_policies","legal_information"]},
    "related_party":         {"label":"Related Party Disclosure Schedule",   "aliases":"RPT Schedule · Related Transactions",                    "icon":"🤝","category":"statements","sources":["related_party_transaction_data"]},
    "deferred_tax":          {"label":"Deferred Tax Schedule",               "aliases":"DTA · DTL · Tax Timing Differences",                     "icon":"🧮","category":"statements","sources":["tax_calculations","tax_returns","temporary_differences"]},
    "tax_provision":         {"label":"Tax Provision Workpapers",            "aliases":"Current & Deferred Tax · ETR Analysis",                  "icon":"💼","category":"statements","sources":["tax_returns","tax_adjustments","tax_rates"]},
    "revenue_recognition":   {"label":"Revenue Recognition Schedule",        "aliases":"ASC 606 · IFRS 15 · Contract Revenue",                   "icon":"📊","category":"statements","sources":["contracts","invoices","performance_obligations"]},
    "prepaid_expense":       {"label":"Prepaid Expense Schedule",            "aliases":"Prepayments · Deferred Charges",                         "icon":"⏳","category":"statements","sources":["payment_records","amortization_periods"]},
    "accrual_schedule":      {"label":"Accrual Schedule",                    "aliases":"Accrued Liabilities · Outstanding Obligations",          "icon":"📋","category":"statements","sources":["outstanding_invoices","contracts","unpaid_obligations"]},
    "employee_benefits":     {"label":"Employee Benefit Schedule",           "aliases":"Payroll · Pension · Benefits",                           "icon":"👥","category":"statements","sources":["payroll_records","pension_data","benefit_data"]},
    "forecast_growth":       {"label":"Growth Rate Forecast",                "aliases":"CAGR · Weighted · Trend · 1/3/5 Year",                   "icon":"🚀","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_driver":       {"label":"Driver-Based Forecast",               "aliases":"Revenue Drivers · Expense Drivers · Working Capital",    "icon":"🎯","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_three_stmt":   {"label":"Three-Statement Forecast Model",      "aliases":"Linked IS · BS · Cash Flow",                             "icon":"🔗","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_cashflow":     {"label":"Cash Flow Forecast",                  "aliases":"Operating · Investing · Financing · Runway",             "icon":"💵","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_scenarios":    {"label":"Scenario Analysis",                   "aliases":"Base · Best · Worst Case · Probability",                 "icon":"🎭","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_valuation":    {"label":"Valuation Model",                     "aliases":"DCF · Terminal Value · Enterprise Value · Equity Value",  "icon":"💎","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_risk":         {"label":"Risk Analysis Report",                "aliases":"Concentration · Liquidity · Debt · Burn · Score",        "icon":"⚠️","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_stakeholder":  {"label":"Stakeholder Analysis",                "aliases":"Owner · CFO · Investor · Bank · Auditor",                "icon":"👥","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_narrative":    {"label":"AI Narrative Insights",               "aliases":"Explainable Forecast Commentary",                       "icon":"🤖","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_full":         {"label":"Full Forecasting Package",            "aliases":"All 15 Phases — Complete Model",                         "icon":"📦","category":"forecast","sources":["trial_balance_3yr_min"]},
}

SOURCE_LABELS = {
    "trial_balance":                  ("Trial Balance",                  "The adjusted or unadjusted trial balance for the period"),
    "prior_balance_sheet":            ("Prior-Year Balance Sheet",       "Balance sheet from the previous fiscal year"),
    "current_balance_sheet":          ("Current-Year Balance Sheet",     "Balance sheet for the current fiscal year"),
    "income_statement":               ("Income Statement",               "Income statement / P&L for the period"),
    "transaction_details":            ("Transaction Details",            "Detailed list of cash transactions for the period"),
    "customer_invoices":              ("Customer Invoices",              "All outstanding customer invoices"),
    "due_dates":                      ("Due Dates",                      "Invoice or obligation due dates"),
    "customer_balances":              ("Customer Balances",              "Aged balances per customer"),
    "supplier_invoices":              ("Supplier Invoices",              "All outstanding supplier/vendor invoices"),
    "supplier_balances":              ("Supplier Balances",              "Aged balances per supplier"),
    "fixed_asset_register":           ("Fixed Asset Register",           "Complete register of all fixed assets"),
    "purchase_dates":                 ("Purchase Dates",                 "Acquisition dates for each asset"),
    "depreciation_rates":             ("Depreciation Rates",             "Depreciation rate or method per asset class"),
    "useful_lives":                   ("Useful Lives",                   "Estimated useful life per asset"),
    "inventory_records":              ("Inventory Records",              "Inventory listing with quantities and costs"),
    "quantities":                     ("Quantities on Hand",             "Physical count or system quantity per SKU"),
    "costing_method":                 ("Costing Method",                 "FIFO, Weighted Average, or Specific Identification"),
    "opening_inventory":              ("Opening Inventory",              "Inventory balance at start of period"),
    "purchases":                      ("Purchases During Period",        "All inventory purchases during the period"),
    "sales":                          ("Sales During Period",            "All inventory sold during the period"),
    "adjustments":                    ("Inventory Adjustments",          "Write-offs, shrinkage, returns, write-ups"),
    "share_issues":                   ("Share Issues",                   "New shares issued during the period"),
    "dividends":                      ("Dividends Declared",             "Dividends declared or paid during the period"),
    "owner_contributions":            ("Owner Contributions",            "Capital contributions by owners/shareholders"),
    "retained_earnings_movements":    ("Retained Earnings Movements",    "Adjustments to retained earnings"),
    "loan_agreements":                ("Loan Agreements",                "Signed loan/credit agreements"),
    "repayment_schedules":            ("Repayment Schedules",            "Amortization tables for all debt"),
    "interest_rates":                 ("Interest Rates",                 "Interest rate per loan (fixed or floating)"),
    "lease_contracts":                ("Lease Contracts",                "Signed lease agreements"),
    "payment_schedules":              ("Payment Schedules",              "Lease payment schedule over term"),
    "lease_terms":                    ("Lease Terms",                    "Lease commencement, end date, renewal options"),
    "bank_statements":                ("Bank Statements",                "Official bank statements for the period"),
    "cash_ledger":                    ("Cash Ledger",                    "Internal cash account / GL entries"),
    "external_statements":            ("External Statements",            "Third-party statements (bank, broker, etc.)"),
    "subledgers":                     ("Subledgers",                     "AR, AP, inventory subledger detail"),
    "supporting_documents":           ("Supporting Documents",           "Invoices, contracts, other evidence"),
    "general_ledger":                 ("General Ledger",                 "Full GL trial listing for the period"),
    "lead_schedules":                 ("Lead Schedules",                 "Audit lead schedules per balance area"),
    "supporting_documentation":       ("Supporting Documentation",       "Audit evidence, confirmations, workpapers"),
    "management_disclosures":         ("Management Disclosures",         "MD&A or notes prepared by management"),
    "accounting_policies":            ("Accounting Policies",            "Summary of significant accounting policies"),
    "legal_information":              ("Legal Information",              "Legal proceedings, contingencies, commitments"),
    "related_party_transaction_data": ("Related Party Transaction Data", "All transactions with related parties"),
    "tax_calculations":               ("Tax Calculations",               "Current and deferred tax computations"),
    "tax_returns":                    ("Tax Returns",                    "Filed or draft corporate tax returns"),
    "temporary_differences":          ("Temporary Differences",          "Taxable vs. accounting timing differences"),
    "tax_adjustments":                ("Tax Adjustments",                "Book-to-tax adjustments and reconciliations"),
    "tax_rates":                      ("Tax Rates",                      "Applicable statutory and effective tax rates"),
    "contracts":                      ("Customer Contracts",             "Signed contracts with performance obligations"),
    "invoices":                       ("Invoices",                       "Revenue invoices issued to customers"),
    "performance_obligations":        ("Performance Obligations",        "Identified POBs per contract"),
    "payment_records":                ("Payment Records",                "Evidence of prepayments made"),
    "amortization_periods":           ("Amortization Periods",           "Period over which prepaid expenses are expensed"),
    "outstanding_invoices":           ("Outstanding Invoices",           "Unpaid invoices at period end"),
    "unpaid_obligations":             ("Unpaid Obligations",             "Accrued but unbilled obligations"),
    "payroll_records":                ("Payroll Records",                "Employee payroll detail for the period"),
    "pension_data":                   ("Pension / Retirement Data",      "Defined benefit or contribution plan data"),
    "benefit_data":                   ("Employee Benefit Data",          "Health, bonus, stock compensation detail"),
    "trial_balance_3yr_min":          ("Trial Balances (Min 3 Years)",   "Upload at least 3 years of trial balances"),
}

TB_ONLY_STATEMENTS = {
    "income_statement","balance_sheet","retained_earnings","equity_statement",
    "trial_balance","ratio_analysis","liquidity_report","solvency_report",
    "profitability_report","working_capital",
}
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Quanto — Financial Intelligence</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#050807;--surface:#0b100d;--surface2:#111713;--surface3:#171f1a;
  --text:#EEF3EF;--text-muted:#93A19A;--text-dim:#5c6b63;
  --accent:#00C853;--accent2:#16A34A;--gold:#D4AF37;
  --green:#22C55E;--red:#EF4444;--orange:#f6993f;
  --border:#1D3125;--border2:#2B4A37;--radius:10px;
  --glow:0 0 30px rgba(0,200,83,0.08);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:'Inter',system-ui,sans-serif;min-height:100vh;overflow-x:hidden}
body::before{content:'';position:fixed;inset:0;z-index:-2;
  background:radial-gradient(ellipse 80% 50% at 50% -10%,rgba(0,200,83,0.07),transparent),
             radial-gradient(ellipse 50% 60% at 90% 50%,rgba(22,163,74,0.04),transparent)}
body::after{content:'';position:fixed;inset:0;z-index:-1;
  background-image:linear-gradient(rgba(29,49,37,0.28) 1px,transparent 1px),
                   linear-gradient(90deg,rgba(29,49,37,0.28) 1px,transparent 1px);
  background-size:48px 48px}

header{padding:3rem 2rem 2.5rem;text-align:center;border-bottom:1px solid var(--border);
  background:linear-gradient(180deg,rgba(11,16,13,0.96),rgba(11,16,13,0.72));backdrop-filter:blur(14px);
  position:sticky;top:0;z-index:100;box-shadow:0 1px 40px rgba(0,200,83,0.05)}
.brand{display:flex;align-items:center;justify-content:center;gap:.75rem;margin-bottom:.5rem}
.brand-name{font-family:'DM Serif Display',serif;font-size:2.1rem;letter-spacing:-.04em;
  background:linear-gradient(135deg,#EEF3EF,var(--accent));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.tagline{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;letter-spacing:.14em;text-transform:uppercase}
.ocr-badge{display:inline-flex;align-items:center;gap:.5rem;margin-top:.9rem;padding:.35rem .85rem;
  border:1px solid var(--border2);border-radius:9999px;font-family:'DM Mono',monospace;font-size:.72rem;color:var(--green);
  background:rgba(34,197,94,0.05)}
.disclaimer{font-size:.68rem;color:var(--text-dim);font-family:'DM Mono',monospace;margin-top:.5rem;letter-spacing:.03em}

.tab-bar{display:flex;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:99}
.tab-btn{flex:1;padding:1rem 1.5rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;font-size:.78rem;
  background:transparent;color:var(--text-muted);border:none;cursor:pointer;transition:all .2s;font-family:'Inter',sans-serif}
.tab-btn:hover{background:var(--surface2);color:var(--text)}
.tab-btn.active{background:var(--surface2);color:var(--accent);border-bottom:2px solid var(--accent)}
.tab-panel{display:none}.tab-panel.active{display:block}

main{max-width:1300px;margin:0 auto;padding:2.5rem 1.5rem 6rem;display:grid;gap:1.75rem}

.step-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);
  padding:2rem;position:relative;transition:border-color .2s,box-shadow .2s;box-shadow:var(--glow)}
.step-card:hover{border-color:var(--border2)}
.step-num{position:absolute;top:-11px;left:22px;background:linear-gradient(135deg,var(--accent),var(--accent2));
  color:#04160c;font-family:'DM Mono',monospace;font-weight:700;padding:.28rem .75rem;
  border-radius:9999px;font-size:.75rem;letter-spacing:.05em}
.step-title{font-size:1.2rem;font-weight:600;margin-bottom:.25rem;color:var(--text)}
.step-sub{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;margin-bottom:1.25rem}

.filter-bar{display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:1.25rem}
.filter-btn{padding:.32rem .8rem;border-radius:9999px;font-size:.75rem;font-weight:500;cursor:pointer;transition:all .2s;border:1px solid var(--border2);background:var(--surface3);color:var(--text-muted)}
.filter-btn.active{background:var(--accent);color:#04160c;border-color:var(--accent)}

.stmt-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem}
.stmt-card{border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem 1.1rem;
  background:var(--surface3);cursor:pointer;transition:all .2s;position:relative}
.stmt-card:hover{border-color:var(--accent);background:#1a2420;transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,200,83,0.1)}
.stmt-card.selected{border-color:var(--accent);background:#1a2420;box-shadow:0 0 0 3px rgba(0,200,83,.14)}
.stmt-card.multisource{border-color:rgba(212,175,55,0.22)}
.stmt-card.multisource:hover,.stmt-card.multisource.selected{border-color:var(--gold)}
.stmt-card.multisource.selected{box-shadow:0 0 0 3px rgba(212,175,55,.14)}
.stmt-icon{font-size:1.75rem;margin-bottom:.6rem}
.stmt-name{font-weight:600;font-size:.95rem;margin-bottom:.2rem;color:var(--text)}
.stmt-alias{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;line-height:1.4}
.source-badge{display:inline-block;margin-top:.5rem;padding:.18rem .45rem;background:rgba(212,175,55,0.1);
  border:1px solid rgba(212,175,55,0.3);border-radius:4px;font-size:.65rem;color:var(--gold)}

.source-required{background:var(--surface3);border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem;margin-top:1rem}
.source-required h4{color:var(--accent);font-size:.8rem;margin-bottom:.875rem;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em}
.source-item{display:flex;align-items:flex-start;gap:.875rem;padding:.65rem;border-radius:7px;background:#1a2420;margin-bottom:.4rem}
.source-item-icon{font-size:1.1rem;flex-shrink:0;margin-top:.1rem}
.source-item-text{flex:1}
.source-item-title{font-weight:600;font-size:.85rem;margin-bottom:.15rem}
.source-item-desc{font-size:.75rem;color:var(--text-muted)}
.source-upload-btn{padding:.38rem .8rem;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;
  color:var(--text-muted);font-size:.75rem;cursor:pointer;transition:all .2s;white-space:nowrap}
.source-upload-btn.uploaded{background:rgba(34,197,94,0.1);border-color:var(--green);color:var(--green)}

.upload-area{border:2px dashed var(--border2);border-radius:var(--radius);padding:2.5rem 1.5rem;text-align:center;
  transition:all .2s;background:var(--surface3);cursor:pointer}
.upload-area:hover{border-color:var(--accent);background:#151f19;box-shadow:inset 0 0 0 1px rgba(0,200,83,0.12)}
.file-tag{display:inline-flex;align-items:center;gap:.5rem;padding:.35rem .7rem;background:var(--surface3);
  border:1px solid var(--border2);border-radius:6px;font-size:.8rem;margin:.2rem}

.submit-btn{width:100%;padding:.9rem 2rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;font-size:.95rem;letter-spacing:.04em;text-transform:uppercase;border:none;
  border-radius:var(--radius);cursor:pointer;transition:all .2s;margin-top:.75rem;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.submit-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}
.submit-btn.forecast-btn{background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400;box-shadow:0 4px 20px rgba(212,175,55,0.25)}
.submit-btn.forecast-btn:hover{box-shadow:0 8px 28px rgba(212,175,55,0.35)}
.submit-btn:disabled{opacity:.4;cursor:not-allowed;transform:none;box-shadow:none}

.progress-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:1.75rem}
.result-card{background:var(--surface2);border:1px solid var(--accent);border-radius:var(--radius);padding:2.5rem;text-align:center;
  box-shadow:0 0 40px rgba(0,200,83,0.1)}
.result-card.forecast-result{border-color:var(--gold);box-shadow:0 0 40px rgba(212,175,55,0.1)}
.download-btn{display:inline-flex;align-items:center;gap:.75rem;padding:.9rem 2.1rem;
  background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;border-radius:var(--radius);text-decoration:none;margin-top:1.5rem;transition:all .2s;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.download-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}

.step-item{padding:.65rem 1rem;border-radius:6px;font-family:'DM Mono',monospace;font-size:.82rem;margin-bottom:.2rem}
.step-item.active{color:var(--accent);background:rgba(0,200,83,0.06)}
.step-item.done{color:var(--green)}.step-item.error{color:var(--red)}
.spinner{display:inline-block;width:.85rem;height:.85rem;border:2px solid var(--border2);border-top-color:var(--accent);
  border-radius:50%;animation:spin 0.7s linear infinite;vertical-align:middle;margin-right:.5rem}
@keyframes spin{to{transform:rotate(360deg)}}

.notice-bar{padding:.7rem 1.1rem;border-radius:7px;font-size:.8rem;font-family:'DM Mono',monospace;margin-bottom:.875rem}
.notice-bar.info{background:rgba(0,200,83,0.06);border:1px solid rgba(0,200,83,0.22);color:var(--accent)}
.notice-bar.warn{background:rgba(246,153,63,0.06);border:1px solid rgba(246,153,63,0.2);color:var(--orange)}

.analytics-panel{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;margin-top:0;box-shadow:var(--glow)}
.analytics-header{padding:1.5rem 2rem 1rem;background:linear-gradient(135deg,var(--surface3),#1a2420);border-bottom:1px solid var(--border)}
.analytics-header h3{font-size:1.1rem;font-weight:600;color:var(--text);margin-bottom:.25rem}
.analytics-header p{font-size:.8rem;color:var(--text-muted);font-family:'DM Mono',monospace}

.analytics-tabs{display:flex;border-bottom:1px solid var(--border);background:var(--surface2)}
.analytics-tab{padding:.75rem 1.25rem;font-size:.78rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;
  color:var(--text-muted);cursor:pointer;border:none;background:transparent;transition:all .2s;border-bottom:2px solid transparent}
.analytics-tab:hover{color:var(--text);background:var(--surface3)}
.analytics-tab.active{color:var(--accent);border-bottom-color:var(--accent);background:var(--surface3)}

.analytics-content{padding:1.75rem 2rem}
.analytics-sub-panel{display:none}.analytics-sub-panel.active{display:block}

.kpi-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem;margin-bottom:1.5rem}
.kpi-card{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.1rem;transition:border-color .2s,transform .2s}
.kpi-card:hover{border-color:var(--border2);transform:translateY(-1px)}
.kpi-label{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.35rem}
.kpi-value{font-size:1.4rem;font-weight:700;color:var(--text);line-height:1}
.kpi-value.positive{color:var(--green)}.kpi-value.negative{color:var(--red)}.kpi-value.neutral{color:var(--accent)}
.kpi-note{font-size:.7rem;color:var(--text-dim);margin-top:.3rem;font-family:'DM Mono',monospace}

.insight-block{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.25rem;margin-bottom:1rem}
.insight-block h4{font-size:.82rem;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem;font-family:'DM Mono',monospace}
.insight-block p,.insight-block li{font-size:.88rem;color:var(--text);line-height:1.7}
.insight-block ul{padding-left:1.1rem}
.insight-block li{margin-bottom:.25rem}

.ratio-section-title{font-size:.75rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;
  letter-spacing:.1em;font-family:'DM Mono',monospace;padding:.5rem 0;margin-top:1.25rem;margin-bottom:.5rem;
  border-bottom:1px solid var(--border)}
.ratio-table{width:100%;border-collapse:collapse;margin-bottom:1rem}
.ratio-table th{font-size:.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:.06em;
  padding:.5rem .75rem;text-align:left;border-bottom:1px solid var(--border);font-family:'DM Mono',monospace}
.ratio-table th:last-child{text-align:right}
.ratio-table td{padding:.55rem .75rem;font-size:.83rem;border-bottom:1px solid rgba(29,49,37,0.5)}
.ratio-table tr:last-child td{border-bottom:none}
.ratio-table tr:hover td{background:rgba(0,200,83,0.03)}
.ratio-table td:last-child{text-align:right;font-weight:600;font-family:'DM Mono',monospace}
.ratio-val.good{color:var(--green)}.ratio-val.warn{color:var(--orange)}.ratio-val.bad{color:var(--red)}.ratio-val.neutral{color:var(--accent)}

.chat-container{display:flex;flex-direction:column;height:520px;background:var(--surface3);border-radius:0 0 var(--radius) var(--radius)}
.chat-messages{flex:1;overflow-y:auto;padding:1.5rem;display:flex;flex-direction:column;gap:.875rem;scroll-behavior:smooth}
.chat-messages::-webkit-scrollbar{width:4px}
.chat-messages::-webkit-scrollbar-track{background:transparent}
.chat-messages::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px}
.chat-msg{max-width:85%;border-radius:10px;padding:.875rem 1.1rem;line-height:1.65;font-size:.875rem}
.chat-msg.user{align-self:flex-end;background:linear-gradient(135deg,rgba(0,200,83,0.22),rgba(22,163,74,0.16));
  border:1px solid rgba(0,200,83,0.28);color:var(--text)}
.chat-msg.assistant{align-self:flex-start;background:#1a2420;border:1px solid var(--border);color:var(--text)}
.chat-msg.assistant .msg-label{font-size:.68rem;font-weight:700;color:var(--accent);font-family:'DM Mono',monospace;
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:.4rem}
.chat-msg.typing{padding:.875rem 1.1rem}
.typing-dot{display:inline-block;width:7px;height:7px;background:var(--text-muted);border-radius:50%;margin:0 2px;
  animation:bounce .9s ease-in-out infinite}
.typing-dot:nth-child(2){animation-delay:.15s}.typing-dot:nth-child(3){animation-delay:.3s}
@keyframes bounce{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-5px)}}
.chat-input-row{display:flex;gap:.5rem;padding:1rem 1.25rem;border-top:1px solid var(--border);background:var(--surface2)}
.chat-input{flex:1;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;
  padding:.65rem 1rem;color:var(--text);font-size:.875rem;font-family:'Inter',sans-serif;resize:none;outline:none;
  transition:border-color .2s,box-shadow .2s;min-height:40px;max-height:100px}
.chat-input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,200,83,0.14)}
.chat-send-btn{padding:.65rem 1.1rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  border:none;border-radius:8px;cursor:pointer;font-weight:600;font-size:.82rem;transition:all .2s;white-space:nowrap}
.chat-send-btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,200,83,0.35)}
.chat-send-btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
.chat-empty{text-align:center;padding:2.5rem;color:var(--text-muted);font-family:'DM Mono',monospace;font-size:.82rem}
.chat-empty .chat-empty-icon{font-size:2rem;margin-bottom:.75rem;opacity:.5}

.ledger-upload-row{display:flex;gap:.75rem;align-items:center;padding:1rem 1.25rem;border-bottom:1px solid var(--border);background:var(--surface2);flex-wrap:wrap}
.ledger-state-block{background:#1a2420;border:1px solid var(--border);border-radius:8px;padding:.9rem 1.1rem;margin-bottom:.6rem;font-family:'DM Mono',monospace;font-size:.8rem;white-space:pre-wrap;line-height:1.6}
.ledger-badge{display:inline-block;padding:.15rem .5rem;border-radius:4px;font-size:.68rem;font-family:'DM Mono',monospace;background:rgba(0,200,83,0.12);color:var(--accent);border:1px solid rgba(0,200,83,0.25)}

footer{text-align:center;padding:2rem;color:var(--text-dim);font-family:'DM Mono',monospace;font-size:.72rem;border-top:1px solid var(--border);letter-spacing:.03em}
footer span{color:var(--text-muted)}

input:focus, textarea:focus, select:focus{outline:none}
::selection{background:rgba(0,200,83,0.28);color:#fff}
</style>
</head>
<body>
<header>
  <div class="brand">
    <div class="brand-name">Quanto</div>
  </div>
  <p class="tagline">Financial Intelligence Platform</p>
  <div class="ocr-badge" id="ocrBadge">⬜ Loading OCR status...</div>
  <p class="disclaimer">Quanto is not responsible for financial decisions.</p>
</header>

<div class="tab-bar">
  <button class="tab-btn active" onclick="switchTab('statements',this)">📋 Statement Generator</button>
  <button class="tab-btn" onclick="switchTab('forecasting',this)">🚀 Forecasting Engine</button>
  <button class="tab-btn" onclick="switchTab('ledger',this)">📒 Interactive Ledger</button>
</div>

<div id="tab-statements" class="tab-panel active">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose a Financial Statement</div>
    <div class="notice-bar info">📋 Statements with a purple badge require multiple source documents — required uploads will appear automatically.</div>
    <div class="filter-bar">
      <button class="filter-btn active" id="fAll" onclick="filterCards('all',this)">All</button>
      <button class="filter-btn" id="fTb" onclick="filterCards('tb',this)">Trial Balance Only</button>
      <button class="filter-btn" id="fMulti" onclick="filterCards('multi',this)">Multi-Source</button>
    </div>
    <div class="stmt-grid" id="stmtGrid"></div>
  </section>

  <section class="step-card" id="uploadSection">
    <span class="step-num">02</span>
    <div class="step-title" id="uploadTitle">Upload Required Documents</div>
    <div class="step-sub" id="uploadSub">Select a statement type above to see required documents.</div>
    <div id="sourceRequiredArea"></div>
    <div id="tbUploadArea" style="display:none">
      <div class="upload-area" id="uploadArea">
        <input type="file" id="fileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none"/>
        <div style="font-size:2rem;margin:.75rem 0">📎</div>
        <p style="font-weight:600;font-size:.95rem">Drop Trial Balance here or click to browse</p>
        <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">PDF, JPG, PNG, TIFF supported · AI Vision OCR</p>
      </div>
      <div id="tbFileChosen" style="display:none;margin-top:.6rem;padding:.65rem 1rem;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;align-items:center">
        <span id="tbFileName" style="flex:1;font-size:.85rem"></span>
        <button onclick="removeTbFile()" style="background:none;border:none;color:var(--red);font-size:1.1rem;cursor:pointer;margin-left:1rem">✕</button>
      </div>
    </div>
  </section>

  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn" id="generateBtn" onclick="generate()" disabled>⚡ Generate Statement</button>
  </section>

  <div id="status-area" style="display:none"></div>
  <div id="result-area"></div>

  <div id="analytics-area" style="display:none">
    <div class="analytics-panel">
      <div class="analytics-header">
        <h3 id="analytics-company-title">Financial Analysis</h3>
        <p id="analytics-period-sub">Generated by Quanto Intelligence Engine</p>
      </div>
      <div class="analytics-tabs">
        <button class="analytics-tab active" onclick="switchAnalyticsTab('overview',this)">📊 Overview</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('ratios',this)">🔢 Ratio Analysis</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('insights',this)">💡 Insights</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('chat',this)">💬 Ask Quanto</button>
      </div>
      <div class="analytics-content">
        <div id="panel-overview" class="analytics-sub-panel active">
          <div id="kpi-grid-container"></div>
        </div>
        <div id="panel-ratios" class="analytics-sub-panel">
          <div id="ratios-container"></div>
        </div>
        <div id="panel-insights" class="analytics-sub-panel">
          <div id="insights-container">
            <div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem">
              <span class="spinner"></span> Generating AI insights...
            </div>
          </div>
        </div>
        <div id="panel-chat" class="analytics-sub-panel">
          <div class="chat-container">
            <div class="chat-messages" id="chatMessages">
              <div class="chat-empty">
                <div class="chat-empty-icon">💬</div>
                <p>Ask anything about this financial statement.</p>
                <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is driving the revenue growth?" or "Is this company financially healthy?"</p>
              </div>
            </div>
            <div class="chat-input-row">
              <textarea class="chat-input" id="chatInput" placeholder="Ask Quanto about this financial statement..." rows="1"
                onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendChat()}"></textarea>
              <button class="chat-send-btn" id="chatSendBtn" onclick="sendChat()">Send ↑</button>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</main>
</div>

<div id="tab-forecasting" class="tab-panel">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose Forecast Type</div>
    <div class="stmt-grid" id="fstmtGrid"></div>
  </section>
  <section class="step-card">
    <span class="step-num">02</span>
    <div class="step-title">Upload Historical Trial Balances</div>
    <div class="notice-bar warn">⚠ Minimum 3 fiscal years of trial balances required. Upload one file per year.</div>
    <div id="forecastDropZone" class="upload-area" style="margin-bottom:1rem">
      <input type="file" id="forecastFileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" multiple style="display:none"/>
      <div style="font-size:2rem;margin:.75rem 0">📁</div>
      <p style="font-weight:600;font-size:.95rem">Drop files here or click to browse</p>
      <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">Select multiple files — one per fiscal year</p>
    </div>
    <div id="forecastFileList" style="display:flex;flex-wrap:wrap;gap:.4rem"></div>
  </section>
  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn forecast-btn" id="fgenerateBtn" onclick="generateForecast()" disabled>🚀 Run Forecast Engine (15 Phases)</button>
  </section>
  <div id="fstatus-area" style="display:none"></div>
  <div id="fresult-area"></div>
</main>
</div>

<div id="tab-ledger" class="tab-panel">
<main>
  <section class="step-card">
    <span class="step-num">📒</span>
    <div class="step-title">Interactive Financial Ledger</div>
    <div class="step-sub">Upload a trial balance / ledger document to start a stateful accounting session, then issue adjustments in plain English (e.g. "increase salaries to 50,000", "remove rent expense", "undo").</div>
    <div class="notice-bar info">All figures are computed deterministically in Python from the maintained ledger state — nothing is estimated by the AI.</div>
    <div class="analytics-panel">
      <div class="ledger-upload-row">
        <input type="file" id="ledgerFileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none"/>
        <button class="submit-btn" style="width:auto;margin:0" onclick="document.getElementById('ledgerFileInput').click()">📎 Upload Trial Balance</button>
        <span id="ledgerFileStatus" style="font-size:.8rem;color:var(--text-muted)">No document loaded yet.</span>
        <span id="ledgerSessionBadge" style="display:none" class="ledger-badge">Session Active</span>
      </div>
      <div class="chat-container" style="height:600px">
        <div class="chat-messages" id="ledgerMessages">
          <div class="chat-empty">
            <div class="chat-empty-icon">📒</div>
            <p>Upload a trial balance to initialize the ledger session.</p>
            <p style="margin-top:.35rem;color:var(--text-dim)">Then try: "increase rent to 15,000" · "remove utilities" · "add fuel expense of 2,500" · "undo"</p>
          </div>
        </div>
        <div class="chat-input-row">
          <textarea class="chat-input" id="ledgerInput" placeholder="Issue an accounting command..." rows="1" disabled
            onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendLedgerCommand()}"></textarea>
          <button class="chat-send-btn" id="ledgerSendBtn" onclick="sendLedgerCommand()" disabled>Send ↑</button>
        </div>
      </div>
    </div>
  </section>
</main>
</div>

<footer>
  <span>Quanto Financial Intelligence Platform v1.0</span> &nbsp;·&nbsp; AI Vision OCR &nbsp;·&nbsp; Llama 3.1 Narrative Engine &nbsp;·&nbsp; 100% Local Processing
  <br><span style="color:var(--text-dim)">Quanto is not responsible for financial decisions.</span>
</footer>
<script>
const TB_ONLY = new Set([
  "income_statement","balance_sheet","retained_earnings","equity_statement",
  "trial_balance","ratio_analysis","liquidity_report","solvency_report",
  "profitability_report","working_capital"
]);

let selectedStmt = null;
let selectedFStmt = null;
let tbFile = null;
let multiSourceFiles = {};
let forecastFiles = [];
let allStatements = {};
let currentAnalyticsData = null;
let chatHistory = [];
let ledgerSessionId = null;

const SOURCE_LABELS_JS = """ + json.dumps({k: list(v) for k,v in SOURCE_LABELS.items()}) + r""";

fetch('/api/ocr-status').then(r=>r.json()).then(d=>{
  const b = document.getElementById('ocrBadge');
  const icons = {claude:'🟣', gemini:'🟢', openai:'🔵', local:'🟡'};
  b.textContent = `${icons[d.provider]||'⚪'} OCR: ${d.label} — ${d.note}`;
});

fetch('/api/statements').then(r=>r.json()).then(data=>{
  allStatements = data;
  const sg = document.getElementById('stmtGrid');
  const fg = document.getElementById('fstmtGrid');
  Object.entries(data).forEach(([key, info])=>{
    const isForecast = info.category === 'forecast';
    const isMulti = !TB_ONLY.has(key) && !isForecast;
    const card = document.createElement('label');
    card.className = 'stmt-card' + (isForecast?' forecast-card':'') + (isMulti?' multisource':'');
    card.setAttribute('data-key', key);
    card.setAttribute('data-category', isForecast ? 'forecast' : (TB_ONLY.has(key) ? 'tb' : 'multi'));
    card.innerHTML = `<input type="radio" class="stmt-radio" name="${isForecast?'fstatement':'statement'}" value="${key}" style="display:none"/>
      <div class="stmt-icon">${info.icon}</div>
      <div class="stmt-name">${info.label}</div>
      <div class="stmt-alias">${info.aliases}</div>
      ${isMulti ? '<span class="source-badge">Multi-Source</span>' : ''}`;
    card.addEventListener('click', ()=>{
      if(isForecast){
        document.querySelectorAll('.forecast-card').forEach(c=>c.classList.remove('selected'));
        selectedFStmt = key;
        document.getElementById('fgenerateBtn').disabled = forecastFiles.length < 3;
      } else {
        document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>c.classList.remove('selected'));
        selectedStmt = key;
        renderUploadSection(key, info);
        checkGenerateReady();
      }
      card.classList.add('selected');
    });
    (isForecast ? fg : sg).appendChild(card);
  });
});

function filterCards(type, btn){
  document.querySelectorAll('.filter-btn').forEach(b=>{b.classList.remove('active')});
  btn.classList.add('active');
  document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>{
    const cat = c.getAttribute('data-category');
    c.style.display = (type==='all' || cat===type) ? '' : 'none';
  });
}

function renderUploadSection(key, info){
  const sources = info.sources || [];
  const isTbOnly = TB_ONLY.has(key);
  document.getElementById('uploadTitle').textContent = isTbOnly ? 'Upload Trial Balance' : 'Upload Required Source Documents';
  document.getElementById('uploadSub').textContent = isTbOnly
    ? 'This statement is generated from a single trial balance.'
    : `This statement requires ${sources.length} source document(s). Upload all to proceed.`;
  document.getElementById('tbUploadArea').style.display = isTbOnly ? 'block' : 'none';
  document.getElementById('sourceRequiredArea').style.display = isTbOnly ? 'none' : 'block';
  if(isTbOnly) return;
  multiSourceFiles = {};
  const area = document.getElementById('sourceRequiredArea');
  area.innerHTML = '';
  const box = document.createElement('div');
  box.className = 'source-required';
  box.innerHTML = '<h4>📎 Required Documents</h4>';
  sources.forEach(srcKey=>{
    const [title, desc] = SOURCE_LABELS_JS[srcKey] || [srcKey, ''];
    const item = document.createElement('div');
    item.className = 'source-item';
    item.id = 'src-item-' + srcKey;
    item.innerHTML = `
      <div class="source-item-icon">📄</div>
      <div class="source-item-text">
        <div class="source-item-title">${title}</div>
        <div class="source-item-desc">${desc}</div>
      </div>
      <div>
        <input type="file" id="src-file-${srcKey}" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none" onchange="handleSourceFile('${srcKey}', this)"/>
        <button class="source-upload-btn" id="src-btn-${srcKey}" onclick="document.getElementById('src-file-${srcKey}').click()">Upload ↑</button>
      </div>`;
    box.appendChild(item);
  });
  area.appendChild(box);
}

function handleSourceFile(srcKey, input){
  if(input.files[0]){
    multiSourceFiles[srcKey] = input.files[0];
    const btn = document.getElementById('src-btn-'+srcKey);
    btn.className = 'source-upload-btn uploaded';
    btn.textContent = '✓ ' + input.files[0].name.substring(0,22);
    checkGenerateReady();
  }
}

const fileInput = document.getElementById('fileInput');
const uploadArea = document.getElementById('uploadArea');
uploadArea.addEventListener('click', ()=>fileInput.click());
fileInput.addEventListener('change', e=>{ if(e.target.files[0]) pickTbFile(e.target.files[0]); });
uploadArea.addEventListener('dragover', e=>{e.preventDefault();uploadArea.style.borderColor='var(--accent)'});
uploadArea.addEventListener('dragleave', ()=>{uploadArea.style.borderColor=''});
uploadArea.addEventListener('drop', e=>{e.preventDefault();uploadArea.style.borderColor='';if(e.dataTransfer.files[0])pickTbFile(e.dataTransfer.files[0])});

function pickTbFile(f){
  tbFile = f;
  document.getElementById('tbFileName').textContent = f.name;
  const fc = document.getElementById('tbFileChosen');
  fc.style.display = 'flex';
  checkGenerateReady();
}
function removeTbFile(){
  tbFile = null;
  document.getElementById('tbFileChosen').style.display = 'none';
  fileInput.value = '';
  checkGenerateReady();
}
function checkGenerateReady(){
  if(!selectedStmt){document.getElementById('generateBtn').disabled=true;return;}
  if(TB_ONLY.has(selectedStmt)){
    document.getElementById('generateBtn').disabled = !tbFile;
  } else {
    const info = allStatements[selectedStmt]||{};
    const required = info.sources||[];
    document.getElementById('generateBtn').disabled = !required.every(s=>multiSourceFiles[s]);
  }
}

const forecastDz = document.getElementById('forecastDropZone');
const forecastFi = document.getElementById('forecastFileInput');
forecastDz.addEventListener('click', ()=>forecastFi.click());
forecastFi.addEventListener('change', e=>{
  Array.from(e.target.files).forEach(f=>{if(!forecastFiles.find(x=>x.name===f.name))forecastFiles.push(f)});
  renderForecastFiles();
});
function renderForecastFiles(){
  const list = document.getElementById('forecastFileList');
  list.innerHTML = '';
  forecastFiles.forEach((f,i)=>{
    const tag = document.createElement('div');
    tag.className = 'file-tag';
    tag.innerHTML = `📄 ${f.name} <button onclick="removeForecastFile(${i})" style="background:none;border:none;color:var(--red);cursor:pointer;margin-left:.3rem">✕</button>`;
    list.appendChild(tag);
  });
  document.getElementById('fgenerateBtn').disabled = !(selectedFStmt && forecastFiles.length >= 3);
  if(forecastFiles.length>0 && forecastFiles.length<3){
    const warn = document.createElement('div');
    warn.style.cssText='width:100%;color:var(--orange);font-size:.77rem;margin-top:.4rem;font-family:"DM Mono",monospace';
    warn.textContent = `⚠ ${forecastFiles.length}/3 files — minimum 3 required`;
    list.appendChild(warn);
  }
}
function removeForecastFile(i){forecastFiles.splice(i,1);renderForecastFiles();}

function switchTab(name, btn){
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  btn.classList.add('active');
}

function switchAnalyticsTab(name, btn){
  document.querySelectorAll('.analytics-sub-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.analytics-tab').forEach(b=>b.classList.remove('active'));
  document.getElementById('panel-'+name).classList.add('active');
  btn.classList.add('active');
}

function fmt(v, type){
  if(v===null||v===undefined||v==='N/A')return'N/A';
  const n = parseFloat(v);
  if(isNaN(n)) return String(v);
  if(type==='pct') return (n*100).toFixed(1)+'%';
  if(type==='x') return n.toFixed(2)+'x';
  if(type==='$') return '$'+n.toLocaleString('en-US',{minimumFractionDigits:0,maximumFractionDigits:0});
  if(type==='days') return n.toFixed(1)+' days';
  return n.toFixed(2);
}
function scoreClass(v, good, warn){
  if(v===null||v===undefined||isNaN(parseFloat(v)))return'neutral';
  const n=parseFloat(v);
  if(n>=good)return'good';if(n>=warn)return'warn';return'bad';
}

function renderKPIs(d){
  const g = document.getElementById('kpi-grid-container');
  g.innerHTML = '';
  const grid = document.createElement('div'); grid.className='kpi-grid';
  const kpis = [
    {label:'Revenue',val:fmt(d.revenue,'$'),cls:'neutral',note:'Total revenue for period'},
    {label:'Net Income',val:fmt(d.net_income,'$'),cls:parseFloat(d.net_income)>=0?'positive':'negative',note:'After all expenses'},
    {label:'Gross Profit',val:fmt(d.gross_profit,'$'),cls:parseFloat(d.gross_profit)>=0?'positive':'negative',note:'Revenue minus COGS'},
    {label:'Gross Margin',val:fmt(d.gross_margin,'pct'),cls:scoreClass(parseFloat(d.gross_margin)*100,30,15),note:'Gross Profit / Revenue'},
    {label:'Net Margin',val:fmt(d.net_margin,'pct'),cls:scoreClass(parseFloat(d.net_margin)*100,10,3),note:'Net Income / Revenue'},
    {label:'Total Assets',val:fmt(d.total_assets,'$'),cls:'neutral',note:'Balance sheet total'},
    {label:'Total Liabilities',val:fmt(d.total_liabilities,'$'),cls:'neutral',note:'All obligations'},
    {label:'Total Equity',val:fmt(d.total_equity,'$'),cls:parseFloat(d.total_equity)>=0?'positive':'negative',note:'Net assets'},
  ];
  if(d.ebitda!==null&&d.ebitda!==undefined)
    kpis.splice(3,0,{label:'EBITDA',val:fmt(d.ebitda,'$'),cls:parseFloat(d.ebitda)>=0?'positive':'negative',note:'Earnings before I/T/D/A'});
  kpis.forEach(k=>{
    const card=document.createElement('div');card.className='kpi-card';
    card.innerHTML=`<div class="kpi-label">${k.label}</div><div class="kpi-value ${k.cls}">${k.val}</div><div class="kpi-note">${k.note}</div>`;
    grid.appendChild(card);
  });
  g.appendChild(grid);
}

function renderRatios(d){
  const c = document.getElementById('ratios-container');
  c.innerHTML = '';
  const sections = [
    {title:'Liquidity Ratios', rows:[
      ['Current Ratio', fmt(d.current_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
      ['Quick Ratio', fmt(d.quick_ratio,'x'), 'x', '(Cash + AR) / Current Liabilities', 1.0, 0.8],
      ['Cash Ratio', fmt(d.cash_ratio,'x'), 'x', 'Cash / Current Liabilities', 0.5, 0.2],
      ['Working Capital', fmt(d.working_capital,'$'), '$', 'Current Assets − Current Liabilities', 1, 0],
      ['Working Capital Ratio', fmt(d.working_capital_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
    ]},
    {title:'Profitability Ratios', rows:[
      ['Gross Profit Margin', fmt(d.gross_margin,'pct'), 'pct', 'Gross Profit / Revenue', 0.35, 0.15],
      ['Operating Profit Margin', fmt(d.operating_margin,'pct'), 'pct', 'Operating Income / Revenue', 0.15, 0.05],
      ['Net Profit Margin', fmt(d.net_margin,'pct'), 'pct', 'Net Income / Revenue', 0.10, 0.03],
      ['EBITDA Margin', fmt(d.ebitda_margin,'pct'), 'pct', 'EBITDA / Revenue', 0.20, 0.08],
      ['Return on Assets (ROA)', fmt(d.roa,'pct'), 'pct', 'Net Income / Total Assets', 0.05, 0.02],
      ['Return on Equity (ROE)', fmt(d.roe,'pct'), 'pct', 'Net Income / Total Equity', 0.15, 0.08],
      ['Return on Invested Capital (ROIC)', fmt(d.roic,'pct'), 'pct', 'NOPAT / Invested Capital', 0.12, 0.06],
    ]},
    {title:'Efficiency Ratios', rows:[
      ['Asset Turnover Ratio', fmt(d.asset_turnover,'x'), 'x', 'Revenue / Total Assets', 1.0, 0.5],
      ['Inventory Turnover Ratio', fmt(d.inventory_turnover,'x'), 'x', 'COGS / Average Inventory', 6.0, 3.0],
      ['Accounts Receivable Turnover', fmt(d.ar_turnover,'x'), 'x', 'Revenue / Accounts Receivable', 8.0, 4.0],
      ['Accounts Payable Turnover', fmt(d.ap_turnover,'x'), 'x', 'COGS / Accounts Payable', 8.0, 4.0],
      ['Working Capital Turnover', fmt(d.wc_turnover,'x'), 'x', 'Revenue / Working Capital', 4.0, 2.0],
      ['Fixed Asset Turnover', fmt(d.fixed_asset_turnover,'x'), 'x', 'Revenue / Net Fixed Assets', 3.0, 1.5],
    ]},
    {title:'Leverage / Solvency Ratios', rows:[
      ['Debt-to-Equity Ratio', fmt(d.debt_to_equity,'x'), 'x', 'Total Liabilities / Total Equity', null, null, true],
      ['Debt Ratio', fmt(d.debt_ratio,'x'), 'x', 'Total Liabilities / Total Assets', null, null, true],
      ['Interest Coverage Ratio', fmt(d.interest_coverage,'x'), 'x', 'EBIT / Interest Expense', 3.0, 1.5],
      ['Debt Service Coverage (DSCR)', fmt(d.dscr,'x'), 'x', 'Operating CF / Debt Service', 1.25, 1.0],
      ['Equity Ratio', fmt(d.equity_ratio,'pct'), 'pct', 'Total Equity / Total Assets', 0.50, 0.30],
    ]},
    {title:'Cash Flow Ratios', rows:[
      ['Operating Cash Flow Ratio', fmt(d.ocf_ratio,'x'), 'x', 'Operating CF / Current Liabilities', 1.0, 0.5],
      ['Cash Flow Coverage Ratio', fmt(d.cf_coverage,'x'), 'x', 'Operating CF / Total Liabilities', 0.3, 0.15],
      ['Free Cash Flow Ratio', fmt(d.fcf_ratio,'x'), 'x', 'FCF / Revenue', 0.10, 0.03],
      ['Cash Conversion Ratio', fmt(d.cash_conversion,'x'), 'x', 'Cash from Ops / Net Income', 1.0, 0.7],
    ]},
    {title:'Growth Ratios (Year-over-Year)', rows:[
      ['Revenue Growth Rate', fmt(d.rev_growth,'pct'), 'pct', 'YoY Revenue Change', 0.10, 0.0],
      ['Gross Profit Growth Rate', fmt(d.gp_growth,'pct'), 'pct', 'YoY Gross Profit Change', 0.10, 0.0],
      ['EBITDA Growth Rate', fmt(d.ebitda_growth,'pct'), 'pct', 'YoY EBITDA Change', 0.10, 0.0],
      ['Net Income Growth Rate', fmt(d.ni_growth,'pct'), 'pct', 'YoY Net Income Change', 0.10, 0.0],
      ['Cash Flow Growth Rate', fmt(d.cf_growth,'pct'), 'pct', 'YoY Cash Flow Change', 0.10, 0.0],
    ]},
    {title:'Valuation Ratios (Market-Based — Requires Share Price)', rows:[
      ['Price-to-Earnings (P/E)', d.pe_ratio!==null?fmt(d.pe_ratio,'x'):'— (needs share price)', 'x', 'Share Price / EPS', null, null],
      ['Price-to-Sales (P/S)', d.ps_ratio!==null?fmt(d.ps_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Revenue', null, null],
      ['EV/EBITDA', d.ev_ebitda!==null?fmt(d.ev_ebitda,'x'):'— (needs share price)', 'x', 'Enterprise Value / EBITDA', null, null],
      ['Price-to-Book (P/B)', d.pb_ratio!==null?fmt(d.pb_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Book Value', null, null],
    ]},
    {title:'SaaS / Recurring Revenue Metrics', rows:[
      ['Monthly Recurring Revenue (MRR)', d.mrr!==null?fmt(d.mrr,'$'):'— (not applicable)', '$', 'Estimated from revenue pattern', null, null],
      ['Annual Recurring Revenue (ARR)', d.arr!==null?fmt(d.arr,'$'):'— (not applicable)', '$', 'MRR × 12', null, null],
      ['Customer Acquisition Cost (CAC)', '— (requires customer data)', null, 'Total Sales & Mktg / New Customers', null, null],
      ['Customer Lifetime Value (LTV)', '— (requires customer data)', null, 'Avg Revenue per User × Lifetime', null, null],
      ['LTV/CAC Ratio', '— (requires customer data)', null, 'Target ≥ 3x', null, null],
      ['Churn Rate', '— (requires customer data)', null, 'Customers Lost / Total Customers', null, null],
      ['Net Revenue Retention (NRR)', '— (requires customer data)', null, 'Expansion / Total Revenue', null, null],
    ]},
  ];
  sections.forEach(sec=>{
    const title = document.createElement('div'); title.className='ratio-section-title'; title.textContent=sec.title;
    c.appendChild(title);
    const table = document.createElement('table'); table.className='ratio-table';
    table.innerHTML=`<thead><tr><th>Ratio</th><th>Description</th><th>Value</th></tr></thead><tbody></tbody>`;
    const tbody = table.querySelector('tbody');
    sec.rows.forEach(([name, val, type, desc, good, warn, invertScore])=>{
      const tr=document.createElement('tr');
      let cls='neutral';
      if(type&&val&&val!=='N/A'&&!val.includes('—')&&!val.includes('requires')){
        const raw=parseFloat(val.replace(/[%x$,days\s]/g,''));
        if(!isNaN(raw)&&good!==null){
          if(invertScore){cls=raw<2?'good':raw<4?'warn':'bad';}
          else{cls=raw>=good?'good':raw>=warn?'warn':'bad';}
        }
      }
      tr.innerHTML=`<td>${name}</td><td style="font-size:.78rem;color:var(--text-muted);font-family:'DM Mono',monospace">${desc}</td><td class="ratio-val ${cls}">${val}</td>`;
      tbody.appendChild(tr);
    });
    c.appendChild(table);
  });
}

function renderInsights(insights){
  const c = document.getElementById('insights-container');
  c.innerHTML = '';
  if(typeof insights === 'string'){
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>AI Financial Commentary</h4><p>${insights.replace(/\n/g,'<br>')}</p>`;
    c.appendChild(block);
    return;
  }
  const sections = [
    {key:'financial_analysis', title:'📈 Financial Analysis'},
    {key:'management_insights', title:'💼 Management Insights & Commentary'},
    {key:'risks_and_opportunities', title:'⚠️ Risks & Opportunities'},
  ];
  sections.forEach(({key, title})=>{
    if(!insights[key]) return;
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>${title}</h4>`;
    const content = insights[key];
    if(Array.isArray(content)){
      const ul=document.createElement('ul');
      content.forEach(item=>{const li=document.createElement('li');li.textContent=item;ul.appendChild(li);});
      block.appendChild(ul);
    } else {
      const p=document.createElement('p');p.textContent=content;block.appendChild(p);
    }
    c.appendChild(block);
  });
  if(!c.children.length){
    c.innerHTML='<p style="color:var(--text-muted);font-size:.85rem">No insights available for this statement type.</p>';
  }
}

async function loadAnalytics(data){
  currentAnalyticsData = data;
  chatHistory = [];
  document.getElementById('analytics-company-title').textContent = data.company + ' — Financial Analysis';
  document.getElementById('analytics-period-sub').textContent = data.period + ' · ' + data.statement_label;
  document.getElementById('analytics-area').style.display = 'block';
  document.getElementById('chatMessages').innerHTML = `<div class="chat-empty">
    <div class="chat-empty-icon">💬</div>
    <p>Ask anything about <strong>${data.company}</strong>'s financials.</p>
    <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is the biggest expense?" · "Is liquidity a concern?" · "How is profitability trending?"</p>
  </div>`;
  document.getElementById('insights-container').innerHTML='<div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem"><span class="spinner"></span> Generating AI insights...</div>';
  if(data.analytics){
    renderKPIs(data.analytics);
    renderRatios(data.analytics);
  }
  try{
    const iResp = await fetch('/api/insights', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({analytics: data.analytics, company: data.company, period: data.period, statement_type: data.statement_type})
    });
    const iData = await iResp.json();
    if(iData.insights) renderInsights(iData.insights);
    else document.getElementById('insights-container').innerHTML='<p style="color:var(--text-muted);font-size:.85rem">Insights unavailable.</p>';
  } catch(e){
    document.getElementById('insights-container').innerHTML='<p style="color:var(--red);font-size:.85rem">Could not generate insights — ensure Ollama is running.</p>';
  }
}

async function sendChat(){
  const input = document.getElementById('chatInput');
  const msg = input.value.trim();
  if(!msg || !currentAnalyticsData) return;
  input.value = '';
  const msgs = document.getElementById('chatMessages');
  const firstChild = msgs.querySelector('.chat-empty');
  if(firstChild) msgs.innerHTML='';
  const userBubble = document.createElement('div');
  userBubble.className='chat-msg user'; userBubble.textContent=msg;
  msgs.appendChild(userBubble);
  const typingBubble = document.createElement('div');
  typingBubble.className='chat-msg assistant typing';
  typingBubble.innerHTML='<span class="typing-dot"></span><span class="typing-dot"></span><span class="typing-dot"></span>';
  msgs.appendChild(typingBubble);
  msgs.scrollTop=msgs.scrollHeight;
  document.getElementById('chatSendBtn').disabled=true;
  chatHistory.push({role:'user', content:msg});
  try{
    const resp = await fetch('/api/chat', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({message:msg, history:chatHistory, analytics:currentAnalyticsData.analytics, company:currentAnalyticsData.company, period:currentAnalyticsData.period})
    });
    const data = await resp.json();
    const reply = data.reply || 'I was unable to generate a response. Please try again.';
    chatHistory.push({role:'assistant',content:reply});
    typingBubble.remove();
    const asBubble=document.createElement('div');
    asBubble.className='chat-msg assistant';
    asBubble.innerHTML=`<div class="msg-label">Quanto AI</div>${reply.replace(/\n/g,'<br>')}`;
    msgs.appendChild(asBubble);
    msgs.scrollTop=msgs.scrollHeight;
  }catch(e){
    typingBubble.remove();
    const errBubble=document.createElement('div');
    errBubble.className='chat-msg assistant';
    errBubble.innerHTML='<div class="msg-label">Quanto AI</div>Error connecting to AI engine. Ensure Ollama is running.';
    msgs.appendChild(errBubble);
  }
  document.getElementById('chatSendBtn').disabled=false;
  input.focus();
}

async function generate(){
  if(!selectedStmt) return;
  const btn = document.getElementById('generateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('status-area');
  statusArea.style.display='block';
  statusArea.innerHTML=`<div class="progress-card">
    <div class="step-item active"><span class="spinner"></span>Extracting document text with AI Vision OCR...</div>
    <div class="step-item active"><span class="spinner"></span>Parsing financial accounts & ledger...</div>
    <div class="step-item active"><span class="spinner"></span>Generating professional Excel workbook...</div>
    <div class="step-item active"><span class="spinner"></span>Computing financial analytics...</div>
  </div>`;
  document.getElementById('result-area').innerHTML='';
  document.getElementById('analytics-area').style.display='none';
  const fd = new FormData();
  fd.append('statement_type', selectedStmt);
  if(TB_ONLY.has(selectedStmt)){
    fd.append('files', tbFile, tbFile.name);
  } else {
    const info = allStatements[selectedStmt]||{};
    (info.sources||[]).forEach(src=>{
      if(multiSourceFiles[src]) fd.append('files', multiSourceFiles[src], `${src}::${multiSourceFiles[src].name}`);
    });
  }
  try{
    const resp = await fetch('/api/generate', {method:'POST', body:fd});
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`
      <div class="result-card">
        <div style="font-size:2.2rem">✅</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.statement_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">${data.accounts_found} accounts extracted · OCR: ${data.ocr_provider}</p>
        <a href="/api/download/${data.filename}" class="download-btn">⬇ Download Excel</a>
      </div>`;
    if(data.analytics) await loadAnalytics(data);
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}

async function generateForecast(){
  if(!selectedFStmt||forecastFiles.length<3) return;
  const btn = document.getElementById('fgenerateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('fstatus-area');
  statusArea.style.display='block';
  const phases=['Phase 1-2: Validating and quality-checking documents…','Phase 3: Normalizing account names across years…','Phase 4: Computing historical ratios…','Phase 5-6: Identifying drivers and running forecast methods…','Phase 7: Building three-statement linked model…','Phase 8: Projecting cash flows and runway…','Phase 9: Running scenario analysis (Base/Best/Worst)…','Phase 10: Generating stakeholder analysis…','Phase 11: Running DCF valuation model…','Phase 12: Scoring risk across 6 dimensions…','Phase 13: Generating AI narrative via Llama 3.1…','Phase 14: Computing forecast confidence scores…','Phase 15: Building 13-tab Excel workbook…'];
  statusArea.innerHTML='<div class="progress-card">'+phases.map(p=>`<div class="step-item active"><span class="spinner"></span>${p}</div>`).join('')+'</div>';
  document.getElementById('fresult-area').innerHTML='';
  const fd=new FormData();
  fd.append('forecast_type', selectedFStmt);
  forecastFiles.forEach(f=>fd.append('files', f, f.name));
  try{
    const resp=await fetch('/api/forecast',{method:'POST',body:fd});
    const data=await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    const cs=data.confidence_scores||{};
    document.getElementById('fresult-area').innerHTML=`
      <div class="result-card forecast-result">
        <div style="font-size:2.2rem">🚀</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.forecast_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">
          ${data.years_analyzed} years analyzed · ${data.phases_run} phases complete ·
          Revenue confidence: <strong>${cs.revenue}</strong> · Margins: <strong>${cs.margins}</strong>
        </p>
        ${(data.validation_warnings||[]).map(w=>`<p style="color:var(--orange);font-size:.78rem;margin-top:.35rem">⚠ ${w}</p>`).join('')}
        <a href="/api/download/${data.filename}" class="download-btn" style="background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400">⬇ Download Forecast Workbook</a>
      </div>`;
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('fresult-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}

// ---- Interactive Ledger tab ----
const ledgerFileInput = document.getElementById('ledgerFileInput');
ledgerFileInput.addEventListener('change', e=>{ if(e.target.files[0]) initLedgerSession(e.target.files[0]); });

function ledgerAppendBubble(role, text){
  const msgs = document.getElementById('ledgerMessages');
  const empty = msgs.querySelector('.chat-empty');
  if(empty) msgs.innerHTML = '';
  const bubble = document.createElement('div');
  bubble.className = 'chat-msg ' + role;
  if(role==='assistant'){
    const label = document.createElement('div');
    label.className = 'msg-label'; label.textContent = 'Quanto Ledger';
    const block = document.createElement('div');
    block.className = 'ledger-state-block'; block.textContent = text;
    bubble.appendChild(label); bubble.appendChild(block);
  } else {
    bubble.textContent = text;
  }
  msgs.appendChild(bubble);
  msgs.scrollTop = msgs.scrollHeight;
  return bubble;
}

async function initLedgerSession(file){
  document.getElementById('ledgerFileStatus').textContent = 'Uploading & extracting: ' + file.name + '...';
  const fd = new FormData();
  fd.append('file', file, file.name);
  try{
    const resp = await fetch('/api/ledger/init', {method:'POST', body:fd});
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    ledgerSessionId = data.session_id;
    document.getElementById('ledgerFileStatus').textContent = '✓ Loaded: ' + file.name;
    document.getElementById('ledgerSessionBadge').style.display = 'inline-block';
    document.getElementById('ledgerInput').disabled = false;
    document.getElementById('ledgerSendBtn').disabled = false;
    ledgerAppendBubble('assistant', data.formatted_state);
  }catch(err){
    document.getElementById('ledgerFileStatus').textContent = '❌ ' + err.message;
  }
}

async function sendLedgerCommand(){
  const input = document.getElementById('ledgerInput');
  const msg = input.value.trim();
  if(!msg || !ledgerSessionId) return;
  input.value = '';
  ledgerAppendBubble('user', msg);
  document.getElementById('ledgerSendBtn').disabled = true;
  try{
    const resp = await fetch('/api/ledger/command', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({session_id: ledgerSessionId, command: msg})
    });
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    ledgerAppendBubble('assistant', data.formatted_state);
  }catch(err){
    ledgerAppendBubble('assistant', '⚠ ' + err.message);
  }
  document.getElementById('ledgerSendBtn').disabled = false;
  input.focus();
}
</script>
</body>
</html>"""


# =====================================================================================
# SECTION 1: OCR / DOCUMENT EXTRACTION LAYER
# =====================================================================================

def _pdf_to_images_base64(file_bytes: bytes) -> List[str]:
    """Render PDF pages to base64 PNGs for vision-model OCR. Returns [] if pdf2image unavailable."""
    try:
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(file_bytes, dpi=300)
        result = []
        for img in images[:8]:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(base64.b64encode(buf.getvalue()).decode())
        return result
    except Exception:
        return []


def _image_to_base64(file_bytes: bytes, filename: str) -> str:
    return base64.b64encode(file_bytes).decode()


def _guess_mime(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {
        ".pdf": "application/pdf", ".png": "image/png",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".tif": "image/tiff", ".tiff": "image/tiff", ".bmp": "image/bmp",
    }.get(ext, "application/octet-stream")


OCR_EXTRACTION_PROMPT = """You are a meticulous financial document OCR and extraction engine.
Extract EVERY line item / account you can find in this financial document image(s).

Return ONLY valid JSON (no markdown fences, no commentary) in this exact shape:
{
  "company_name": "string or null",
  "period_label": "string describing the fiscal period, e.g. 'FY2024' or 'Year Ended Dec 31, 2024', or null",
  "currency": "3-letter currency code guess, default USD",
  "accounts": [
    {"account_name": "string", "category": "asset|liability|equity|revenue|cogs|expense|other", "amount": number, "subcategory": "current_asset|non_current_asset|current_liability|non_current_liability|other|null"}
  ],
  "notes": "any caveats about illegible or ambiguous figures, or empty string"
}

Rules:
- Numbers must be plain numbers (no $ signs, no commas, no parentheses — convert (1,000) to -1000).
- Include subtotal/total lines too if present, but prioritize line-level detail.
- If a value is illegible, omit that line and mention it in notes.
- Category must be one of: asset, liability, equity, revenue, cogs, expense, other.
- Do not invent figures that are not visibly present in the document.
"""


def _extract_json_from_text(text: str) -> dict:
    """Best-effort extraction of a JSON object from a model response that may include stray text/fences."""
    text = text.strip()
    text = re.sub(r"^```(?:json)?", "", text).strip()

    text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError("No JSON object found in OCR response")
    return json.loads(match.group(0))


def _ocr_with_claude(file_bytes: bytes, filename: str) -> dict:
    """Uses Anthropic's Claude vision model for OCR/extraction. Claude accepts PDFs natively
    as 'document' content blocks (no pdf2image/poppler conversion needed), which makes this
    provider simpler to set up than the Gemini/OpenAI paths while remaining highly accurate
    on financial documents."""
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    mime = _guess_mime(filename)
    b64 = base64.b64encode(file_bytes).decode()

    if mime == "application/pdf":
        content_block: Dict[str, Any] = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    else:
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": mime, "data": b64},
        }

    message = client.messages.create(
        model="claude-sonnet-5",
        max_tokens=4096,
        temperature=0,
        messages=[{
            "role": "user",
            "content": [content_block, {"type": "text", "text": OCR_EXTRACTION_PROMPT}],
        }],
    )
    text = "".join(block.text for block in message.content if getattr(block, "type", None) == "text")
    return _extract_json_from_text(text)


def _ocr_with_gemini(file_bytes: bytes, filename: str) -> dict:
    from google import genai
    from google.genai import types

    client = genai.Client(api_key=GEMINI_API_KEY)

    mime = _guess_mime(filename)

    response = client.models.generate_content(
        model="gemini-3.5-flash",
        contents=[
            OCR_EXTRACTION_PROMPT,
            types.Part.from_bytes(
                data=file_bytes,
                mime_type=mime,
            ),
        ],
    )

    if not response.text:
        print("Gemini Response:")
        print(response.text)
        raise RuntimeError("Gemini returned an empty response.")


    return _extract_json_from_text(response.text)



def _ocr_with_openai(file_bytes: bytes, filename: str) -> dict:
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    mime = _guess_mime(filename)
    content: List[Dict[str, Any]] = [{"type": "text", "text": OCR_EXTRACTION_PROMPT}]

    if mime == "application/pdf":
        images_b64 = _pdf_to_images_base64(file_bytes)
        if not images_b64:
            raise RuntimeError("PDF page rendering unavailable (install pdf2image + poppler) for OpenAI vision path")
        for img_b64 in images_b64:
            content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}})
    else:
        b64 = _image_to_base64(file_bytes, filename)
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})

    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": content}],
        temperature=0.0,
        max_tokens=4096,
    )
    return _extract_json_from_text(resp.choices[0].message.content)


def _ocr_with_local(file_bytes: bytes, filename: str) -> dict:
    """Local fallback: pdfplumber for PDFs, pytesseract for images, then regex-based line parsing."""
    raw_text = ""
    mime = _guess_mime(filename)
    if mime == "application/pdf":
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                raw_text += (page.extract_text() or "") + "\n"
        if not raw_text.strip():
            images_b64 = _pdf_to_images_base64(file_bytes)
            if images_b64:
                import pytesseract
                from PIL import Image as PILImage
                for img_b64 in images_b64:
                    img = PILImage.open(io.BytesIO(base64.b64decode(img_b64)))
                    raw_text += pytesseract.image_to_string(img) + "\n"
    else:
        import pytesseract
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(file_bytes))
        raw_text = pytesseract.image_to_string(img)

    return _parse_raw_text_to_accounts(raw_text)


LINE_ITEM_RE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z0-9&,\.\'/\-\s]{2,80}?)\s{1,}\$?\(?(?P<amount>-?[\d,]+(?:\.\d{1,2})?)\)?\s*$"
)

ASSET_HINTS = ["cash", "receivable", "inventory", "prepaid", "asset", "equipment", "property", "investment", "goodwill", "intangible"]
LIABILITY_HINTS = ["payable", "liability", "liabilities", "loan", "debt", "accrued", "deferred revenue", "unearned", "note payable", "lease liability"]
EQUITY_HINTS = ["equity", "retained earnings", "common stock", "paid-in capital", "treasury stock", "owner's capital", "shareholders"]
REVENUE_HINTS = ["revenue", "sales", "income from", "service income", "fees earned"]
COGS_HINTS = ["cost of goods", "cogs", "cost of sales", "cost of revenue"]
EXPENSE_HINTS = ["expense", "salaries", "wages", "rent", "utilities", "depreciation", "amortization", "interest expense", "tax expense", "advertising", "insurance", "supplies"]


def _classify_account(name: str) -> str:
    n = name.lower()
    if any(h in n for h in COGS_HINTS): return "cogs"
    if any(h in n for h in REVENUE_HINTS): return "revenue"
    if any(h in n for h in LIABILITY_HINTS): return "liability"
    if any(h in n for h in EQUITY_HINTS): return "equity"
    if any(h in n for h in EXPENSE_HINTS): return "expense"
    if any(h in n for h in ASSET_HINTS): return "asset"
    return "other"


def _parse_raw_text_to_accounts(raw_text: str) -> dict:
    accounts = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line or len(line) < 4:
            continue
        m = LINE_ITEM_RE.match(line)
        if not m:
            continue
        name = re.sub(r"\s{2,}", " ", m.group("name")).strip(" .:-")
        amount_str = m.group("amount").replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        if "(" in line and ")" in line and amount > 0:
            amount = -amount
        if not name or name.lower() in {"total", "subtotal"}:
            continue
        accounts.append({
            "account_name": name,
            "category": _classify_account(name),
            "amount": amount,
            "subcategory": None,
        })
    period_match = re.search(r"(?:year ended|period ended|as of|fiscal year)\s+([A-Za-z0-9,\s]+\d{4})", raw_text, re.IGNORECASE)
    return {
        "company_name": None,
        "period_label": period_match.group(1).strip() if period_match else None,
        "currency": "USD",
        "accounts": accounts,
        "notes": "Extracted via local OCR (pdfplumber/pytesseract) — lower fidelity than AI Vision OCR. Review figures carefully." if accounts else "No line items could be confidently parsed from this document.",
    }


def extract_document(file_bytes: bytes, filename: str) -> dict:
    """
    OCR dispatcher.
    Tries the selected provider first, then falls back to the others.
    """

    available = []

    if ANTHROPIC_API_KEY:
        available.append(("claude", _ocr_with_claude))

    if GEMINI_API_KEY:
        available.append(("gemini", _ocr_with_gemini))

    if OPENAI_API_KEY:
        available.append(("openai", _ocr_with_openai))

    # Put the configured provider first
    provider_chain = sorted(
        available,
        key=lambda x: 0 if x[0] == OCR_PROVIDER else 1
    )

    # Local OCR is always the final fallback
    provider_chain.append(("local", _ocr_with_local))

    last_error = None

    for provider_name, fn in provider_chain:

        print(f"\n===== Trying OCR Provider: {provider_name} =====")

        try:
            result = fn(file_bytes, filename)
            result["_ocr_provider_used"] = provider_name

            print(f"SUCCESS using {provider_name}")

            return result

        except Exception as e:

            print("=" * 60)
            print(f"OCR Provider Failed: {provider_name}")
            print(f"Error Type: {type(e).__name__}")
            print(f"Error: {e}")
            print("=" * 60)

            last_error = e

    raise HTTPException(
        status_code=422,
        detail=f"All OCR providers failed to extract this document. Last error: {last_error}"
    )


def get_ocr_status() -> dict:
    labels = {
        "claude": "Claude Sonnet 5 Vision",
        "gemini": "Gemini 1.5 Pro Vision",
        "openai": "OpenAI GPT-4o Vision",
    }
    keys = {"claude": ANTHROPIC_API_KEY, "gemini": GEMINI_API_KEY, "openai": OPENAI_API_KEY}
    if OCR_PROVIDER in labels and keys[OCR_PROVIDER]:
        return {"provider": OCR_PROVIDER, "label": labels[OCR_PROVIDER], "note": "AI Vision OCR active"}
    for provider, key in keys.items():
        if key:
            return {"provider": provider, "label": labels[provider], "note": "AI Vision OCR active (auto-selected)"}
    return {"provider": "local", "label": "Local OCR (pdfplumber/Tesseract)", "note": "No API key set — using local fallback"}


# =====================================================================================
# SECTION 2: ACCOUNT NORMALIZATION & FINANCIAL CALCULATIONS
# =====================================================================================

def _sum_by_category(accounts: List[Dict[str, Any]], category: str) -> float:
    return sum(a["amount"] for a in accounts if a.get("category") == category)


def _find_account(accounts: List[Dict[str, Any]], *keywords: str) -> float:
    """Sum amounts of accounts whose name contains any of the given (lowercase) keywords."""
    total = 0.0
    for a in accounts:
        name = a.get("account_name", "").lower()
        if any(kw in name for kw in keywords):
            total += a.get("amount", 0.0)
    return total


def compute_core_financials(accounts: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Derive core statement totals (Revenue, COGS, Expenses, Assets, Liabilities, Equity)
    from a flat list of classified accounts. This is the foundation all ratio/KPI math builds on."""
    revenue = _sum_by_category(accounts, "revenue")
    cogs = abs(_sum_by_category(accounts, "cogs"))
    opex = abs(_sum_by_category(accounts, "expense"))
    total_assets = _sum_by_category(accounts, "asset")
    total_liabilities = abs(_sum_by_category(accounts, "liability"))
    total_equity = _sum_by_category(accounts, "equity")

    # If equity wasn't classified/extracted but balance sheet should balance, derive it.
    if total_equity == 0 and total_assets > 0 and total_liabilities > 0:
        total_equity = total_assets - total_liabilities

    gross_profit = revenue - cogs
    depreciation_amort = _find_account(accounts, "depreciation", "amortization")
    interest_expense = abs(_find_account(accounts, "interest expense"))
    tax_expense = abs(_find_account(accounts, "tax expense", "income tax"))

    operating_income = gross_profit - opex
    ebit = operating_income
    ebitda = ebit + abs(depreciation_amort)
    net_income = operating_income - interest_expense - tax_expense

    cash = _find_account(accounts, "cash", "bank")
    accounts_receivable = _find_account(accounts, "accounts receivable", "receivable")
    inventory = _find_account(accounts, "inventory")
    accounts_payable = abs(_find_account(accounts, "accounts payable", "payable"))

    current_assets = cash + accounts_receivable + inventory
    if current_assets == 0:
        current_assets = total_assets * 0.5  # heuristic fallback when subcategory tagging absent
    current_liabilities = accounts_payable
    if current_liabilities == 0:
        current_liabilities = total_liabilities * 0.5

    return {
        "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
        "operating_expenses": opex, "operating_income": operating_income,
        "ebit": ebit, "ebitda": ebitda, "net_income": net_income,
        "interest_expense": interest_expense, "tax_expense": tax_expense,
        "depreciation_amortization": abs(depreciation_amort),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,
        "cash": cash, "accounts_receivable": accounts_receivable, "inventory": inventory,
        "accounts_payable": accounts_payable,
        "current_assets": current_assets, "current_liabilities": current_liabilities,
    }


def _safe_div(numerator: float, denominator: float) -> Optional[float]:
    if denominator == 0 or denominator is None:
        return None
    return numerator / denominator


def compute_full_analytics(accounts: List[Dict[str, Any]], prior_accounts: Optional[List[Dict[str, Any]]] = None,
                            share_price: Optional[float] = None) -> Dict[str, Any]:
    """Computes the full ratio suite shown in the frontend's Overview + Ratio Analysis tabs."""
    f = compute_core_financials(accounts)

    revenue, cogs, gross_profit = f["revenue"], f["cogs"], f["gross_profit"]
    opex, op_income, ebit, ebitda = f["operating_expenses"], f["operating_income"], f["ebit"], f["ebitda"]
    net_income = f["net_income"]
    total_assets, total_liabilities, total_equity = f["total_assets"], f["total_liabilities"], f["total_equity"]
    cash, ar, inv, ap = f["cash"], f["accounts_receivable"], f["inventory"], f["accounts_payable"]
    current_assets, current_liabilities = f["current_assets"], f["current_liabilities"]
    interest_expense = f["interest_expense"]

    quick_assets = cash + ar
    working_capital = current_assets - current_liabilities

    # Approximate operating cash flow (no full CF statement available from a single TB).
    operating_cf = net_income + f["depreciation_amortization"]

    analytics: Dict[str, Any] = {
        "revenue": revenue, "net_income": net_income, "gross_profit": gross_profit,
        "gross_margin": _safe_div(gross_profit, revenue),
        "net_margin": _safe_div(net_income, revenue),
        "operating_margin": _safe_div(op_income, revenue),
        "ebitda": ebitda, "ebitda_margin": _safe_div(ebitda, revenue),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,

        # Liquidity
        "current_ratio": _safe_div(current_assets, current_liabilities),
        "quick_ratio": _safe_div(quick_assets, current_liabilities),
        "cash_ratio": _safe_div(cash, current_liabilities),
        "working_capital": working_capital,
        "working_capital_ratio": _safe_div(current_assets, current_liabilities),

        # Profitability
        "roa": _safe_div(net_income, total_assets),
        "roe": _safe_div(net_income, total_equity),
        "roic": _safe_div(ebit * 0.79, (total_equity + total_liabilities)) if (total_equity + total_liabilities) else None,

        # Efficiency
        "asset_turnover": _safe_div(revenue, total_assets),
        "inventory_turnover": _safe_div(cogs, inv) if inv else None,
        "ar_turnover": _safe_div(revenue, ar) if ar else None,
        "ap_turnover": _safe_div(cogs, ap) if ap else None,
        "wc_turnover": _safe_div(revenue, working_capital) if working_capital else None,
        "fixed_asset_turnover": _safe_div(revenue, total_assets - current_assets) if (total_assets - current_assets) > 0 else None,

        # Leverage / Solvency
        "debt_to_equity": _safe_div(total_liabilities, total_equity),
        "debt_ratio": _safe_div(total_liabilities, total_assets),
        "interest_coverage": _safe_div(ebit, interest_expense) if interest_expense else None,
        "dscr": _safe_div(operating_cf, interest_expense) if interest_expense else None,
        "equity_ratio": _safe_div(total_equity, total_assets),

        # Cash flow (approximated)
        "ocf_ratio": _safe_div(operating_cf, current_liabilities),
        "cf_coverage": _safe_div(operating_cf, total_liabilities),
        "fcf_ratio": _safe_div(operating_cf, revenue),
        "cash_conversion": _safe_div(operating_cf, net_income) if net_income else None,

        # Growth (requires prior period)
        "rev_growth": None, "gp_growth": None, "ebitda_growth": None, "ni_growth": None, "cf_growth": None,

        # Valuation (requires share price / market data — not available from financial statements alone)
        "pe_ratio": None, "ps_ratio": None, "ev_ebitda": None, "pb_ratio": None,

        # SaaS metrics (not derivable from a generic TB without subscription data)
        "mrr": None, "arr": None,
    }

    if prior_accounts:
        pf = compute_core_financials(prior_accounts)
        analytics["rev_growth"] = _safe_div(revenue - pf["revenue"], pf["revenue"]) if pf["revenue"] else None
        analytics["gp_growth"] = _safe_div(gross_profit - pf["gross_profit"], pf["gross_profit"]) if pf["gross_profit"] else None
        prior_ebitda = pf["ebit"] + pf["depreciation_amortization"]
        analytics["ebitda_growth"] = _safe_div(ebitda - prior_ebitda, prior_ebitda) if prior_ebitda else None
        analytics["ni_growth"] = _safe_div(net_income - pf["net_income"], pf["net_income"]) if pf["net_income"] else None
        prior_ocf = pf["net_income"] + pf["depreciation_amortization"]
        analytics["cf_growth"] = _safe_div(operating_cf - prior_ocf, prior_ocf) if prior_ocf else None

    if share_price is not None and net_income and total_equity:
        shares_outstanding = None  # Not derivable without cap table data; left as None deliberately.

    return analytics


def generate_ai_insights(analytics: Dict[str, Any], company: str, period: str, statement_type: str) -> Any:
    """Calls the local Ollama Llama 3.1 model to produce narrative commentary on the analytics.
    Falls back to a deterministic rules-based summary if Ollama is unavailable."""
    try:
        import ollama
        prompt = f"""You are a senior financial analyst writing commentary for {company}'s {period} {STATEMENTS.get(statement_type, {}).get('label', 'financial statement')}.

Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None}, indent=2)}

Quanto is not responsible for financial decisions — keep that spirit in mind (informative, not advisory).

Respond ONLY with valid JSON in this exact shape, no markdown fences:
{{
  "financial_analysis": "2-4 sentences on overall financial position and performance",
  "management_insights": ["3-5 short bullet observations a CFO would care about"],
  "risks_and_opportunities": ["3-5 short bullet items, mixing risks and opportunities"]
}}"""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.3})
        content = resp["message"]["content"]
        return _extract_json_from_text(content)
    except Exception:
        return _fallback_insights(analytics)


def _fallback_insights(a: Dict[str, Any]) -> Dict[str, Any]:
    """Deterministic, rules-based insights used when Ollama is not running."""
    obs = []
    risks = []
    gm = a.get("gross_margin")
    nm = a.get("net_margin")
    cr = a.get("current_ratio")
    de = a.get("debt_to_equity")

    if gm is not None:
        obs.append(f"Gross margin stands at {gm*100:.1f}%, {'a healthy level' if gm > 0.3 else 'on the thinner side for most industries'}.")
    if nm is not None:
        obs.append(f"Net margin of {nm*100:.1f}% reflects {'solid' if nm > 0.1 else 'modest'} bottom-line conversion.")
    if cr is not None:
        if cr < 1.0:
            risks.append(f"Current ratio of {cr:.2f}x is below 1.0 — short-term obligations may exceed liquid assets.")
        else:
            obs.append(f"Current ratio of {cr:.2f}x indicates the company can cover short-term liabilities.")
    if de is not None and de > 2.0:
        risks.append(f"Debt-to-equity of {de:.2f}x signals elevated leverage relative to the equity base.")
    if not obs:
        obs.append("Key margin and liquidity figures were derived from the available data; review the ratio tables for full detail.")
    if not risks:
        risks.append("No major red flags surfaced from the ratios computed; continue monitoring trends period-over-period.")

    return {
        "financial_analysis": " ".join(obs[:2]) if obs else "Financial figures have been computed from the uploaded statement.",
        "management_insights": obs,
        "risks_and_opportunities": risks,
    }


def chat_about_financials(message: str, history: List[Dict[str, str]], analytics: Dict[str, Any],
                           company: str, period: str) -> str:
    """Powers the 'Ask Quanto' chat tab using local Ollama, with a deterministic fallback."""
    try:
        import ollama
        system_msg = f"""You are Quanto AI, a financial analyst assistant embedded in the Quanto platform.
You are discussing {company}'s financials for {period}.
Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None})}
Be concise, factual, and grounded only in the figures provided. Remind the user that Quanto is not responsible
for financial decisions if they ask for advice on what action to take. Do not invent figures not present above."""
        messages = [{"role": "system", "content": system_msg}]
        for h in history[-10:]:
            messages.append({"role": h.get("role", "user"), "content": h.get("content", "")})
        messages.append({"role": "user", "content": message})
        resp = ollama.chat(model=OLLAMA_MODEL, messages=messages, options={"temperature": 0.4})
        return resp["message"]["content"]
    except Exception:
        return ("I can't reach the local Ollama AI engine right now, so I can only point you to the numbers "
                "directly: check the Overview and Ratio Analysis tabs for the figures relevant to your question. "
                "Make sure Ollama is running (`ollama serve`) with the llama3.1 model pulled to enable full chat.")


# =====================================================================================
# SECTION 3: EXCEL WORKBOOK GENERATION (openpyxl, commercial-grade formatting)
# =====================================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XL_NAVY = "1F2D4A"
XL_ACCENT = "4F8EF7"
XL_LIGHT = "EAF1FE"
XL_GREEN = "2E8B57"
XL_RED = "C0392B"
XL_GREY = "7B8DB0"

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=False, color="FFFFFF", italic=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=XL_NAVY)
LABEL_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=XL_NAVY)
NUMBER_FONT = Font(name="Calibri", size=10, color="333333")
DISCLAIMER_FONT = Font(name="Calibri", size=8, italic=True, color=XL_GREY)

HEADER_FILL = PatternFill(start_color=XL_NAVY, end_color=XL_NAVY, fill_type="solid")
SECTION_FILL = PatternFill(start_color=XL_LIGHT, end_color=XL_LIGHT, fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E4F5", end_color="D9E4F5", fill_type="solid")

THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
TOTAL_BORDER = Border(top=Side(style="thin", color=XL_NAVY), bottom=Side(style="double", color=XL_NAVY))

CURRENCY_FMT = '#,##0;[Red](#,##0)'
PCT_FMT = '0.0%'
X_FMT = '0.00"x"'


def _new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_statement_header(ws: Worksheet, company: str, statement_label: str, period: str, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    c = ws.cell(row=1, column=start_col, value=company or "Company Name")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    c = ws.cell(row=2, column=start_col, value=statement_label)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
    c = ws.cell(row=3, column=start_col, value=period or "")
    c.font = SUBHEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 20 if r != 1 else 24

    return 5  # next free row


def _write_disclaimer_footer(ws: Worksheet, row: int, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value="Quanto is not responsible for financial decisions. Generated by Quanto Financial Intelligence Platform.")
    c.font = DISCLAIMER_FONT
    c.alignment = Alignment(horizontal="center")


def _autosize_columns(ws: Worksheet, widths: Dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_section_title(ws: Worksheet, row: int, title: str, span: int = 4, start_col: int = 1):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=title)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_line_item(ws: Worksheet, row: int, label: str, value: Optional[float], indent: int = 1,
                      bold: bool = False, currency: bool = True, label_col: int = 1, value_col: int = 4):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = BOLD_LABEL_FONT if bold else LABEL_FONT
    lc.alignment = Alignment(indent=indent)
    lc.border = THIN_BORDER
    for col in range(label_col + 1, value_col):
        ws.cell(row=row, column=col).border = THIN_BORDER
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT if bold else NUMBER_FONT
    if currency and isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    vc.border = THIN_BORDER
    return row + 1


def _write_total_row(ws: Worksheet, row: int, label: str, value: Optional[float], label_col: int = 1, value_col: int = 4):
    for col in range(label_col, value_col + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = TOTAL_BORDER
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = TOTAL_FONT
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT
    if isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    return row + 2


# ---- Individual statement-tab builders -------------------------------------------------

def _build_income_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Income Statement")
    row = _write_statement_header(ws, company, "Income Statement", period)
    row = _write_section_title(ws, row, "Revenue")
    for a in accounts:
        if a["category"] == "revenue":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Revenue", f["revenue"])

    row = _write_section_title(ws, row, "Cost of Goods Sold")
    for a in accounts:
        if a["category"] == "cogs":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total COGS", f["cogs"])
    row = _write_total_row(ws, row, "Gross Profit", f["gross_profit"])

    row = _write_section_title(ws, row, "Operating Expenses")
    for a in accounts:
        if a["category"] == "expense":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Operating Expenses", f["operating_expenses"])
    row = _write_total_row(ws, row, "Operating Income (EBIT)", f["ebit"])

    row = _write_line_item(ws, row, "Interest Expense", f["interest_expense"], bold=False)
    row = _write_line_item(ws, row, "Income Tax Expense", f["tax_expense"], bold=False)
    row = _write_total_row(ws, row, "NET INCOME", f["net_income"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_balance_sheet_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Balance Sheet")
    row = _write_statement_header(ws, company, "Balance Sheet", period)
    row = _write_section_title(ws, row, "Assets")
    for a in accounts:
        if a["category"] == "asset":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Assets", f["total_assets"])

    row = _write_section_title(ws, row, "Liabilities")
    for a in accounts:
        if a["category"] == "liability":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Liabilities", f["total_liabilities"])

    row = _write_section_title(ws, row, "Equity")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Equity", f["total_equity"])
    row = _write_total_row(ws, row, "Total Liabilities & Equity", f["total_liabilities"] + f["total_equity"])

    balance_check = abs(f["total_assets"] - (f["total_liabilities"] + f["total_equity"]))
    note_row = row
    c = ws.cell(row=note_row, column=1, value=("✓ Balance sheet balances." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f} — review extracted figures."))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, note_row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_retained_earnings_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Retained Earnings")
    row = _write_statement_header(ws, company, "Statement of Retained Earnings", period)
    opening_re = _find_account(accounts, "retained earnings") - f["net_income"]
    dividends = abs(_find_account(accounts, "dividend"))
    row = _write_line_item(ws, row, "Retained Earnings — Beginning of Period", opening_re)
    row = _write_line_item(ws, row, "Add: Net Income", f["net_income"])
    row = _write_line_item(ws, row, "Less: Dividends Declared", -dividends if dividends else 0)
    row = _write_total_row(ws, row, "Retained Earnings — End of Period", opening_re + f["net_income"] - dividends)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


def _build_equity_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Equity Statement")
    row = _write_statement_header(ws, company, "Statement of Shareholders' Equity", period)
    row = _write_section_title(ws, row, "Equity Components")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Shareholders' Equity", f["total_equity"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_trial_balance_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Trial Balance")
    row = _write_statement_header(ws, company, "Trial Balance", period, span=5)
    headers = ["Account", "Category", "Debit", "Credit"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    row += 1
    total_debit = total_credit = 0.0
    debit_categories = {"asset", "cogs", "expense"}
    for a in accounts:
        is_debit = a["category"] in debit_categories
        amt = abs(a["amount"])
        ws.cell(row=row, column=1, value=a["account_name"]).font = LABEL_FONT
        ws.cell(row=row, column=2, value=a["category"].title()).font = LABEL_FONT
        debit_cell = ws.cell(row=row, column=3, value=amt if is_debit else None)
        credit_cell = ws.cell(row=row, column=4, value=None if is_debit else amt)
        debit_cell.number_format = CURRENCY_FMT
        credit_cell.number_format = CURRENCY_FMT
        if is_debit:
            total_debit += amt
        else:
            total_credit += amt
        row += 1
    row = _write_total_row(ws, row, "TOTAL", None, value_col=4)
    ws.cell(row=row - 2, column=3, value=total_debit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=3).font = TOTAL_FONT
    ws.cell(row=row - 2, column=4, value=total_credit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=4).font = TOTAL_FONT
    balance_check = abs(total_debit - total_credit)
    c = ws.cell(row=row, column=1, value=("✓ Trial balance is in balance." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f}"))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, row + 1, span=4)
    _autosize_columns(ws, {1: 38, 2: 16, 3: 16, 4: 16})
    return ws


def _build_ratio_tab(wb: Workbook, analytics: Dict[str, Any], company, period, sheet_name="Ratio Analysis"):
    ws = wb.create_sheet(sheet_name)
    row = _write_statement_header(ws, company, sheet_name, period)
    sections = [
        ("Liquidity Ratios", [
            ("Current Ratio", analytics.get("current_ratio"), X_FMT),
            ("Quick Ratio", analytics.get("quick_ratio"), X_FMT),
            ("Cash Ratio", analytics.get("cash_ratio"), X_FMT),
            ("Working Capital", analytics.get("working_capital"), CURRENCY_FMT),
        ]),
        ("Profitability Ratios", [
            ("Gross Margin", analytics.get("gross_margin"), PCT_FMT),
            ("Operating Margin", analytics.get("operating_margin"), PCT_FMT),
            ("Net Margin", analytics.get("net_margin"), PCT_FMT),
            ("EBITDA Margin", analytics.get("ebitda_margin"), PCT_FMT),
            ("Return on Assets (ROA)", analytics.get("roa"), PCT_FMT),
            ("Return on Equity (ROE)", analytics.get("roe"), PCT_FMT),
        ]),
        ("Efficiency Ratios", [
            ("Asset Turnover", analytics.get("asset_turnover"), X_FMT),
            ("Inventory Turnover", analytics.get("inventory_turnover"), X_FMT),
            ("AR Turnover", analytics.get("ar_turnover"), X_FMT),
            ("AP Turnover", analytics.get("ap_turnover"), X_FMT),
        ]),
        ("Leverage / Solvency Ratios", [
            ("Debt-to-Equity", analytics.get("debt_to_equity"), X_FMT),
            ("Debt Ratio", analytics.get("debt_ratio"), X_FMT),
            ("Interest Coverage", analytics.get("interest_coverage"), X_FMT),
            ("Equity Ratio", analytics.get("equity_ratio"), PCT_FMT),
        ]),
    ]
    for title, rows_data in sections:
        row = _write_section_title(ws, row, title)
        for label, value, fmt in rows_data:
            lc = ws.cell(row=row, column=1, value=label); lc.font = LABEL_FONT; lc.alignment = Alignment(indent=1)
            vc = ws.cell(row=row, column=4, value=value if value is not None else "N/A")
            vc.font = NUMBER_FONT
            if isinstance(value, (int, float)):
                vc.number_format = fmt
            vc.alignment = Alignment(horizontal="right")
            row += 1
        row += 1
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 32, 2: 10, 3: 10, 4: 16})
    return ws


def _build_generic_schedule_tab(wb: Workbook, statement_key: str, statement_label: str,
                                 extracted_docs: Dict[str, dict], company: str, period: str):
    """Generic builder for the 25+ multi-source schedules (AR aging, fixed asset schedule, debt schedule, etc.)
    Lists all extracted line items per source document in clearly labeled sections, since each schedule type
    has bespoke business logic that depends on real-world source documents Quanto can refine further per type."""
    ws = wb.create_sheet(statement_label[:31])
    row = _write_statement_header(ws, company, statement_label, period)
    info = STATEMENTS.get(statement_key, {})
    for source_key in info.get("sources", []):
        doc = extracted_docs.get(source_key)
        src_title = SOURCE_LABELS.get(source_key, (source_key, ""))[0]
        row = _write_section_title(ws, row, f"Source: {src_title}")
        if not doc or not doc.get("accounts"):
            row = _write_line_item(ws, row, "(No line items extracted from this document)", None, currency=False)
            continue
        section_total = 0.0
        for a in doc["accounts"]:
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
            section_total += a["amount"]
        row = _write_total_row(ws, row, f"Subtotal — {src_title}", section_total)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


# =====================================================================================
# SECTION 4: STATEMENT GENERATION ORCHESTRATION
# =====================================================================================

def build_statement_workbook(statement_type: str, extracted_docs: Dict[str, dict],
                              company: str, period: str) -> Tuple[Workbook, Dict[str, Any], int]:
    """Builds the full output workbook for any of the 40 statement types and returns
    (workbook, analytics_dict_or_None, total_accounts_found)."""
    wb = _new_workbook()
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    is_tb_only = statement_type in TB_ONLY_STATEMENTS
    analytics = None
    total_accounts = 0

    if is_tb_only:
        tb_doc = extracted_docs.get("trial_balance")
        if not tb_doc:
            raise HTTPException(status_code=400, detail="Trial balance document is required but was not found.")
        accounts = tb_doc.get("accounts", [])
        total_accounts = len(accounts)
        f = compute_core_financials(accounts)
        analytics = compute_full_analytics(accounts)

        if statement_type == "income_statement":
            _build_income_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "balance_sheet":
            _build_balance_sheet_tab(wb, accounts, company, period, f)
        elif statement_type == "retained_earnings":
            _build_retained_earnings_tab(wb, accounts, company, period, f)
        elif statement_type == "equity_statement":
            _build_equity_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "trial_balance":
            _build_trial_balance_tab(wb, accounts, company, period, f)
        elif statement_type in ("ratio_analysis", "liquidity_report", "solvency_report",
                                 "profitability_report", "working_capital"):
            label_map = {
                "ratio_analysis": "Financial Ratio Analysis",
                "liquidity_report": "Liquidity Report",
                "solvency_report": "Solvency Report",
                "profitability_report": "Profitability Report",
                "working_capital": "Working Capital Report",
            }
            _build_ratio_tab(wb, analytics, company, period, sheet_name=label_map[statement_type])
        # Always include the source trial balance for traceability
        _build_trial_balance_tab(wb, accounts, company, period, f) if statement_type != "trial_balance" else None

    else:
        # Multi-source schedules: 25+ specialized statements built generically from their
        # required source documents, organized into clearly labeled sections per source.
        for doc in extracted_docs.values():
            total_accounts += len(doc.get("accounts", []))
        _build_generic_schedule_tab(wb, statement_type, info["label"], extracted_docs, company, period)

        # If an income statement and/or balance sheet were among the sources, compute analytics too
        combined_accounts = []
        for doc in extracted_docs.values():
            combined_accounts.extend(doc.get("accounts", []))
        if combined_accounts:
            analytics = compute_full_analytics(combined_accounts)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Statement")
        _write_statement_header(ws, company, info["label"], period)

    return wb, analytics, total_accounts


def determine_company_and_period(extracted_docs: Dict[str, dict]) -> Tuple[str, str]:
    company = None
    period = None
    for doc in extracted_docs.values():
        if not company and doc.get("company_name"):
            company = doc["company_name"]
        if not period and doc.get("period_label"):
            period = doc["period_label"]
    return company or "Unnamed Company", period or datetime.now().strftime("FY%Y")


def save_workbook_and_get_filename(wb: Workbook, prefix: str) -> str:
    filename = f"{prefix}_{uuid.uuid4().hex[:10]}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return filename


# =====================================================================================
# SECTION 5: FORECASTING ENGINE — PHASES 1-8 (Historical Analysis & Projection)
# =====================================================================================

FORECAST_YEARS_OUT = 5


def fc_phase1_2_validate(yearly_docs: List[dict]) -> Tuple[List[dict], List[str]]:
    """Phase 1-2: Validate and quality-check uploaded trial balances. Returns
    (sorted list of {year, accounts, label}, list of warning strings)."""
    warnings: List[str] = []
    parsed = []
    for i, doc in enumerate(yearly_docs):
        accounts = doc.get("accounts", [])
        if not accounts:
            warnings.append(f"Document {i+1} ({doc.get('period_label') or 'unknown period'}) yielded no extractable line items.")
        label = doc.get("period_label") or f"Year {i+1}"
        year_match = re.search(r"(20\d{2}|19\d{2})", label)
        year = int(year_match.group(1)) if year_match else (2020 + i)
        parsed.append({"year": year, "label": label, "accounts": accounts})

    parsed.sort(key=lambda x: x["year"])
    if len(parsed) < 3:
        warnings.append(f"Only {len(parsed)} fiscal years provided — minimum 3 recommended for reliable trend forecasting.")

    years_seen = [p["year"] for p in parsed]
    if len(set(years_seen)) != len(years_seen):
        warnings.append("Duplicate or indeterminate fiscal years detected — year labels were inferred from document content where possible.")

    return parsed, warnings


def fc_phase3_normalize(parsed_years: List[dict]) -> List[Dict[str, Any]]:
    """Phase 3: Normalize account names/categories across years into a consistent
    per-year core-financials series for trend computation."""
    series = []
    for p in parsed_years:
        f = compute_core_financials(p["accounts"])
        f["year"] = p["year"]
        f["label"] = p["label"]
        series.append(f)
    return series


def fc_phase4_historical_ratios(series: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Phase 4: Compute historical ratios for each year in the series for trend analysis."""
    ratio_history = []
    for yr in series:
        gm = _safe_div(yr["gross_profit"], yr["revenue"])
        nm = _safe_div(yr["net_income"], yr["revenue"])
        cr = _safe_div(yr["current_assets"], yr["current_liabilities"])
        de = _safe_div(yr["total_liabilities"], yr["total_equity"])
        ratio_history.append({"year": yr["year"], "gross_margin": gm, "net_margin": nm,
                               "current_ratio": cr, "debt_to_equity": de})
    return ratio_history


def _cagr(start: float, end: float, periods: int) -> Optional[float]:
    if start is None or end is None or start <= 0 or periods <= 0:
        return None
    try:
        return (end / start) ** (1 / periods) - 1
    except (ValueError, ZeroDivisionError):
        return None


def fc_phase5_6_drivers_and_growth(series: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 5-6: Identify revenue/expense drivers and compute growth forecasts via
    CAGR, weighted-average growth, and simple linear trend — the three methods referenced
    in the 'Growth Rate Forecast' statement type."""
    revenues = [s["revenue"] for s in series]
    n_periods = len(series) - 1

    cagr = _cagr(revenues[0], revenues[-1], n_periods) if n_periods > 0 else None

    yoy_growth_rates = []
    for i in range(1, len(revenues)):
        g = _safe_div(revenues[i] - revenues[i-1], revenues[i-1])
        if g is not None:
            yoy_growth_rates.append(g)
    weighted_growth = None
    if yoy_growth_rates:
        weights = list(range(1, len(yoy_growth_rates) + 1))
        weighted_growth = sum(g * w for g, w in zip(yoy_growth_rates, weights)) / sum(weights)

    # Simple linear trend (least squares slope) expressed as an implied growth rate off the latest year
    trend_growth = None
    if len(revenues) >= 2:
        x = list(range(len(revenues)))
        x_mean = statistics.mean(x)
        y_mean = statistics.mean(revenues)
        denom = sum((xi - x_mean) ** 2 for xi in x)
        if denom > 0:
            slope = sum((xi - x_mean) * (yi - y_mean) for xi, yi in zip(x, revenues)) / denom
            trend_growth = _safe_div(slope, revenues[-1])

    expense_ratios = [_safe_div(s["operating_expenses"], s["revenue"]) for s in series]
    expense_ratios = [e for e in expense_ratios if e is not None]
    avg_expense_ratio = statistics.mean(expense_ratios) if expense_ratios else 0.3

    cogs_ratios = [_safe_div(s["cogs"], s["revenue"]) for s in series]
    cogs_ratios = [c for c in cogs_ratios if c is not None]
    avg_cogs_ratio = statistics.mean(cogs_ratios) if cogs_ratios else 0.4

    chosen_growth = weighted_growth if weighted_growth is not None else (cagr if cagr is not None else 0.05)

    return {
        "cagr": cagr, "weighted_growth": weighted_growth, "trend_growth": trend_growth,
        "chosen_growth_rate": chosen_growth,
        "avg_expense_ratio": avg_expense_ratio, "avg_cogs_ratio": avg_cogs_ratio,
        "yoy_growth_rates": yoy_growth_rates,
    }


def fc_phase7_three_statement_model(series: List[Dict[str, Any]], drivers: Dict[str, Any],
                                     years_out: int = FORECAST_YEARS_OUT) -> List[Dict[str, Any]]:
    """Phase 7: Build a linked 3-statement (simplified) forecast for N years forward,
    driven by the chosen revenue growth rate and historical expense/COGS ratios."""
    last = series[-1]
    growth = drivers["chosen_growth_rate"]
    cogs_ratio = drivers["avg_cogs_ratio"]
    expense_ratio = drivers["avg_expense_ratio"]

    projections = []
    prev_revenue = last["revenue"]
    prev_assets = last["total_assets"]
    prev_liabilities = last["total_liabilities"]
    prev_equity = last["total_equity"]
    prev_cash = last["cash"]

    for i in range(1, years_out + 1):
        revenue = prev_revenue * (1 + growth)
        cogs = revenue * cogs_ratio
        gross_profit = revenue - cogs
        opex = revenue * expense_ratio
        operating_income = gross_profit - opex
        tax_rate = _safe_div(last["tax_expense"], (operating_income if operating_income else 1)) or 0.21
        tax_rate = min(max(tax_rate, 0.0), 0.40)
        interest_expense = last["interest_expense"]
        net_income = (operating_income - interest_expense) * (1 - tax_rate)

        assets = prev_assets * (1 + growth * 0.6)
        liabilities = prev_liabilities * (1 + growth * 0.5)
        equity = prev_equity + net_income
        cash = prev_cash + net_income * 0.7

        projections.append({
            "year_offset": i, "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
            "operating_expenses": opex, "operating_income": operating_income,
            "interest_expense": interest_expense, "net_income": net_income,
            "total_assets": assets, "total_liabilities": liabilities, "total_equity": equity,
            "cash": cash,
        })
        prev_revenue, prev_assets, prev_liabilities, prev_equity, prev_cash = revenue, assets, liabilities, equity, cash

    return projections


def fc_phase8_cashflow_forecast(projections: List[Dict[str, Any]], current_cash: float) -> List[Dict[str, Any]]:
    """Phase 8: Project operating/investing/financing cash flows and resulting cash runway."""
    cf_rows = []
    running_cash = current_cash
    for p in projections:
        operating_cf = p["net_income"] * 0.85
        investing_cf = -p["revenue"] * 0.03
        financing_cf = 0.0
        net_change = operating_cf + investing_cf + financing_cf
        running_cash += net_change
        monthly_burn = abs(net_change) / 12 if net_change < 0 else 0
        runway_months = (running_cash / monthly_burn) if monthly_burn > 0 else None
        cf_rows.append({
            "year_offset": p["year_offset"], "operating_cf": operating_cf, "investing_cf": investing_cf,
            "financing_cf": financing_cf, "net_change_in_cash": net_change, "ending_cash": running_cash,
            "runway_months": runway_months,
        })
    return cf_rows


# =====================================================================================
# SECTION 6: FORECASTING ENGINE — PHASES 9-15 (Scenarios, Valuation, Risk, Narrative)
# =====================================================================================

def fc_phase9_scenarios(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """Phase 9: Run Base / Best / Worst case scenario analysis by flexing the growth rate
    and expense ratio assumptions, each tagged with an illustrative probability weight."""
    base_growth = drivers["chosen_growth_rate"]
    scenarios = {}
    scenario_defs = [
        ("base", base_growth, drivers["avg_expense_ratio"], 0.50),
        ("best", base_growth + 0.08, drivers["avg_expense_ratio"] * 0.92, 0.25),
        ("worst", max(base_growth - 0.12, -0.10), drivers["avg_expense_ratio"] * 1.10, 0.25),
    ]
    for name, growth, expense_ratio, probability in scenario_defs:
        flexed_drivers = dict(drivers)
        flexed_drivers["chosen_growth_rate"] = growth
        flexed_drivers["avg_expense_ratio"] = expense_ratio
        projections = fc_phase7_three_statement_model(series, flexed_drivers, years_out=FORECAST_YEARS_OUT)
        scenarios[name] = {"projections": projections, "growth_assumed": growth,
                            "expense_ratio_assumed": expense_ratio, "probability": probability}
    return scenarios


def fc_phase10_stakeholder_analysis(series: List[Dict[str, Any]], projections: List[Dict[str, Any]],
                                     analytics_latest: Dict[str, Any]) -> Dict[str, str]:
    """Phase 10: Generate stakeholder-specific summaries (Owner, CFO, Investor, Bank, Auditor)."""
    last = series[-1]
    y1, y5 = projections[0], projections[-1]
    rev_cagr_fwd = _cagr(last["revenue"], y5["revenue"], len(projections))

    owner = (f"Revenue is projected to grow from {last['revenue']:,.0f} to {y5['revenue']:,.0f} "
             f"over {len(projections)} years (~{(rev_cagr_fwd or 0)*100:.1f}% CAGR), with net income reaching "
             f"{y5['net_income']:,.0f} by year {len(projections)}.")
    cfo = (f"Year 1 net income of {y1['net_income']:,.0f} assumes a {(analytics_latest.get('gross_margin') or 0)*100:.1f}% "
           f"gross margin held roughly flat; working capital and capex assumptions are simplified and should be "
           f"refined with a full budget once available.")
    investor = (f"Implied forward growth of ~{(rev_cagr_fwd or 0)*100:.1f}% CAGR with equity growing from "
                f"{last['total_equity']:,.0f} to {y5['total_equity']:,.0f}; see the Valuation tab for DCF-based "
                f"enterprise and equity value estimates.")
    bank = (f"Debt service capacity depends on interest coverage; current leverage shows debt-to-equity of "
            f"{analytics_latest.get('debt_to_equity') or 0:.2f}x. Cash flow forecast tab details projected runway "
            f"and coverage ratios.")
    auditor = ("This forecast is a model-based projection derived from historical trial balances using simplified "
               "linear/CAGR-based assumptions; it does not constitute audited financial statements and carries "
               "material estimation uncertainty around working capital, capex, and financing assumptions.")
    return {"owner": owner, "cfo": cfo, "investor": investor, "bank": bank, "auditor": auditor}


def fc_phase11_dcf_valuation(projections: List[Dict[str, Any]], discount_rate: float = 0.12,
                              terminal_growth: float = 0.025) -> Dict[str, Any]:
    """Phase 11: DCF valuation — discounts a simplified unlevered FCF proxy (NI + back of envelope
    add-backs) and computes terminal value via the Gordon Growth method."""
    fcf_series = [p["net_income"] * 0.85 - p["revenue"] * 0.03 for p in projections]
    pv_sum = 0.0
    pv_detail = []
    for i, fcf in enumerate(fcf_series, start=1):
        pv = fcf / ((1 + discount_rate) ** i)
        pv_sum += pv
        pv_detail.append({"year_offset": i, "fcf": fcf, "present_value": pv})

    terminal_fcf = fcf_series[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (discount_rate - terminal_growth) if discount_rate > terminal_growth else None
    pv_terminal = terminal_value / ((1 + discount_rate) ** len(fcf_series)) if terminal_value else None

    enterprise_value = pv_sum + (pv_terminal or 0)
    net_debt = projections[0]["total_liabilities"] - projections[0]["cash"]
    equity_value = enterprise_value - net_debt

    return {
        "discount_rate": discount_rate, "terminal_growth": terminal_growth,
        "pv_detail": pv_detail, "pv_of_explicit_fcf": pv_sum,
        "terminal_value": terminal_value, "pv_of_terminal_value": pv_terminal,
        "enterprise_value": enterprise_value, "net_debt_estimate": net_debt,
        "equity_value": equity_value,
    }


def fc_phase12_risk_scoring(series: List[Dict[str, Any]], analytics_latest: Dict[str, Any],
                             cashflow_forecast: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 12: Score risk across 6 dimensions (0-100, higher = riskier) and produce an overall score."""
    def clamp(v, lo=0, hi=100): return max(lo, min(hi, v))

    revenues = [s["revenue"] for s in series]
    rev_volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if len(revenues) > 1 and statistics.mean(revenues) else 0
    concentration_risk = clamp(rev_volatility * 100)

    cr = analytics_latest.get("current_ratio") or 1.0
    liquidity_risk = clamp((1.5 - cr) * 50) if cr < 1.5 else clamp((1.5 - cr) * 20)

    de = analytics_latest.get("debt_to_equity") or 0.5
    debt_risk = clamp(de * 30)

    negative_cf_years = sum(1 for cf in cashflow_forecast if cf["net_change_in_cash"] < 0)
    burn_risk = clamp((negative_cf_years / max(len(cashflow_forecast), 1)) * 100)

    nm = analytics_latest.get("net_margin")
    profitability_risk = clamp((0.1 - nm) * 300) if nm is not None and nm < 0.1 else 0

    n_years = len(series)
    data_quality_risk = clamp((3 - n_years) * 20) if n_years < 3 else 5

    dimensions = {
        "revenue_concentration_volatility": round(concentration_risk, 1),
        "liquidity_risk": round(liquidity_risk, 1),
        "debt_leverage_risk": round(debt_risk, 1),
        "cash_burn_risk": round(burn_risk, 1),
        "profitability_risk": round(profitability_risk, 1),
        "data_quality_risk": round(data_quality_risk, 1),
    }
    overall = round(statistics.mean(dimensions.values()), 1)
    if overall < 25: band = "Low Risk"
    elif overall < 50: band = "Moderate Risk"
    elif overall < 75: band = "Elevated Risk"
    else: band = "High Risk"

    return {"dimensions": dimensions, "overall_score": overall, "risk_band": band}


def fc_phase13_narrative(series, projections, drivers, risk: Dict[str, Any], company: str) -> str:
    """Phase 13: AI narrative commentary via local Ollama (Llama 3.1), with deterministic fallback."""
    try:
        import ollama
        prompt = f"""You are a financial forecasting analyst. Write a concise narrative (4-6 sentences) summarizing
this forecast for {company}.

Historical revenue: {[round(s['revenue']) for s in series]}
Forecast growth rate assumed: {drivers['chosen_growth_rate']*100:.1f}%
Year 1 forecast revenue: {projections[0]['revenue']:,.0f}, Year {len(projections)} forecast revenue: {projections[-1]['revenue']:,.0f}
Overall risk score: {risk['overall_score']}/100 ({risk['risk_band']})

Quanto is not responsible for financial decisions — write as an informative analyst, not as advice to act on.
Respond with plain text only, no markdown, no JSON."""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.4})
        return resp["message"]["content"].strip()
    except Exception:
        return (f"{company}'s historical revenue trend implies a forward growth assumption of "
                f"{drivers['chosen_growth_rate']*100:.1f}% annually, projecting revenue from "
                f"{projections[0]['revenue']:,.0f} in year 1 to {projections[-1]['revenue']:,.0f} by year "
                f"{len(projections)}. The model's overall risk score of {risk['overall_score']}/100 places this "
                f"forecast in the '{risk['risk_band']}' band, driven primarily by the dimensions with the highest "
                f"individual scores. As with any model-based projection, actual results will depend on factors "
                f"such as market conditions, execution, and financing decisions not captured in the historical "
                f"trial balances alone. (Narrative generated via fallback — start Ollama with llama3.1 for richer AI commentary.)")


def fc_phase14_confidence_scores(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, str]:
    """Phase 14: Qualitative confidence scoring for revenue and margin assumptions based on
    data sufficiency and historical volatility."""
    n_years = len(series)
    revenues = [s["revenue"] for s in series]
    volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if n_years > 1 and statistics.mean(revenues) else 1

    def score(n_years_local, volatility_local):
        if n_years_local >= 5 and volatility_local < 0.15: return "High"
        if n_years_local >= 3 and volatility_local < 0.30: return "Medium"
        return "Low"

    revenue_confidence = score(n_years, volatility)
    margin_volatility = statistics.pstdev([_safe_div(s["gross_profit"], s["revenue"]) or 0 for s in series]) if n_years > 1 else 0.2
    margins_confidence = score(n_years, margin_volatility * 2)

    return {"revenue": revenue_confidence, "margins": margins_confidence}


def chat_about_financials_unused_placeholder():
    pass


def fc_phase15_build_workbook(company: str, period: str, series: List[Dict[str, Any]],
                               drivers: Dict[str, Any], projections: List[Dict[str, Any]],
                               cashflow_forecast: List[Dict[str, Any]], scenarios: Dict[str, Any],
                               stakeholder: Dict[str, str], dcf: Dict[str, Any], risk: Dict[str, Any],
                               narrative: str, confidence: Dict[str, str]) -> Workbook:
    """Phase 15: Assemble the full 13-tab forecast Excel workbook."""
    wb = _new_workbook()

    # Tab 1: Historical Summary
    ws = wb.create_sheet("Historical Summary")
    row = _write_statement_header(ws, company, "Historical Financial Summary", period)
    headers = ["Metric"] + [str(s["year"]) for s in series]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    metrics = ["revenue", "cogs", "gross_profit", "operating_expenses", "ebit", "net_income",
               "total_assets", "total_liabilities", "total_equity", "cash"]
    for m in metrics:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, s in enumerate(series):
            cell = ws.cell(row=row, column=2 + j, value=s.get(m))
            cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(series))}})

    # Tab 2: Growth & Driver Assumptions
    ws = wb.create_sheet("Growth Assumptions")
    row = _write_statement_header(ws, company, "Growth Rate & Driver Assumptions", period)
    rows_data = [
        ("CAGR (Historical)", drivers["cagr"], PCT_FMT),
        ("Weighted-Average Growth", drivers["weighted_growth"], PCT_FMT),
        ("Linear Trend-Implied Growth", drivers["trend_growth"], PCT_FMT),
        ("Chosen Forward Growth Rate", drivers["chosen_growth_rate"], PCT_FMT),
        ("Average COGS / Revenue Ratio", drivers["avg_cogs_ratio"], PCT_FMT),
        ("Average Opex / Revenue Ratio", drivers["avg_expense_ratio"], PCT_FMT),
    ]
    for label, value, fmt in rows_data:
        row = _write_line_item(ws, row, label, value, currency=False)
        ws.cell(row=row - 1, column=4).number_format = fmt
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 3: Three-Statement Forecast
    ws = wb.create_sheet("3-Statement Forecast")
    row = _write_statement_header(ws, company, "Three-Statement Forecast Model", period, span=len(projections) + 1)
    headers = ["Metric"] + [f"Year +{p['year_offset']}" for p in projections]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["revenue", "cogs", "gross_profit", "operating_expenses", "operating_income",
              "net_income", "total_assets", "total_liabilities", "total_equity", "cash"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, p in enumerate(projections):
            cell = ws.cell(row=row, column=2 + j, value=p.get(m)); cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(projections))}})

    # Tab 4: Cash Flow Forecast
    ws = wb.create_sheet("Cash Flow Forecast")
    row = _write_statement_header(ws, company, "Cash Flow Forecast", period, span=len(cashflow_forecast) + 1)
    headers = ["Metric"] + [f"Year +{c['year_offset']}" for c in cashflow_forecast]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["operating_cf", "investing_cf", "financing_cf", "net_change_in_cash", "ending_cash", "runway_months"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, cf in enumerate(cashflow_forecast):
            val = cf.get(m)
            cell = ws.cell(row=row, column=2 + j, value=val)
            if m != "runway_months":
                cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(cashflow_forecast))}})

    # Tab 5: Scenario Analysis
    ws = wb.create_sheet("Scenario Analysis")
    row = _write_statement_header(ws, company, "Scenario Analysis (Base / Best / Worst)", period, span=4)
    headers = ["Scenario", "Growth Assumed", "Probability", f"Revenue Yr+{FORECAST_YEARS_OUT}"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for name, sc in scenarios.items():
        ws.cell(row=row, column=1, value=name.title()).font = BOLD_LABEL_FONT
        gc = ws.cell(row=row, column=2, value=sc["growth_assumed"]); gc.number_format = PCT_FMT
        pc = ws.cell(row=row, column=3, value=sc["probability"]); pc.number_format = PCT_FMT
        rc = ws.cell(row=row, column=4, value=sc["projections"][-1]["revenue"]); rc.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 16, 2: 16, 3: 14, 4: 18})

    # Tab 6: Valuation (DCF)
    ws = wb.create_sheet("Valuation (DCF)")
    row = _write_statement_header(ws, company, "DCF Valuation Model", period)
    row = _write_line_item(ws, row, "Discount Rate (WACC proxy)", dcf["discount_rate"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "Terminal Growth Rate", dcf["terminal_growth"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "PV of Explicit-Period FCF", dcf["pv_of_explicit_fcf"])
    row = _write_line_item(ws, row, "Terminal Value", dcf["terminal_value"])
    row = _write_line_item(ws, row, "PV of Terminal Value", dcf["pv_of_terminal_value"])
    row = _write_total_row(ws, row, "Enterprise Value", dcf["enterprise_value"])
    row = _write_line_item(ws, row, "Less: Net Debt (Estimate)", -dcf["net_debt_estimate"] if dcf["net_debt_estimate"] else 0)
    row = _write_total_row(ws, row, "Equity Value", dcf["equity_value"])
    _autosize_columns(ws, {1: 32, 4: 18})

    # Tab 7: Risk Analysis
    ws = wb.create_sheet("Risk Analysis")
    row = _write_statement_header(ws, company, "Risk Analysis Report", period)
    for dim, score in risk["dimensions"].items():
        row = _write_line_item(ws, row, dim.replace("_", " ").title(), score, currency=False)
        ws.cell(row=row - 1, column=4).number_format = '0.0'
    row = _write_total_row(ws, row, "Overall Risk Score (0-100)", risk["overall_score"])
    c = ws.cell(row=row, column=1, value=f"Risk Band: {risk['risk_band']}")
    c.font = Font(bold=True, color=(XL_GREEN if risk["overall_score"] < 25 else XL_RED if risk["overall_score"] >= 75 else "B8860B"))
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 8: Stakeholder Analysis
    ws = wb.create_sheet("Stakeholder Analysis")
    row = _write_statement_header(ws, company, "Stakeholder Analysis", period)
    for role, text in stakeholder.items():
        row = _write_section_title(ws, row, role.title())
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=text)
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 45
        row += 2
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 9: AI Narrative Insights
    ws = wb.create_sheet("AI Narrative")
    row = _write_statement_header(ws, company, "AI Narrative Insights", period)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=narrative)
    c.font = LABEL_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 120
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 10: Confidence Scores
    ws = wb.create_sheet("Confidence Scores")
    row = _write_statement_header(ws, company, "Forecast Confidence Scores", period)
    row = _write_line_item(ws, row, "Revenue Forecast Confidence", confidence["revenue"], currency=False)
    row = _write_line_item(ws, row, "Margin Forecast Confidence", confidence["margins"], currency=False)
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 11: Historical Ratios Trend
    ws = wb.create_sheet("Historical Ratio Trend")
    row = _write_statement_header(ws, company, "Historical Ratio Trend", period, span=len(series) + 1)
    ratio_hist = fc_phase4_historical_ratios(series)
    headers = ["Ratio"] + [str(r["year"]) for r in ratio_hist]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m, fmt in [("gross_margin", PCT_FMT), ("net_margin", PCT_FMT), ("current_ratio", X_FMT), ("debt_to_equity", X_FMT)]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, r in enumerate(ratio_hist):
            cell = ws.cell(row=row, column=2 + j, value=r.get(m))
            if r.get(m) is not None: cell.number_format = fmt
        row += 1
    _autosize_columns(ws, {1: 22, **{i: 12 for i in range(2, 2 + len(series))}})

    # Tab 12: Methodology & Assumptions Notes
    ws = wb.create_sheet("Methodology Notes")
    row = _write_statement_header(ws, company, "Methodology & Assumptions", period)
    notes = [
        "Revenue forecast uses a weighted-average year-over-year growth rate (more weight on recent years), "
        "falling back to historical CAGR if insufficient growth history exists.",
        "COGS and operating expenses are forecast as a constant percentage of revenue based on the historical average ratio.",
        "Balance sheet items (assets, liabilities) are scaled at a fraction of the revenue growth rate as a simplifying assumption.",
        "Cash flow forecast approximates operating cash flow from net income and assumes capex of 3% of revenue with no financing activity.",
        "DCF valuation discounts an unlevered free-cash-flow proxy at the specified discount rate, with Gordon Growth terminal value.",
        "This model is for planning and directional insight only — it is not a substitute for a full FP&A build-out or professional valuation.",
    ]
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"• {note}")
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 13: Cover / Disclaimer
    ws = wb.create_sheet("Cover", 0)
    ws.merge_cells("A1:D3")
    c = ws.cell(row=1, column=1, value=f"{company}\nFull Forecasting Package\n{period}")
    c.font = Font(size=16, bold=True, color=XL_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A5:D7")
    c2 = ws.cell(row=5, column=1, value="Generated by Quanto Financial Intelligence Platform — 15-Phase Forecasting Engine.\n\n"
                                         "Quanto is not responsible for financial decisions. This document is a model-based "
                                         "forecast built on historical trial balances and simplified assumptions; actual "
                                         "results will vary.")
    c2.font = Font(size=10, italic=True, color=XL_GREY)
    c2.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    _autosize_columns(ws, {1: 24, 2: 24, 3: 24, 4: 24})

    return wb


# =====================================================================================
# SECTION 7: FASTAPI ROUTES
# =====================================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    return HTMLResponse(content=HTML)


@app.get("/api/ocr-status")
async def api_ocr_status():
    return JSONResponse(get_ocr_status())


@app.get("/api/statements")
async def api_statements():
    return JSONResponse(STATEMENTS)


@app.post("/api/generate")
async def api_generate(statement_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    _enforce_free_plan_limit("financial_statements_generated", FREE_STATEMENT_LIMIT, STATEMENT_LIMIT_MESSAGE)

    extracted_docs: Dict[str, dict] = {}
    is_tb_only = statement_type in TB_ONLY_STATEMENTS

    if is_tb_only:
        if not files:
            raise HTTPException(status_code=400, detail="A trial balance file is required.")
        file = files[0]
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        extracted_docs["trial_balance"] = extract_document(content, file.filename or "upload")
    else:
        required_sources = set(info.get("sources", []))
        for file in files:
            raw_name = file.filename or ""
            if "::" in raw_name:
                source_key, original_name = raw_name.split("::", 1)
            else:
                source_key, original_name = (required_sources.pop() if required_sources else "unknown"), raw_name
            content = await file.read()
            if not content:
                continue
            extracted_docs[source_key] = extract_document(content, original_name or raw_name)

        missing = [s for s in info.get("sources", []) if s not in extracted_docs]
        if missing:
            missing_labels = [SOURCE_LABELS.get(m, (m, ""))[0] for m in missing]
            raise HTTPException(status_code=400, detail=f"Missing required source document(s): {', '.join(missing_labels)}")

    company, period = determine_company_and_period(extracted_docs)
    wb, analytics, accounts_found = build_statement_workbook(statement_type, extracted_docs, company, period)
    filename = save_workbook_and_get_filename(wb, prefix=statement_type)

    ocr_provider_used = next(iter(extracted_docs.values()), {}).get("_ocr_provider_used", get_ocr_status()["provider"])

    response_payload = {
        "statement_type": statement_type,
        "statement_label": info["label"],
        "company": company,
        "period": period,
        "accounts_found": accounts_found,
        "ocr_provider": ocr_provider_used,
        "filename": filename,
        "analytics": analytics,
    }

    if QUANTO_PLAN != "paid":
        _increment_usage("financial_statements_generated")

    return JSONResponse(response_payload)


@app.post("/api/insights")
async def api_insights(payload: Dict[str, Any]):
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    statement_type = payload.get("statement_type", "")
    insights = generate_ai_insights(analytics, company, period, statement_type)
    return JSONResponse({"insights": insights})


@app.post("/api/chat")
async def api_chat(payload: Dict[str, Any]):
    message = payload.get("message", "")
    history = payload.get("history", [])
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    if not message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty.")
    reply = chat_about_financials(message, history, analytics, company, period)
    return JSONResponse({"reply": reply})


@app.post("/api/forecast")
async def api_forecast(forecast_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(forecast_type)
    if not info or info.get("category") != "forecast":
        raise HTTPException(status_code=400, detail=f"Unknown forecast type: {forecast_type}")
    if len(files) < 3:
        raise HTTPException(status_code=400, detail="At least 3 fiscal years of trial balances are required.")

    _enforce_free_plan_limit("forecasts_generated", FREE_FORECAST_LIMIT, FORECAST_LIMIT_MESSAGE)

    yearly_docs = []
    for file in files:
        content = await file.read()
        if not content:
            continue
        yearly_docs.append(extract_document(content, file.filename or "upload"))

    if len(yearly_docs) < 3:
        raise HTTPException(status_code=400, detail="At least 3 valid (non-empty) trial balance files are required.")

    parsed_years, warnings = fc_phase1_2_validate(yearly_docs)
    series = fc_phase3_normalize(parsed_years)
    drivers = fc_phase5_6_drivers_and_growth(series)
    projections = fc_phase7_three_statement_model(series, drivers, years_out=FORECAST_YEARS_OUT)
    cashflow_forecast = fc_phase8_cashflow_forecast(projections, series[-1]["cash"])
    scenarios = fc_phase9_scenarios(series, drivers)

    latest_accounts = parsed_years[-1]["accounts"]
    analytics_latest = compute_full_analytics(latest_accounts)

    stakeholder = fc_phase10_stakeholder_analysis(series, projections, analytics_latest)
    dcf = fc_phase11_dcf_valuation(projections)
    risk = fc_phase12_risk_scoring(series, analytics_latest, cashflow_forecast)
    company, _ = determine_company_and_period({"_": yearly_docs[-1]})
    narrative = fc_phase13_narrative(series, projections, drivers, risk, company)
    confidence = fc_phase14_confidence_scores(series, drivers)

    period_label = f"{series[0]['year']}–{series[-1]['year']} Historical · {series[-1]['year']+1}–{series[-1]['year']+FORECAST_YEARS_OUT} Forecast"

    wb = fc_phase15_build_workbook(company, period_label, series, drivers, projections, cashflow_forecast,
                                    scenarios, stakeholder, dcf, risk, narrative, confidence)
    filename = save_workbook_and_get_filename(wb, prefix=forecast_type)

    response_payload = {
        "forecast_type": forecast_type,
        "forecast_label": info["label"],
        "company": company,
        "period": period_label,
        "years_analyzed": len(series),
        "phases_run": 15,
        "confidence_scores": confidence,
        "validation_warnings": warnings,
        "filename": filename,
    }

    if QUANTO_PLAN != "paid":
        _increment_usage("forecasts_generated")

    return JSONResponse(response_payload)

"""
core.py — Quanto Financial Intelligence Platform v1.0
=======================================================
Run:  uvicorn core:app --reload --port 8000
Then: http://localhost:8000

Requirements:
    python -m pip install fastapi uvicorn python-multipart pdfplumber pytesseract pillow openpyxl ollama numpy openai google-generativeai anthropic

OCR PROVIDERS (configure via environment variables):
    QUANTO_OCR_PROVIDER  = "claude" | "gemini" | "openai" | "local"  (default: claude)
    ANTHROPIC_API_KEY    = your Anthropic API key (recommended — most accurate, and PDFs are
                            sent natively without needing pdf2image/poppler installed)
    GEMINI_API_KEY       = your Google Gemini API key
    OPENAI_API_KEY       = your OpenAI API key
    (if no key is set, falls back to local pdfplumber + pytesseract)

NOTE: This build contains only the Statement Generator and Forecasting Engine.
The multi-entity Financial Consolidation Engine has been removed.
"""
from typing import Optional
import io, re, json, math, tempfile, statistics, base64, os, uuid
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse


app = FastAPI(title="Quanto", version="1.0.0")
OLLAMA_MODEL = "gemini-3-flash-preview:latest"
OUTPUT_DIR   = Path(tempfile.gettempdir()) / "quanto_outputs"
OUTPUT_DIR.mkdir(exist_ok=True)
from google import genai

OCR_PROVIDER = "gemini"

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
GEMINI_API_KEY = os.environ.get(
    "GEMINI_API_KEY",
    "AQ.Ab8RN6LnFOwbXFwSvH0K1A4X13d85DuEM4PtZ3bcQU6wiqhzrQ"
)

print("OCR_PROVIDER:", OCR_PROVIDER)
print("GEMINI KEY EXISTS:", bool(GEMINI_API_KEY))





print("OCR_PROVIDER:", OCR_PROVIDER)
print("GEMINI KEY EXISTS:", bool(GEMINI_API_KEY))
# =====================================================================================
# FREE-PLAN USAGE LIMITS
# =====================================================================================
# QUANTO_PLAN controls whether usage limits are enforced. "paid" plans are unrestricted.
QUANTO_PLAN = os.environ.get("QUANTO_PLAN", "free").lower()

FREE_STATEMENT_LIMIT = 100000
FREE_FORECAST_LIMIT = 100000

# Usage counters are persisted to disk (next to this script) so they survive app restarts.
USAGE_FILE = Path(__file__).resolve().parent / "quanto_usage.json"

STATEMENT_LIMIT_MESSAGE = "You have reached the free plan limit of 2 financial statements. Please upgrade your plan to continue."
FORECAST_LIMIT_MESSAGE = "You have reached the free plan limit of 2 forecasts. Please upgrade your plan to continue."


def _load_usage() -> Dict[str, int]:
    """Load persisted usage counters from disk. Returns defaults if the file is missing or unreadable."""
    default = {"financial_statements_generated": 0, "forecasts_generated": 0}
    if USAGE_FILE.exists():
        try:
            with open(USAGE_FILE, "r") as f:
                data = json.load(f)
            default.update({k: data.get(k, default[k]) for k in default})
        except Exception:
            pass
    return default


def _save_usage(usage: Dict[str, int]) -> None:
    """Persist usage counters to disk."""
    try:
        with open(USAGE_FILE, "w") as f:
            json.dump(usage, f)
    except Exception:
        pass


def _increment_usage(key: str) -> None:
    """Increment a single usage counter and persist it. Only called after successful generation."""
    usage = _load_usage()
    usage[key] = usage.get(key, 0) + 1
    _save_usage(usage)


def _enforce_free_plan_limit(usage_key: str, limit: int, error_message: str) -> None:
    """Raises an HTTPException (stopping execution) if a free-plan user has hit the given limit.
    Paid plans are never restricted."""
    if QUANTO_PLAN == "paid":
        return
    usage = _load_usage()
    if usage.get(usage_key, 0) >= limit:
        raise HTTPException(status_code=403, detail=error_message)

STATEMENTS = {
    "income_statement":      {"label":"Income Statement",                    "aliases":"P&L · Statement of Earnings · Statement of Operations","icon":"📈","category":"statements","sources":["trial_balance"]},
    "balance_sheet":         {"label":"Balance Sheet",                       "aliases":"Statement of Financial Position",                       "icon":"⚖️","category":"statements","sources":["trial_balance"]},
    "retained_earnings":     {"label":"Statement of Retained Earnings",      "aliases":"RE Rollforward",                                        "icon":"🔄","category":"statements","sources":["trial_balance"]},
    "equity_statement":      {"label":"Statement of Shareholders' Equity",   "aliases":"Changes in Equity · Owners' Equity",                    "icon":"🏛️","category":"statements","sources":["trial_balance"]},
    "trial_balance":         {"label":"Trial Balance",                       "aliases":"Adjusted / Unadjusted TB",                              "icon":"📋","category":"statements","sources":["trial_balance"]},
    "ratio_analysis":        {"label":"Financial Ratio Analysis",            "aliases":"Liquidity · Solvency · Profitability · Efficiency",      "icon":"📊","category":"statements","sources":["trial_balance"]},
    "liquidity_report":      {"label":"Liquidity Report",                    "aliases":"Current Ratio · Quick Ratio · Cash Ratio",               "icon":"💧","category":"statements","sources":["trial_balance"]},
    "solvency_report":       {"label":"Solvency Report",                     "aliases":"Debt-to-Equity · Leverage Analysis",                     "icon":"🏗️","category":"statements","sources":["trial_balance"]},
    "profitability_report":  {"label":"Profitability Report",                "aliases":"Margins · ROA · ROE",                                    "icon":"💰","category":"statements","sources":["trial_balance"]},
    "working_capital":       {"label":"Working Capital Report",              "aliases":"Cash Conversion Cycle · DIO · DPO",                      "icon":"⚙️","category":"statements","sources":["trial_balance"]},
    "cash_flow_statement":   {"label":"Cash Flow Statement",                 "aliases":"Operating · Investing · Financing Activities",           "icon":"💵","category":"statements","sources":["prior_balance_sheet","current_balance_sheet","income_statement","transaction_details"]},
    "ar_aging":              {"label":"Accounts Receivable Aging",           "aliases":"Customer Aging · Receivables Schedule",                  "icon":"📥","category":"statements","sources":["customer_invoices","due_dates","customer_balances"]},
    "ap_aging":              {"label":"Accounts Payable Aging",              "aliases":"Supplier Aging · Payables Schedule",                     "icon":"📤","category":"statements","sources":["supplier_invoices","due_dates","supplier_balances"]},
    "fixed_asset_schedule":  {"label":"Fixed Asset Schedule",                "aliases":"PPE Schedule · Depreciation Schedule",                   "icon":"🏭","category":"statements","sources":["fixed_asset_register","purchase_dates","depreciation_rates","useful_lives"]},
    "inventory_schedule":    {"label":"Inventory Schedule",                  "aliases":"Stock Schedule · FIFO / Weighted Average",               "icon":"📦","category":"statements","sources":["inventory_records","quantities","costing_method"]},
    "inventory_rollforward": {"label":"Inventory Rollforward",               "aliases":"Inventory Movement · Opening/Closing Stock",             "icon":"🔁","category":"statements","sources":["opening_inventory","purchases","sales","adjustments"]},
    "equity_rollforward":    {"label":"Equity Rollforward",                  "aliases":"Capital Account Movement",                               "icon":"📈","category":"statements","sources":["share_issues","dividends","owner_contributions","retained_earnings_movements"]},
    "debt_schedule":         {"label":"Debt Schedule",                       "aliases":"Loan Schedule · Debt Repayment Plan",                    "icon":"🏦","category":"statements","sources":["loan_agreements","repayment_schedules","interest_rates"]},
    "lease_schedule":        {"label":"Lease Schedule",                      "aliases":"IFRS 16 · Right-of-Use Assets",                          "icon":"🏢","category":"statements","sources":["lease_contracts","payment_schedules","lease_terms"]},
    "bank_reconciliation":   {"label":"Bank Reconciliation",                 "aliases":"Cash Reconciliation · Bank Recon",                       "icon":"🏧","category":"statements","sources":["bank_statements","cash_ledger"]},
    "account_reconciliation":{"label":"Account Reconciliations",            "aliases":"Subledger Reconciliation · Control Account",             "icon":"🔍","category":"statements","sources":["external_statements","subledgers","supporting_documents"]},
    "audit_working_papers":  {"label":"Audit Working Papers",                "aliases":"Lead Schedules · Audit File",                            "icon":"📝","category":"statements","sources":["general_ledger","lead_schedules","supporting_documentation"]},
    "notes_financial_stmts": {"label":"Notes to Financial Statements",       "aliases":"Disclosures · Accounting Policies",                      "icon":"📄","category":"statements","sources":["management_disclosures","accounting_policies","legal_information"]},
    "related_party":         {"label":"Related Party Disclosure Schedule",   "aliases":"RPT Schedule · Related Transactions",                    "icon":"🤝","category":"statements","sources":["related_party_transaction_data"]},
    "deferred_tax":          {"label":"Deferred Tax Schedule",               "aliases":"DTA · DTL · Tax Timing Differences",                     "icon":"🧮","category":"statements","sources":["tax_calculations","tax_returns","temporary_differences"]},
    "tax_provision":         {"label":"Tax Provision Workpapers",            "aliases":"Current & Deferred Tax · ETR Analysis",                  "icon":"💼","category":"statements","sources":["tax_returns","tax_adjustments","tax_rates"]},
    "revenue_recognition":   {"label":"Revenue Recognition Schedule",        "aliases":"ASC 606 · IFRS 15 · Contract Revenue",                   "icon":"📊","category":"statements","sources":["contracts","invoices","performance_obligations"]},
    "prepaid_expense":       {"label":"Prepaid Expense Schedule",            "aliases":"Prepayments · Deferred Charges",                         "icon":"⏳","category":"statements","sources":["payment_records","amortization_periods"]},
    "accrual_schedule":      {"label":"Accrual Schedule",                    "aliases":"Accrued Liabilities · Outstanding Obligations",          "icon":"📋","category":"statements","sources":["outstanding_invoices","contracts","unpaid_obligations"]},
    "employee_benefits":     {"label":"Employee Benefit Schedule",           "aliases":"Payroll · Pension · Benefits",                           "icon":"👥","category":"statements","sources":["payroll_records","pension_data","benefit_data"]},
    "forecast_growth":       {"label":"Growth Rate Forecast",                "aliases":"CAGR · Weighted · Trend · 1/3/5 Year",                   "icon":"🚀","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_driver":       {"label":"Driver-Based Forecast",               "aliases":"Revenue Drivers · Expense Drivers · Working Capital",    "icon":"🎯","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_three_stmt":   {"label":"Three-Statement Forecast Model",      "aliases":"Linked IS · BS · Cash Flow",                             "icon":"🔗","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_cashflow":     {"label":"Cash Flow Forecast",                  "aliases":"Operating · Investing · Financing · Runway",             "icon":"💵","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_scenarios":    {"label":"Scenario Analysis",                   "aliases":"Base · Best · Worst Case · Probability",                 "icon":"🎭","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_valuation":    {"label":"Valuation Model",                     "aliases":"DCF · Terminal Value · Enterprise Value · Equity Value",  "icon":"💎","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_risk":         {"label":"Risk Analysis Report",                "aliases":"Concentration · Liquidity · Debt · Burn · Score",        "icon":"⚠️","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_stakeholder":  {"label":"Stakeholder Analysis",                "aliases":"Owner · CFO · Investor · Bank · Auditor",                "icon":"👥","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_narrative":    {"label":"AI Narrative Insights",               "aliases":"Explainable Forecast Commentary",                       "icon":"🤖","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_full":         {"label":"Full Forecasting Package",            "aliases":"All 15 Phases — Complete Model",                         "icon":"📦","category":"forecast","sources":["trial_balance_3yr_min"]},
}

SOURCE_LABELS = {
    "trial_balance":                  ("Trial Balance",                  "The adjusted or unadjusted trial balance for the period"),
    "prior_balance_sheet":            ("Prior-Year Balance Sheet",       "Balance sheet from the previous fiscal year"),
    "current_balance_sheet":          ("Current-Year Balance Sheet",     "Balance sheet for the current fiscal year"),
    "income_statement":               ("Income Statement",               "Income statement / P&L for the period"),
    "transaction_details":            ("Transaction Details",            "Detailed list of cash transactions for the period"),
    "customer_invoices":              ("Customer Invoices",              "All outstanding customer invoices"),
    "due_dates":                      ("Due Dates",                      "Invoice or obligation due dates"),
    "customer_balances":              ("Customer Balances",              "Aged balances per customer"),
    "supplier_invoices":              ("Supplier Invoices",              "All outstanding supplier/vendor invoices"),
    "supplier_balances":              ("Supplier Balances",              "Aged balances per supplier"),
    "fixed_asset_register":           ("Fixed Asset Register",           "Complete register of all fixed assets"),
    "purchase_dates":                 ("Purchase Dates",                 "Acquisition dates for each asset"),
    "depreciation_rates":             ("Depreciation Rates",             "Depreciation rate or method per asset class"),
    "useful_lives":                   ("Useful Lives",                   "Estimated useful life per asset"),
    "inventory_records":              ("Inventory Records",              "Inventory listing with quantities and costs"),
    "quantities":                     ("Quantities on Hand",             "Physical count or system quantity per SKU"),
    "costing_method":                 ("Costing Method",                 "FIFO, Weighted Average, or Specific Identification"),
    "opening_inventory":              ("Opening Inventory",              "Inventory balance at start of period"),
    "purchases":                      ("Purchases During Period",        "All inventory purchases during the period"),
    "sales":                          ("Sales During Period",            "All inventory sold during the period"),
    "adjustments":                    ("Inventory Adjustments",          "Write-offs, shrinkage, returns, write-ups"),
    "share_issues":                   ("Share Issues",                   "New shares issued during the period"),
    "dividends":                      ("Dividends Declared",             "Dividends declared or paid during the period"),
    "owner_contributions":            ("Owner Contributions",            "Capital contributions by owners/shareholders"),
    "retained_earnings_movements":    ("Retained Earnings Movements",    "Adjustments to retained earnings"),
    "loan_agreements":                ("Loan Agreements",                "Signed loan/credit agreements"),
    "repayment_schedules":            ("Repayment Schedules",            "Amortization tables for all debt"),
    "interest_rates":                 ("Interest Rates",                 "Interest rate per loan (fixed or floating)"),
    "lease_contracts":                ("Lease Contracts",                "Signed lease agreements"),
    "payment_schedules":              ("Payment Schedules",              "Lease payment schedule over term"),
    "lease_terms":                    ("Lease Terms",                    "Lease commencement, end date, renewal options"),
    "bank_statements":                ("Bank Statements",                "Official bank statements for the period"),
    "cash_ledger":                    ("Cash Ledger",                    "Internal cash account / GL entries"),
    "external_statements":            ("External Statements",            "Third-party statements (bank, broker, etc.)"),
    "subledgers":                     ("Subledgers",                     "AR, AP, inventory subledger detail"),
    "supporting_documents":           ("Supporting Documents",           "Invoices, contracts, other evidence"),
    "general_ledger":                 ("General Ledger",                 "Full GL trial listing for the period"),
    "lead_schedules":                 ("Lead Schedules",                 "Audit lead schedules per balance area"),
    "supporting_documentation":       ("Supporting Documentation",       "Audit evidence, confirmations, workpapers"),
    "management_disclosures":         ("Management Disclosures",         "MD&A or notes prepared by management"),
    "accounting_policies":            ("Accounting Policies",            "Summary of significant accounting policies"),
    "legal_information":              ("Legal Information",              "Legal proceedings, contingencies, commitments"),
    "related_party_transaction_data": ("Related Party Transaction Data", "All transactions with related parties"),
    "tax_calculations":               ("Tax Calculations",               "Current and deferred tax computations"),
    "tax_returns":                    ("Tax Returns",                    "Filed or draft corporate tax returns"),
    "temporary_differences":          ("Temporary Differences",          "Taxable vs. accounting timing differences"),
    "tax_adjustments":                ("Tax Adjustments",                "Book-to-tax adjustments and reconciliations"),
    "tax_rates":                      ("Tax Rates",                      "Applicable statutory and effective tax rates"),
    "contracts":                      ("Customer Contracts",             "Signed contracts with performance obligations"),
    "invoices":                       ("Invoices",                       "Revenue invoices issued to customers"),
    "performance_obligations":        ("Performance Obligations",        "Identified POBs per contract"),
    "payment_records":                ("Payment Records",                "Evidence of prepayments made"),
    "amortization_periods":           ("Amortization Periods",           "Period over which prepaid expenses are expensed"),
    "outstanding_invoices":           ("Outstanding Invoices",           "Unpaid invoices at period end"),
    "unpaid_obligations":             ("Unpaid Obligations",             "Accrued but unbilled obligations"),
    "payroll_records":                ("Payroll Records",                "Employee payroll detail for the period"),
    "pension_data":                   ("Pension / Retirement Data",      "Defined benefit or contribution plan data"),
    "benefit_data":                   ("Employee Benefit Data",          "Health, bonus, stock compensation detail"),
    "trial_balance_3yr_min":          ("Trial Balances (Min 3 Years)",   "Upload at least 3 years of trial balances"),
}

TB_ONLY_STATEMENTS = {
    "income_statement","balance_sheet","retained_earnings","equity_statement",
    "trial_balance","ratio_analysis","liquidity_report","solvency_report",
    "profitability_report","working_capital",
}
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Quanto — Financial Intelligence</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#050807;--surface:#0b100d;--surface2:#111713;--surface3:#171f1a;
  --text:#EEF3EF;--text-muted:#93A19A;--text-dim:#5c6b63;
  --accent:#00C853;--accent2:#16A34A;--gold:#D4AF37;
  --green:#22C55E;--red:#EF4444;--orange:#f6993f;
  --border:#1D3125;--border2:#2B4A37;--radius:10px;
  --glow:0 0 30px rgba(0,200,83,0.08);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:'Inter',system-ui,sans-serif;min-height:100vh;overflow-x:hidden}
body::before{content:'';position:fixed;inset:0;z-index:-2;
  background:radial-gradient(ellipse 80% 50% at 50% -10%,rgba(0,200,83,0.07),transparent),
             radial-gradient(ellipse 50% 60% at 90% 50%,rgba(22,163,74,0.04),transparent)}
body::after{content:'';position:fixed;inset:0;z-index:-1;
  background-image:linear-gradient(rgba(29,49,37,0.28) 1px,transparent 1px),
                   linear-gradient(90deg,rgba(29,49,37,0.28) 1px,transparent 1px);
  background-size:48px 48px}

header{padding:3rem 2rem 2.5rem;text-align:center;border-bottom:1px solid var(--border);
  background:linear-gradient(180deg,rgba(11,16,13,0.96),rgba(11,16,13,0.72));backdrop-filter:blur(14px);
  position:sticky;top:0;z-index:100;box-shadow:0 1px 40px rgba(0,200,83,0.05)}
.brand{display:flex;align-items:center;justify-content:center;gap:.75rem;margin-bottom:.5rem}
.brand-name{font-family:'DM Serif Display',serif;font-size:2.1rem;letter-spacing:-.04em;
  background:linear-gradient(135deg,#EEF3EF,var(--accent));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.tagline{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;letter-spacing:.14em;text-transform:uppercase}
.ocr-badge{display:inline-flex;align-items:center;gap:.5rem;margin-top:.9rem;padding:.35rem .85rem;
  border:1px solid var(--border2);border-radius:9999px;font-family:'DM Mono',monospace;font-size:.72rem;color:var(--green);
  background:rgba(34,197,94,0.05)}
.disclaimer{font-size:.68rem;color:var(--text-dim);font-family:'DM Mono',monospace;margin-top:.5rem;letter-spacing:.03em}

.tab-bar{display:flex;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:99}
.tab-btn{flex:1;padding:1rem 1.5rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;font-size:.78rem;
  background:transparent;color:var(--text-muted);border:none;cursor:pointer;transition:all .2s;font-family:'Inter',sans-serif}
.tab-btn:hover{background:var(--surface2);color:var(--text)}
.tab-btn.active{background:var(--surface2);color:var(--accent);border-bottom:2px solid var(--accent)}
.tab-panel{display:none}.tab-panel.active{display:block}

main{max-width:1300px;margin:0 auto;padding:2.5rem 1.5rem 6rem;display:grid;gap:1.75rem}

.step-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);
  padding:2rem;position:relative;transition:border-color .2s,box-shadow .2s;box-shadow:var(--glow)}
.step-card:hover{border-color:var(--border2)}
.step-num{position:absolute;top:-11px;left:22px;background:linear-gradient(135deg,var(--accent),var(--accent2));
  color:#04160c;font-family:'DM Mono',monospace;font-weight:700;padding:.28rem .75rem;
  border-radius:9999px;font-size:.75rem;letter-spacing:.05em}
.step-title{font-size:1.2rem;font-weight:600;margin-bottom:.25rem;color:var(--text)}
.step-sub{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;margin-bottom:1.25rem}

.filter-bar{display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:1.25rem}
.filter-btn{padding:.32rem .8rem;border-radius:9999px;font-size:.75rem;font-weight:500;cursor:pointer;transition:all .2s;border:1px solid var(--border2);background:var(--surface3);color:var(--text-muted)}
.filter-btn.active{background:var(--accent);color:#04160c;border-color:var(--accent)}

.stmt-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem}
.stmt-card{border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem 1.1rem;
  background:var(--surface3);cursor:pointer;transition:all .2s;position:relative}
.stmt-card:hover{border-color:var(--accent);background:#1a2420;transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,200,83,0.1)}
.stmt-card.selected{border-color:var(--accent);background:#1a2420;box-shadow:0 0 0 3px rgba(0,200,83,.14)}
.stmt-card.multisource{border-color:rgba(212,175,55,0.22)}
.stmt-card.multisource:hover,.stmt-card.multisource.selected{border-color:var(--gold)}
.stmt-card.multisource.selected{box-shadow:0 0 0 3px rgba(212,175,55,.14)}
.stmt-icon{font-size:1.75rem;margin-bottom:.6rem}
.stmt-name{font-weight:600;font-size:.95rem;margin-bottom:.2rem;color:var(--text)}
.stmt-alias{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;line-height:1.4}
.source-badge{display:inline-block;margin-top:.5rem;padding:.18rem .45rem;background:rgba(212,175,55,0.1);
  border:1px solid rgba(212,175,55,0.3);border-radius:4px;font-size:.65rem;color:var(--gold)}

.source-required{background:var(--surface3);border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem;margin-top:1rem}
.source-required h4{color:var(--accent);font-size:.8rem;margin-bottom:.875rem;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em}
.source-item{display:flex;align-items:flex-start;gap:.875rem;padding:.65rem;border-radius:7px;background:#1a2420;margin-bottom:.4rem}
.source-item-icon{font-size:1.1rem;flex-shrink:0;margin-top:.1rem}
.source-item-text{flex:1}
.source-item-title{font-weight:600;font-size:.85rem;margin-bottom:.15rem}
.source-item-desc{font-size:.75rem;color:var(--text-muted)}
.source-upload-btn{padding:.38rem .8rem;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;
  color:var(--text-muted);font-size:.75rem;cursor:pointer;transition:all .2s;white-space:nowrap}
.source-upload-btn.uploaded{background:rgba(34,197,94,0.1);border-color:var(--green);color:var(--green)}

.upload-area{border:2px dashed var(--border2);border-radius:var(--radius);padding:2.5rem 1.5rem;text-align:center;
  transition:all .2s;background:var(--surface3);cursor:pointer}
.upload-area:hover{border-color:var(--accent);background:#151f19;box-shadow:inset 0 0 0 1px rgba(0,200,83,0.12)}
.file-tag{display:inline-flex;align-items:center;gap:.5rem;padding:.35rem .7rem;background:var(--surface3);
  border:1px solid var(--border2);border-radius:6px;font-size:.8rem;margin:.2rem}

.submit-btn{width:100%;padding:.9rem 2rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;font-size:.95rem;letter-spacing:.04em;text-transform:uppercase;border:none;
  border-radius:var(--radius);cursor:pointer;transition:all .2s;margin-top:.75rem;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.submit-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}
.submit-btn.forecast-btn{background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400;box-shadow:0 4px 20px rgba(212,175,55,0.25)}
.submit-btn.forecast-btn:hover{box-shadow:0 8px 28px rgba(212,175,55,0.35)}
.submit-btn:disabled{opacity:.4;cursor:not-allowed;transform:none;box-shadow:none}

.progress-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:1.75rem}
.result-card{background:var(--surface2);border:1px solid var(--accent);border-radius:var(--radius);padding:2.5rem;text-align:center;
  box-shadow:0 0 40px rgba(0,200,83,0.1)}
.result-card.forecast-result{border-color:var(--gold);box-shadow:0 0 40px rgba(212,175,55,0.1)}
.download-btn{display:inline-flex;align-items:center;gap:.75rem;padding:.9rem 2.1rem;
  background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;border-radius:var(--radius);text-decoration:none;margin-top:1.5rem;transition:all .2s;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.download-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}

.step-item{padding:.65rem 1rem;border-radius:6px;font-family:'DM Mono',monospace;font-size:.82rem;margin-bottom:.2rem}
.step-item.active{color:var(--accent);background:rgba(0,200,83,0.06)}
.step-item.done{color:var(--green)}.step-item.error{color:var(--red)}
.spinner{display:inline-block;width:.85rem;height:.85rem;border:2px solid var(--border2);border-top-color:var(--accent);
  border-radius:50%;animation:spin 0.7s linear infinite;vertical-align:middle;margin-right:.5rem}
@keyframes spin{to{transform:rotate(360deg)}}

.notice-bar{padding:.7rem 1.1rem;border-radius:7px;font-size:.8rem;font-family:'DM Mono',monospace;margin-bottom:.875rem}
.notice-bar.info{background:rgba(0,200,83,0.06);border:1px solid rgba(0,200,83,0.22);color:var(--accent)}
.notice-bar.warn{background:rgba(246,153,63,0.06);border:1px solid rgba(246,153,63,0.2);color:var(--orange)}

.analytics-panel{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;margin-top:0;box-shadow:var(--glow)}
.analytics-header{padding:1.5rem 2rem 1rem;background:linear-gradient(135deg,var(--surface3),#1a2420);border-bottom:1px solid var(--border)}
.analytics-header h3{font-size:1.1rem;font-weight:600;color:var(--text);margin-bottom:.25rem}
.analytics-header p{font-size:.8rem;color:var(--text-muted);font-family:'DM Mono',monospace}

.analytics-tabs{display:flex;border-bottom:1px solid var(--border);background:var(--surface2)}
.analytics-tab{padding:.75rem 1.25rem;font-size:.78rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;
  color:var(--text-muted);cursor:pointer;border:none;background:transparent;transition:all .2s;border-bottom:2px solid transparent}
.analytics-tab:hover{color:var(--text);background:var(--surface3)}
.analytics-tab.active{color:var(--accent);border-bottom-color:var(--accent);background:var(--surface3)}

.analytics-content{padding:1.75rem 2rem}
.analytics-sub-panel{display:none}.analytics-sub-panel.active{display:block}

.kpi-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem;margin-bottom:1.5rem}
.kpi-card{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.1rem;transition:border-color .2s,transform .2s}
.kpi-card:hover{border-color:var(--border2);transform:translateY(-1px)}
.kpi-label{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.35rem}
.kpi-value{font-size:1.4rem;font-weight:700;color:var(--text);line-height:1}
.kpi-value.positive{color:var(--green)}.kpi-value.negative{color:var(--red)}.kpi-value.neutral{color:var(--accent)}
.kpi-note{font-size:.7rem;color:var(--text-dim);margin-top:.3rem;font-family:'DM Mono',monospace}

.insight-block{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.25rem;margin-bottom:1rem}
.insight-block h4{font-size:.82rem;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem;font-family:'DM Mono',monospace}
.insight-block p,.insight-block li{font-size:.88rem;color:var(--text);line-height:1.7}
.insight-block ul{padding-left:1.1rem}
.insight-block li{margin-bottom:.25rem}

.ratio-section-title{font-size:.75rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;
  letter-spacing:.1em;font-family:'DM Mono',monospace;padding:.5rem 0;margin-top:1.25rem;margin-bottom:.5rem;
  border-bottom:1px solid var(--border)}
.ratio-table{width:100%;border-collapse:collapse;margin-bottom:1rem}
.ratio-table th{font-size:.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:.06em;
  padding:.5rem .75rem;text-align:left;border-bottom:1px solid var(--border);font-family:'DM Mono',monospace}
.ratio-table th:last-child{text-align:right}
.ratio-table td{padding:.55rem .75rem;font-size:.83rem;border-bottom:1px solid rgba(29,49,37,0.5)}
.ratio-table tr:last-child td{border-bottom:none}
.ratio-table tr:hover td{background:rgba(0,200,83,0.03)}
.ratio-table td:last-child{text-align:right;font-weight:600;font-family:'DM Mono',monospace}
.ratio-val.good{color:var(--green)}.ratio-val.warn{color:var(--orange)}.ratio-val.bad{color:var(--red)}.ratio-val.neutral{color:var(--accent)}

.chat-container{display:flex;flex-direction:column;height:520px;background:var(--surface3);border-radius:0 0 var(--radius) var(--radius)}
.chat-messages{flex:1;overflow-y:auto;padding:1.5rem;display:flex;flex-direction:column;gap:.875rem;scroll-behavior:smooth}
.chat-messages::-webkit-scrollbar{width:4px}
.chat-messages::-webkit-scrollbar-track{background:transparent}
.chat-messages::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px}
.chat-msg{max-width:85%;border-radius:10px;padding:.875rem 1.1rem;line-height:1.65;font-size:.875rem}
.chat-msg.user{align-self:flex-end;background:linear-gradient(135deg,rgba(0,200,83,0.22),rgba(22,163,74,0.16));
  border:1px solid rgba(0,200,83,0.28);color:var(--text)}
.chat-msg.assistant{align-self:flex-start;background:#1a2420;border:1px solid var(--border);color:var(--text)}
.chat-msg.assistant .msg-label{font-size:.68rem;font-weight:700;color:var(--accent);font-family:'DM Mono',monospace;
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:.4rem}
.chat-msg.typing{padding:.875rem 1.1rem}
.typing-dot{display:inline-block;width:7px;height:7px;background:var(--text-muted);border-radius:50%;margin:0 2px;
  animation:bounce .9s ease-in-out infinite}
.typing-dot:nth-child(2){animation-delay:.15s}.typing-dot:nth-child(3){animation-delay:.3s}
@keyframes bounce{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-5px)}}
.chat-input-row{display:flex;gap:.5rem;padding:1rem 1.25rem;border-top:1px solid var(--border);background:var(--surface2)}
.chat-input{flex:1;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;
  padding:.65rem 1rem;color:var(--text);font-size:.875rem;font-family:'Inter',sans-serif;resize:none;outline:none;
  transition:border-color .2s,box-shadow .2s;min-height:40px;max-height:100px}
.chat-input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,200,83,0.14)}
.chat-send-btn{padding:.65rem 1.1rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  border:none;border-radius:8px;cursor:pointer;font-weight:600;font-size:.82rem;transition:all .2s;white-space:nowrap}
.chat-send-btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,200,83,0.35)}
.chat-send-btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
.chat-empty{text-align:center;padding:2.5rem;color:var(--text-muted);font-family:'DM Mono',monospace;font-size:.82rem}
.chat-empty .chat-empty-icon{font-size:2rem;margin-bottom:.75rem;opacity:.5}

.ledger-upload-row{display:flex;gap:.75rem;align-items:center;padding:1rem 1.25rem;border-bottom:1px solid var(--border);background:var(--surface2);flex-wrap:wrap}
.ledger-state-block{background:#1a2420;border:1px solid var(--border);border-radius:8px;padding:.9rem 1.1rem;margin-bottom:.6rem;font-family:'DM Mono',monospace;font-size:.8rem;white-space:pre-wrap;line-height:1.6}
.ledger-badge{display:inline-block;padding:.15rem .5rem;border-radius:4px;font-size:.68rem;font-family:'DM Mono',monospace;background:rgba(0,200,83,0.12);color:var(--accent);border:1px solid rgba(0,200,83,0.25)}

footer{text-align:center;padding:2rem;color:var(--text-dim);font-family:'DM Mono',monospace;font-size:.72rem;border-top:1px solid var(--border);letter-spacing:.03em}
footer span{color:var(--text-muted)}

input:focus, textarea:focus, select:focus{outline:none}
::selection{background:rgba(0,200,83,0.28);color:#fff}
</style>
</head>
<body>
<header>
  <div class="brand">
    <div class="brand-name">Quanto</div>
  </div>
  <p class="tagline">Financial Intelligence Platform</p>
  <div class="ocr-badge" id="ocrBadge">⬜ Loading OCR status...</div>
  <p class="disclaimer">Quanto is not responsible for financial decisions.</p>
</header>

<div class="tab-bar">
  <button class="tab-btn active" onclick="switchTab('statements',this)">📋 Statement Generator</button>
  <button class="tab-btn" onclick="switchTab('forecasting',this)">🚀 Forecasting Engine</button>
  <button class="tab-btn" onclick="switchTab('ledger',this)">📒 Interactive Ledger</button>
</div>

<div id="tab-statements" class="tab-panel active">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose a Financial Statement</div>
    <div class="notice-bar info">📋 Statements with a purple badge require multiple source documents — required uploads will appear automatically.</div>
    <div class="filter-bar">
      <button class="filter-btn active" id="fAll" onclick="filterCards('all',this)">All</button>
      <button class="filter-btn" id="fTb" onclick="filterCards('tb',this)">Trial Balance Only</button>
      <button class="filter-btn" id="fMulti" onclick="filterCards('multi',this)">Multi-Source</button>
    </div>
    <div class="stmt-grid" id="stmtGrid"></div>
  </section>

  <section class="step-card" id="uploadSection">
    <span class="step-num">02</span>
    <div class="step-title" id="uploadTitle">Upload Required Documents</div>
    <div class="step-sub" id="uploadSub">Select a statement type above to see required documents.</div>
    <div id="sourceRequiredArea"></div>
    <div id="tbUploadArea" style="display:none">
      <div class="upload-area" id="uploadArea">
        <input type="file" id="fileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none"/>
        <div style="font-size:2rem;margin:.75rem 0">📎</div>
        <p style="font-weight:600;font-size:.95rem">Drop Trial Balance here or click to browse</p>
        <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">PDF, JPG, PNG, TIFF supported · AI Vision OCR</p>
      </div>
      <div id="tbFileChosen" style="display:none;margin-top:.6rem;padding:.65rem 1rem;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;align-items:center">
        <span id="tbFileName" style="flex:1;font-size:.85rem"></span>
        <button onclick="removeTbFile()" style="background:none;border:none;color:var(--red);font-size:1.1rem;cursor:pointer;margin-left:1rem">✕</button>
      </div>
    </div>
  </section>

  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn" id="generateBtn" onclick="generate()" disabled>⚡ Generate Statement</button>
  </section>

  <div id="status-area" style="display:none"></div>
  <div id="result-area"></div>

  <div id="analytics-area" style="display:none">
    <div class="analytics-panel">
      <div class="analytics-header">
        <h3 id="analytics-company-title">Financial Analysis</h3>
        <p id="analytics-period-sub">Generated by Quanto Intelligence Engine</p>
      </div>
      <div class="analytics-tabs">
        <button class="analytics-tab active" onclick="switchAnalyticsTab('overview',this)">📊 Overview</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('ratios',this)">🔢 Ratio Analysis</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('insights',this)">💡 Insights</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('chat',this)">💬 Ask Quanto</button>
      </div>
      <div class="analytics-content">
        <div id="panel-overview" class="analytics-sub-panel active">
          <div id="kpi-grid-container"></div>
        </div>
        <div id="panel-ratios" class="analytics-sub-panel">
          <div id="ratios-container"></div>
        </div>
        <div id="panel-insights" class="analytics-sub-panel">
          <div id="insights-container">
            <div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem">
              <span class="spinner"></span> Generating AI insights...
            </div>
          </div>
        </div>
        <div id="panel-chat" class="analytics-sub-panel">
          <div class="chat-container">
            <div class="chat-messages" id="chatMessages">
              <div class="chat-empty">
                <div class="chat-empty-icon">💬</div>
                <p>Ask anything about this financial statement.</p>
                <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is driving the revenue growth?" or "Is this company financially healthy?"</p>
              </div>
            </div>
            <div class="chat-input-row">
              <textarea class="chat-input" id="chatInput" placeholder="Ask Quanto about this financial statement..." rows="1"
                onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendChat()}"></textarea>
              <button class="chat-send-btn" id="chatSendBtn" onclick="sendChat()">Send ↑</button>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</main>
</div>

<div id="tab-forecasting" class="tab-panel">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose Forecast Type</div>
    <div class="stmt-grid" id="fstmtGrid"></div>
  </section>
  <section class="step-card">
    <span class="step-num">02</span>
    <div class="step-title">Upload Historical Trial Balances</div>
    <div class="notice-bar warn">⚠ Minimum 3 fiscal years of trial balances required. Upload one file per year.</div>
    <div id="forecastDropZone" class="upload-area" style="margin-bottom:1rem">
      <input type="file" id="forecastFileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" multiple style="display:none"/>
      <div style="font-size:2rem;margin:.75rem 0">📁</div>
      <p style="font-weight:600;font-size:.95rem">Drop files here or click to browse</p>
      <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">Select multiple files — one per fiscal year</p>
    </div>
    <div id="forecastFileList" style="display:flex;flex-wrap:wrap;gap:.4rem"></div>
  </section>
  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn forecast-btn" id="fgenerateBtn" onclick="generateForecast()" disabled>🚀 Run Forecast Engine (15 Phases)</button>
  </section>
  <div id="fstatus-area" style="display:none"></div>
  <div id="fresult-area"></div>
</main>
</div>

<div id="tab-ledger" class="tab-panel">
<main>
  <section class="step-card">
    <span class="step-num">📒</span>
    <div class="step-title">Interactive Financial Ledger</div>
    <div class="step-sub">Upload a trial balance / ledger document to start a stateful accounting session, then issue adjustments in plain English (e.g. "increase salaries to 50,000", "remove rent expense", "undo").</div>
    <div class="notice-bar info">All figures are computed deterministically in Python from the maintained ledger state — nothing is estimated by the AI.</div>
    <div class="analytics-panel">
      <div class="ledger-upload-row">
        <input type="file" id="ledgerFileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none"/>
        <button class="submit-btn" style="width:auto;margin:0" onclick="document.getElementById('ledgerFileInput').click()">📎 Upload Trial Balance</button>
        <span id="ledgerFileStatus" style="font-size:.8rem;color:var(--text-muted)">No document loaded yet.</span>
        <span id="ledgerSessionBadge" style="display:none" class="ledger-badge">Session Active</span>
      </div>
      <div class="chat-container" style="height:600px">
        <div class="chat-messages" id="ledgerMessages">
          <div class="chat-empty">
            <div class="chat-empty-icon">📒</div>
            <p>Upload a trial balance to initialize the ledger session.</p>
            <p style="margin-top:.35rem;color:var(--text-dim)">Then try: "increase rent to 15,000" · "remove utilities" · "add fuel expense of 2,500" · "undo"</p>
          </div>
        </div>
        <div class="chat-input-row">
          <textarea class="chat-input" id="ledgerInput" placeholder="Issue an accounting command..." rows="1" disabled
            onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendLedgerCommand()}"></textarea>
          <button class="chat-send-btn" id="ledgerSendBtn" onclick="sendLedgerCommand()" disabled>Send ↑</button>
        </div>
      </div>
    </div>
  </section>
</main>
</div>

<footer>
  <span>Quanto Financial Intelligence Platform v1.0</span> &nbsp;·&nbsp; AI Vision OCR &nbsp;·&nbsp; Llama 3.1 Narrative Engine &nbsp;·&nbsp; 100% Local Processing
  <br><span style="color:var(--text-dim)">Quanto is not responsible for financial decisions.</span>
</footer>
<script>
const TB_ONLY = new Set([
  "income_statement","balance_sheet","retained_earnings","equity_statement",
  "trial_balance","ratio_analysis","liquidity_report","solvency_report",
  "profitability_report","working_capital"
]);

let selectedStmt = null;
let selectedFStmt = null;
let tbFile = null;
let multiSourceFiles = {};
let forecastFiles = [];
let allStatements = {};
let currentAnalyticsData = null;
let chatHistory = [];
let ledgerSessionId = null;

const SOURCE_LABELS_JS = """ + json.dumps({k: list(v) for k,v in SOURCE_LABELS.items()}) + r""";

fetch('/api/ocr-status').then(r=>r.json()).then(d=>{
  const b = document.getElementById('ocrBadge');
  const icons = {claude:'🟣', gemini:'🟢', openai:'🔵', local:'🟡'};
  b.textContent = `${icons[d.provider]||'⚪'} OCR: ${d.label} — ${d.note}`;
});

fetch('/api/statements').then(r=>r.json()).then(data=>{
  allStatements = data;
  const sg = document.getElementById('stmtGrid');
  const fg = document.getElementById('fstmtGrid');
  Object.entries(data).forEach(([key, info])=>{
    const isForecast = info.category === 'forecast';
    const isMulti = !TB_ONLY.has(key) && !isForecast;
    const card = document.createElement('label');
    card.className = 'stmt-card' + (isForecast?' forecast-card':'') + (isMulti?' multisource':'');
    card.setAttribute('data-key', key);
    card.setAttribute('data-category', isForecast ? 'forecast' : (TB_ONLY.has(key) ? 'tb' : 'multi'));
    card.innerHTML = `<input type="radio" class="stmt-radio" name="${isForecast?'fstatement':'statement'}" value="${key}" style="display:none"/>
      <div class="stmt-icon">${info.icon}</div>
      <div class="stmt-name">${info.label}</div>
      <div class="stmt-alias">${info.aliases}</div>
      ${isMulti ? '<span class="source-badge">Multi-Source</span>' : ''}`;
    card.addEventListener('click', ()=>{
      if(isForecast){
        document.querySelectorAll('.forecast-card').forEach(c=>c.classList.remove('selected'));
        selectedFStmt = key;
        document.getElementById('fgenerateBtn').disabled = forecastFiles.length < 3;
      } else {
        document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>c.classList.remove('selected'));
        selectedStmt = key;
        renderUploadSection(key, info);
        checkGenerateReady();
      }
      card.classList.add('selected');
    });
    (isForecast ? fg : sg).appendChild(card);
  });
});

function filterCards(type, btn){
  document.querySelectorAll('.filter-btn').forEach(b=>{b.classList.remove('active')});
  btn.classList.add('active');
  document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>{
    const cat = c.getAttribute('data-category');
    c.style.display = (type==='all' || cat===type) ? '' : 'none';
  });
}

function renderUploadSection(key, info){
  const sources = info.sources || [];
  const isTbOnly = TB_ONLY.has(key);
  document.getElementById('uploadTitle').textContent = isTbOnly ? 'Upload Trial Balance' : 'Upload Required Source Documents';
  document.getElementById('uploadSub').textContent = isTbOnly
    ? 'This statement is generated from a single trial balance.'
    : `This statement requires ${sources.length} source document(s). Upload all to proceed.`;
  document.getElementById('tbUploadArea').style.display = isTbOnly ? 'block' : 'none';
  document.getElementById('sourceRequiredArea').style.display = isTbOnly ? 'none' : 'block';
  if(isTbOnly) return;
  multiSourceFiles = {};
  const area = document.getElementById('sourceRequiredArea');
  area.innerHTML = '';
  const box = document.createElement('div');
  box.className = 'source-required';
  box.innerHTML = '<h4>📎 Required Documents</h4>';
  sources.forEach(srcKey=>{
    const [title, desc] = SOURCE_LABELS_JS[srcKey] || [srcKey, ''];
    const item = document.createElement('div');
    item.className = 'source-item';
    item.id = 'src-item-' + srcKey;
    item.innerHTML = `
      <div class="source-item-icon">📄</div>
      <div class="source-item-text">
        <div class="source-item-title">${title}</div>
        <div class="source-item-desc">${desc}</div>
      </div>
      <div>
        <input type="file" id="src-file-${srcKey}" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none" onchange="handleSourceFile('${srcKey}', this)"/>
        <button class="source-upload-btn" id="src-btn-${srcKey}" onclick="document.getElementById('src-file-${srcKey}').click()">Upload ↑</button>
      </div>`;
    box.appendChild(item);
  });
  area.appendChild(box);
}

function handleSourceFile(srcKey, input){
  if(input.files[0]){
    multiSourceFiles[srcKey] = input.files[0];
    const btn = document.getElementById('src-btn-'+srcKey);
    btn.className = 'source-upload-btn uploaded';
    btn.textContent = '✓ ' + input.files[0].name.substring(0,22);
    checkGenerateReady();
  }
}

const fileInput = document.getElementById('fileInput');
const uploadArea = document.getElementById('uploadArea');
uploadArea.addEventListener('click', ()=>fileInput.click());
fileInput.addEventListener('change', e=>{ if(e.target.files[0]) pickTbFile(e.target.files[0]); });
uploadArea.addEventListener('dragover', e=>{e.preventDefault();uploadArea.style.borderColor='var(--accent)'});
uploadArea.addEventListener('dragleave', ()=>{uploadArea.style.borderColor=''});
uploadArea.addEventListener('drop', e=>{e.preventDefault();uploadArea.style.borderColor='';if(e.dataTransfer.files[0])pickTbFile(e.dataTransfer.files[0])});

function pickTbFile(f){
  tbFile = f;
  document.getElementById('tbFileName').textContent = f.name;
  const fc = document.getElementById('tbFileChosen');
  fc.style.display = 'flex';
  checkGenerateReady();
}
function removeTbFile(){
  tbFile = null;
  document.getElementById('tbFileChosen').style.display = 'none';
  fileInput.value = '';
  checkGenerateReady();
}
function checkGenerateReady(){
  if(!selectedStmt){document.getElementById('generateBtn').disabled=true;return;}
  if(TB_ONLY.has(selectedStmt)){
    document.getElementById('generateBtn').disabled = !tbFile;
  } else {
    const info = allStatements[selectedStmt]||{};
    const required = info.sources||[];
    document.getElementById('generateBtn').disabled = !required.every(s=>multiSourceFiles[s]);
  }
}

const forecastDz = document.getElementById('forecastDropZone');
const forecastFi = document.getElementById('forecastFileInput');
forecastDz.addEventListener('click', ()=>forecastFi.click());
forecastFi.addEventListener('change', e=>{
  Array.from(e.target.files).forEach(f=>{if(!forecastFiles.find(x=>x.name===f.name))forecastFiles.push(f)});
  renderForecastFiles();
});
function renderForecastFiles(){
  const list = document.getElementById('forecastFileList');
  list.innerHTML = '';
  forecastFiles.forEach((f,i)=>{
    const tag = document.createElement('div');
    tag.className = 'file-tag';
    tag.innerHTML = `📄 ${f.name} <button onclick="removeForecastFile(${i})" style="background:none;border:none;color:var(--red);cursor:pointer;margin-left:.3rem">✕</button>`;
    list.appendChild(tag);
  });
  document.getElementById('fgenerateBtn').disabled = !(selectedFStmt && forecastFiles.length >= 3);
  if(forecastFiles.length>0 && forecastFiles.length<3){
    const warn = document.createElement('div');
    warn.style.cssText='width:100%;color:var(--orange);font-size:.77rem;margin-top:.4rem;font-family:"DM Mono",monospace';
    warn.textContent = `⚠ ${forecastFiles.length}/3 files — minimum 3 required`;
    list.appendChild(warn);
  }
}
function removeForecastFile(i){forecastFiles.splice(i,1);renderForecastFiles();}

function switchTab(name, btn){
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  btn.classList.add('active');
}

function switchAnalyticsTab(name, btn){
  document.querySelectorAll('.analytics-sub-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.analytics-tab').forEach(b=>b.classList.remove('active'));
  document.getElementById('panel-'+name).classList.add('active');
  btn.classList.add('active');
}

function fmt(v, type){
  if(v===null||v===undefined||v==='N/A')return'N/A';
  const n = parseFloat(v);
  if(isNaN(n)) return String(v);
  if(type==='pct') return (n*100).toFixed(1)+'%';
  if(type==='x') return n.toFixed(2)+'x';
  if(type==='$') return '$'+n.toLocaleString('en-US',{minimumFractionDigits:0,maximumFractionDigits:0});
  if(type==='days') return n.toFixed(1)+' days';
  return n.toFixed(2);
}
function scoreClass(v, good, warn){
  if(v===null||v===undefined||isNaN(parseFloat(v)))return'neutral';
  const n=parseFloat(v);
  if(n>=good)return'good';if(n>=warn)return'warn';return'bad';
}

function renderKPIs(d){
  const g = document.getElementById('kpi-grid-container');
  g.innerHTML = '';
  const grid = document.createElement('div'); grid.className='kpi-grid';
  const kpis = [
    {label:'Revenue',val:fmt(d.revenue,'$'),cls:'neutral',note:'Total revenue for period'},
    {label:'Net Income',val:fmt(d.net_income,'$'),cls:parseFloat(d.net_income)>=0?'positive':'negative',note:'After all expenses'},
    {label:'Gross Profit',val:fmt(d.gross_profit,'$'),cls:parseFloat(d.gross_profit)>=0?'positive':'negative',note:'Revenue minus COGS'},
    {label:'Gross Margin',val:fmt(d.gross_margin,'pct'),cls:scoreClass(parseFloat(d.gross_margin)*100,30,15),note:'Gross Profit / Revenue'},
    {label:'Net Margin',val:fmt(d.net_margin,'pct'),cls:scoreClass(parseFloat(d.net_margin)*100,10,3),note:'Net Income / Revenue'},
    {label:'Total Assets',val:fmt(d.total_assets,'$'),cls:'neutral',note:'Balance sheet total'},
    {label:'Total Liabilities',val:fmt(d.total_liabilities,'$'),cls:'neutral',note:'All obligations'},
    {label:'Total Equity',val:fmt(d.total_equity,'$'),cls:parseFloat(d.total_equity)>=0?'positive':'negative',note:'Net assets'},
  ];
  if(d.ebitda!==null&&d.ebitda!==undefined)
    kpis.splice(3,0,{label:'EBITDA',val:fmt(d.ebitda,'$'),cls:parseFloat(d.ebitda)>=0?'positive':'negative',note:'Earnings before I/T/D/A'});
  kpis.forEach(k=>{
    const card=document.createElement('div');card.className='kpi-card';
    card.innerHTML=`<div class="kpi-label">${k.label}</div><div class="kpi-value ${k.cls}">${k.val}</div><div class="kpi-note">${k.note}</div>`;
    grid.appendChild(card);
  });
  g.appendChild(grid);
}

function renderRatios(d){
  const c = document.getElementById('ratios-container');
  c.innerHTML = '';
  const sections = [
    {title:'Liquidity Ratios', rows:[
      ['Current Ratio', fmt(d.current_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
      ['Quick Ratio', fmt(d.quick_ratio,'x'), 'x', '(Cash + AR) / Current Liabilities', 1.0, 0.8],
      ['Cash Ratio', fmt(d.cash_ratio,'x'), 'x', 'Cash / Current Liabilities', 0.5, 0.2],
      ['Working Capital', fmt(d.working_capital,'$'), '$', 'Current Assets − Current Liabilities', 1, 0],
      ['Working Capital Ratio', fmt(d.working_capital_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
    ]},
    {title:'Profitability Ratios', rows:[
      ['Gross Profit Margin', fmt(d.gross_margin,'pct'), 'pct', 'Gross Profit / Revenue', 0.35, 0.15],
      ['Operating Profit Margin', fmt(d.operating_margin,'pct'), 'pct', 'Operating Income / Revenue', 0.15, 0.05],
      ['Net Profit Margin', fmt(d.net_margin,'pct'), 'pct', 'Net Income / Revenue', 0.10, 0.03],
      ['EBITDA Margin', fmt(d.ebitda_margin,'pct'), 'pct', 'EBITDA / Revenue', 0.20, 0.08],
      ['Return on Assets (ROA)', fmt(d.roa,'pct'), 'pct', 'Net Income / Total Assets', 0.05, 0.02],
      ['Return on Equity (ROE)', fmt(d.roe,'pct'), 'pct', 'Net Income / Total Equity', 0.15, 0.08],
      ['Return on Invested Capital (ROIC)', fmt(d.roic,'pct'), 'pct', 'NOPAT / Invested Capital', 0.12, 0.06],
    ]},
    {title:'Efficiency Ratios', rows:[
      ['Asset Turnover Ratio', fmt(d.asset_turnover,'x'), 'x', 'Revenue / Total Assets', 1.0, 0.5],
      ['Inventory Turnover Ratio', fmt(d.inventory_turnover,'x'), 'x', 'COGS / Average Inventory', 6.0, 3.0],
      ['Accounts Receivable Turnover', fmt(d.ar_turnover,'x'), 'x', 'Revenue / Accounts Receivable', 8.0, 4.0],
      ['Accounts Payable Turnover', fmt(d.ap_turnover,'x'), 'x', 'COGS / Accounts Payable', 8.0, 4.0],
      ['Working Capital Turnover', fmt(d.wc_turnover,'x'), 'x', 'Revenue / Working Capital', 4.0, 2.0],
      ['Fixed Asset Turnover', fmt(d.fixed_asset_turnover,'x'), 'x', 'Revenue / Net Fixed Assets', 3.0, 1.5],
    ]},
    {title:'Leverage / Solvency Ratios', rows:[
      ['Debt-to-Equity Ratio', fmt(d.debt_to_equity,'x'), 'x', 'Total Liabilities / Total Equity', null, null, true],
      ['Debt Ratio', fmt(d.debt_ratio,'x'), 'x', 'Total Liabilities / Total Assets', null, null, true],
      ['Interest Coverage Ratio', fmt(d.interest_coverage,'x'), 'x', 'EBIT / Interest Expense', 3.0, 1.5],
      ['Debt Service Coverage (DSCR)', fmt(d.dscr,'x'), 'x', 'Operating CF / Debt Service', 1.25, 1.0],
      ['Equity Ratio', fmt(d.equity_ratio,'pct'), 'pct', 'Total Equity / Total Assets', 0.50, 0.30],
    ]},
    {title:'Cash Flow Ratios', rows:[
      ['Operating Cash Flow Ratio', fmt(d.ocf_ratio,'x'), 'x', 'Operating CF / Current Liabilities', 1.0, 0.5],
      ['Cash Flow Coverage Ratio', fmt(d.cf_coverage,'x'), 'x', 'Operating CF / Total Liabilities', 0.3, 0.15],
      ['Free Cash Flow Ratio', fmt(d.fcf_ratio,'x'), 'x', 'FCF / Revenue', 0.10, 0.03],
      ['Cash Conversion Ratio', fmt(d.cash_conversion,'x'), 'x', 'Cash from Ops / Net Income', 1.0, 0.7],
    ]},
    {title:'Growth Ratios (Year-over-Year)', rows:[
      ['Revenue Growth Rate', fmt(d.rev_growth,'pct'), 'pct', 'YoY Revenue Change', 0.10, 0.0],
      ['Gross Profit Growth Rate', fmt(d.gp_growth,'pct'), 'pct', 'YoY Gross Profit Change', 0.10, 0.0],
      ['EBITDA Growth Rate', fmt(d.ebitda_growth,'pct'), 'pct', 'YoY EBITDA Change', 0.10, 0.0],
      ['Net Income Growth Rate', fmt(d.ni_growth,'pct'), 'pct', 'YoY Net Income Change', 0.10, 0.0],
      ['Cash Flow Growth Rate', fmt(d.cf_growth,'pct'), 'pct', 'YoY Cash Flow Change', 0.10, 0.0],
    ]},
    {title:'Valuation Ratios (Market-Based — Requires Share Price)', rows:[
      ['Price-to-Earnings (P/E)', d.pe_ratio!==null?fmt(d.pe_ratio,'x'):'— (needs share price)', 'x', 'Share Price / EPS', null, null],
      ['Price-to-Sales (P/S)', d.ps_ratio!==null?fmt(d.ps_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Revenue', null, null],
      ['EV/EBITDA', d.ev_ebitda!==null?fmt(d.ev_ebitda,'x'):'— (needs share price)', 'x', 'Enterprise Value / EBITDA', null, null],
      ['Price-to-Book (P/B)', d.pb_ratio!==null?fmt(d.pb_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Book Value', null, null],
    ]},
    {title:'SaaS / Recurring Revenue Metrics', rows:[
      ['Monthly Recurring Revenue (MRR)', d.mrr!==null?fmt(d.mrr,'$'):'— (not applicable)', '$', 'Estimated from revenue pattern', null, null],
      ['Annual Recurring Revenue (ARR)', d.arr!==null?fmt(d.arr,'$'):'— (not applicable)', '$', 'MRR × 12', null, null],
      ['Customer Acquisition Cost (CAC)', '— (requires customer data)', null, 'Total Sales & Mktg / New Customers', null, null],
      ['Customer Lifetime Value (LTV)', '— (requires customer data)', null, 'Avg Revenue per User × Lifetime', null, null],
      ['LTV/CAC Ratio', '— (requires customer data)', null, 'Target ≥ 3x', null, null],
      ['Churn Rate', '— (requires customer data)', null, 'Customers Lost / Total Customers', null, null],
      ['Net Revenue Retention (NRR)', '— (requires customer data)', null, 'Expansion / Total Revenue', null, null],
    ]},
  ];
  sections.forEach(sec=>{
    const title = document.createElement('div'); title.className='ratio-section-title'; title.textContent=sec.title;
    c.appendChild(title);
    const table = document.createElement('table'); table.className='ratio-table';
    table.innerHTML=`<thead><tr><th>Ratio</th><th>Description</th><th>Value</th></tr></thead><tbody></tbody>`;
    const tbody = table.querySelector('tbody');
    sec.rows.forEach(([name, val, type, desc, good, warn, invertScore])=>{
      const tr=document.createElement('tr');
      let cls='neutral';
      if(type&&val&&val!=='N/A'&&!val.includes('—')&&!val.includes('requires')){
        const raw=parseFloat(val.replace(/[%x$,days\s]/g,''));
        if(!isNaN(raw)&&good!==null){
          if(invertScore){cls=raw<2?'good':raw<4?'warn':'bad';}
          else{cls=raw>=good?'good':raw>=warn?'warn':'bad';}
        }
      }
      tr.innerHTML=`<td>${name}</td><td style="font-size:.78rem;color:var(--text-muted);font-family:'DM Mono',monospace">${desc}</td><td class="ratio-val ${cls}">${val}</td>`;
      tbody.appendChild(tr);
    });
    c.appendChild(table);
  });
}

function renderInsights(insights){
  const c = document.getElementById('insights-container');
  c.innerHTML = '';
  if(typeof insights === 'string'){
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>AI Financial Commentary</h4><p>${insights.replace(/\n/g,'<br>')}</p>`;
    c.appendChild(block);
    return;
  }
  const sections = [
    {key:'financial_analysis', title:'📈 Financial Analysis'},
    {key:'management_insights', title:'💼 Management Insights & Commentary'},
    {key:'risks_and_opportunities', title:'⚠️ Risks & Opportunities'},
  ];
  sections.forEach(({key, title})=>{
    if(!insights[key]) return;
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>${title}</h4>`;
    const content = insights[key];
    if(Array.isArray(content)){
      const ul=document.createElement('ul');
      content.forEach(item=>{const li=document.createElement('li');li.textContent=item;ul.appendChild(li);});
      block.appendChild(ul);
    } else {
      const p=document.createElement('p');p.textContent=content;block.appendChild(p);
    }
    c.appendChild(block);
  });
  if(!c.children.length){
    c.innerHTML='<p style="color:var(--text-muted);font-size:.85rem">No insights available for this statement type.</p>';
  }
}

async function loadAnalytics(data){
  currentAnalyticsData = data;
  chatHistory = [];
  document.getElementById('analytics-company-title').textContent = data.company + ' — Financial Analysis';
  document.getElementById('analytics-period-sub').textContent = data.period + ' · ' + data.statement_label;
  document.getElementById('analytics-area').style.display = 'block';
  document.getElementById('chatMessages').innerHTML = `<div class="chat-empty">
    <div class="chat-empty-icon">💬</div>
    <p>Ask anything about <strong>${data.company}</strong>'s financials.</p>
    <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is the biggest expense?" · "Is liquidity a concern?" · "How is profitability trending?"</p>
  </div>`;
  document.getElementById('insights-container').innerHTML='<div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem"><span class="spinner"></span> Generating AI insights...</div>';
  if(data.analytics){
    renderKPIs(data.analytics);
    renderRatios(data.analytics);
  }
  try{
    const iResp = await fetch('/api/insights', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({analytics: data.analytics, company: data.company, period: data.period, statement_type: data.statement_type})
    });
    const iData = await iResp.json();
    if(iData.insights) renderInsights(iData.insights);
    else document.getElementById('insights-container').innerHTML='<p style="color:var(--text-muted);font-size:.85rem">Insights unavailable.</p>';
  } catch(e){
    document.getElementById('insights-container').innerHTML='<p style="color:var(--red);font-size:.85rem">Could not generate insights — ensure Ollama is running.</p>';
  }
}

async function sendChat(){
  const input = document.getElementById('chatInput');
  const msg = input.value.trim();
  if(!msg || !currentAnalyticsData) return;
  input.value = '';
  const msgs = document.getElementById('chatMessages');
  const firstChild = msgs.querySelector('.chat-empty');
  if(firstChild) msgs.innerHTML='';
  const userBubble = document.createElement('div');
  userBubble.className='chat-msg user'; userBubble.textContent=msg;
  msgs.appendChild(userBubble);
  const typingBubble = document.createElement('div');
  typingBubble.className='chat-msg assistant typing';
  typingBubble.innerHTML='<span class="typing-dot"></span><span class="typing-dot"></span><span class="typing-dot"></span>';
  msgs.appendChild(typingBubble);
  msgs.scrollTop=msgs.scrollHeight;
  document.getElementById('chatSendBtn').disabled=true;
  chatHistory.push({role:'user', content:msg});
  try{
    const resp = await fetch('/api/chat', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({message:msg, history:chatHistory, analytics:currentAnalyticsData.analytics, company:currentAnalyticsData.company, period:currentAnalyticsData.period})
    });
    const data = await resp.json();
    const reply = data.reply || 'I was unable to generate a response. Please try again.';
    chatHistory.push({role:'assistant',content:reply});
    typingBubble.remove();
    const asBubble=document.createElement('div');
    asBubble.className='chat-msg assistant';
    asBubble.innerHTML=`<div class="msg-label">Quanto AI</div>${reply.replace(/\n/g,'<br>')}`;
    msgs.appendChild(asBubble);
    msgs.scrollTop=msgs.scrollHeight;
  }catch(e){
    typingBubble.remove();
    const errBubble=document.createElement('div');
    errBubble.className='chat-msg assistant';
    errBubble.innerHTML='<div class="msg-label">Quanto AI</div>Error connecting to AI engine. Ensure Ollama is running.';
    msgs.appendChild(errBubble);
  }
  document.getElementById('chatSendBtn').disabled=false;
  input.focus();
}

async function generate(){
  if(!selectedStmt) return;
  const btn = document.getElementById('generateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('status-area');
  statusArea.style.display='block';
  statusArea.innerHTML=`<div class="progress-card">
    <div class="step-item active"><span class="spinner"></span>Extracting document text with AI Vision OCR...</div>
    <div class="step-item active"><span class="spinner"></span>Parsing financial accounts & ledger...</div>
    <div class="step-item active"><span class="spinner"></span>Generating professional Excel workbook...</div>
    <div class="step-item active"><span class="spinner"></span>Computing financial analytics...</div>
  </div>`;
  document.getElementById('result-area').innerHTML='';
  document.getElementById('analytics-area').style.display='none';
  const fd = new FormData();
  fd.append('statement_type', selectedStmt);
  if(TB_ONLY.has(selectedStmt)){
    fd.append('files', tbFile, tbFile.name);
  } else {
    const info = allStatements[selectedStmt]||{};
    (info.sources||[]).forEach(src=>{
      if(multiSourceFiles[src]) fd.append('files', multiSourceFiles[src], `${src}::${multiSourceFiles[src].name}`);
    });
  }
  try{
    const resp = await fetch('/api/generate', {method:'POST', body:fd});
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`
      <div class="result-card">
        <div style="font-size:2.2rem">✅</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.statement_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">${data.accounts_found} accounts extracted · OCR: ${data.ocr_provider}</p>
        <a href="/api/download/${data.filename}" class="download-btn">⬇ Download Excel</a>
      </div>`;
    if(data.analytics) await loadAnalytics(data);
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}

async function generateForecast(){
  if(!selectedFStmt||forecastFiles.length<3) return;
  const btn = document.getElementById('fgenerateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('fstatus-area');
  statusArea.style.display='block';
  const phases=['Phase 1-2: Validating and quality-checking documents…','Phase 3: Normalizing account names across years…','Phase 4: Computing historical ratios…','Phase 5-6: Identifying drivers and running forecast methods…','Phase 7: Building three-statement linked model…','Phase 8: Projecting cash flows and runway…','Phase 9: Running scenario analysis (Base/Best/Worst)…','Phase 10: Generating stakeholder analysis…','Phase 11: Running DCF valuation model…','Phase 12: Scoring risk across 6 dimensions…','Phase 13: Generating AI narrative via Llama 3.1…','Phase 14: Computing forecast confidence scores…','Phase 15: Building 13-tab Excel workbook…'];
  statusArea.innerHTML='<div class="progress-card">'+phases.map(p=>`<div class="step-item active"><span class="spinner"></span>${p}</div>`).join('')+'</div>';
  document.getElementById('fresult-area').innerHTML='';
  const fd=new FormData();
  fd.append('forecast_type', selectedFStmt);
  forecastFiles.forEach(f=>fd.append('files', f, f.name));
  try{
    const resp=await fetch('/api/forecast',{method:'POST',body:fd});
    const data=await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    const cs=data.confidence_scores||{};
    document.getElementById('fresult-area').innerHTML=`
      <div class="result-card forecast-result">
        <div style="font-size:2.2rem">🚀</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.forecast_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">
          ${data.years_analyzed} years analyzed · ${data.phases_run} phases complete ·
          Revenue confidence: <strong>${cs.revenue}</strong> · Margins: <strong>${cs.margins}</strong>
        </p>
        ${(data.validation_warnings||[]).map(w=>`<p style="color:var(--orange);font-size:.78rem;margin-top:.35rem">⚠ ${w}</p>`).join('')}
        <a href="/api/download/${data.filename}" class="download-btn" style="background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400">⬇ Download Forecast Workbook</a>
      </div>`;
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('fresult-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}

// ---- Interactive Ledger tab ----
const ledgerFileInput = document.getElementById('ledgerFileInput');
ledgerFileInput.addEventListener('change', e=>{ if(e.target.files[0]) initLedgerSession(e.target.files[0]); });

function ledgerAppendBubble(role, text){
  const msgs = document.getElementById('ledgerMessages');
  const empty = msgs.querySelector('.chat-empty');
  if(empty) msgs.innerHTML = '';
  const bubble = document.createElement('div');
  bubble.className = 'chat-msg ' + role;
  if(role==='assistant'){
    const label = document.createElement('div');
    label.className = 'msg-label'; label.textContent = 'Quanto Ledger';
    const block = document.createElement('div');
    block.className = 'ledger-state-block'; block.textContent = text;
    bubble.appendChild(label); bubble.appendChild(block);
  } else {
    bubble.textContent = text;
  }
  msgs.appendChild(bubble);
  msgs.scrollTop = msgs.scrollHeight;
  return bubble;
}

async function initLedgerSession(file){
  document.getElementById('ledgerFileStatus').textContent = 'Uploading & extracting: ' + file.name + '...';
  const fd = new FormData();
  fd.append('file', file, file.name);
  try{
    const resp = await fetch('/api/ledger/init', {method:'POST', body:fd});
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    ledgerSessionId = data.session_id;
    document.getElementById('ledgerFileStatus').textContent = '✓ Loaded: ' + file.name;
    document.getElementById('ledgerSessionBadge').style.display = 'inline-block';
    document.getElementById('ledgerInput').disabled = false;
    document.getElementById('ledgerSendBtn').disabled = false;
    ledgerAppendBubble('assistant', data.formatted_state);
  }catch(err){
    document.getElementById('ledgerFileStatus').textContent = '❌ ' + err.message;
  }
}

async function sendLedgerCommand(){
  const input = document.getElementById('ledgerInput');
  const msg = input.value.trim();
  if(!msg || !ledgerSessionId) return;
  input.value = '';
  ledgerAppendBubble('user', msg);
  document.getElementById('ledgerSendBtn').disabled = true;
  try{
    const resp = await fetch('/api/ledger/command', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({session_id: ledgerSessionId, command: msg})
    });
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    ledgerAppendBubble('assistant', data.formatted_state);
  }catch(err){
    ledgerAppendBubble('assistant', '⚠ ' + err.message);
  }
  document.getElementById('ledgerSendBtn').disabled = false;
  input.focus();
}
</script>
</body>
</html>"""


# =====================================================================================
# SECTION 1: OCR / DOCUMENT EXTRACTION LAYER
# =====================================================================================

def _pdf_to_images_base64(file_bytes: bytes) -> List[str]:
    """Render PDF pages to base64 PNGs for vision-model OCR. Returns [] if pdf2image unavailable."""
    try:
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(file_bytes, dpi=300)
        result = []
        for img in images[:8]:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(base64.b64encode(buf.getvalue()).decode())
        return result
    except Exception:
        return []


def _image_to_base64(file_bytes: bytes, filename: str) -> str:
    return base64.b64encode(file_bytes).decode()


def _guess_mime(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {
        ".pdf": "application/pdf", ".png": "image/png",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".tif": "image/tiff", ".tiff": "image/tiff", ".bmp": "image/bmp",
    }.get(ext, "application/octet-stream")


OCR_EXTRACTION_PROMPT = """You are a meticulous financial document OCR and extraction engine.
Extract EVERY line item / account you can find in this financial document image(s).

Return ONLY valid JSON (no markdown fences, no commentary) in this exact shape:
{
  "company_name": "string or null",
  "period_label": "string describing the fiscal period, e.g. 'FY2024' or 'Year Ended Dec 31, 2024', or null",
  "currency": "3-letter currency code guess, default USD",
  "accounts": [
    {"account_name": "string", "category": "asset|liability|equity|revenue|cogs|expense|other", "amount": number, "subcategory": "current_asset|non_current_asset|current_liability|non_current_liability|other|null"}
  ],
  "notes": "any caveats about illegible or ambiguous figures, or empty string"
}

Rules:
- Numbers must be plain numbers (no $ signs, no commas, no parentheses — convert (1,000) to -1000).
- Include subtotal/total lines too if present, but prioritize line-level detail.
- If a value is illegible, omit that line and mention it in notes.
- Category must be one of: asset, liability, equity, revenue, cogs, expense, other.
- Do not invent figures that are not visibly present in the document.
"""


def _extract_json_from_text(text: str) -> dict:
    """Best-effort extraction of a JSON object from a model response that may include stray text/fences."""
    text = text.strip()
    text = re.sub(r"^```(?:json)?", "", text).strip()

    text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError("No JSON object found in OCR response")
    return json.loads(match.group(0))


def _ocr_with_claude(file_bytes: bytes, filename: str) -> dict:
    """Uses Anthropic's Claude vision model for OCR/extraction. Claude accepts PDFs natively
    as 'document' content blocks (no pdf2image/poppler conversion needed), which makes this
    provider simpler to set up than the Gemini/OpenAI paths while remaining highly accurate
    on financial documents."""
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    mime = _guess_mime(filename)
    b64 = base64.b64encode(file_bytes).decode()

    if mime == "application/pdf":
        content_block: Dict[str, Any] = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    else:
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": mime, "data": b64},
        }

    message = client.messages.create(
        model="claude-sonnet-5",
        max_tokens=4096,
        temperature=0,
        messages=[{
            "role": "user",
            "content": [content_block, {"type": "text", "text": OCR_EXTRACTION_PROMPT}],
        }],
    )
    text = "".join(block.text for block in message.content if getattr(block, "type", None) == "text")
    return _extract_json_from_text(text)


def _ocr_with_gemini(file_bytes: bytes, filename: str) -> dict:
    from google import genai
    from google.genai import types

    client = genai.Client(api_key=GEMINI_API_KEY)

    mime = _guess_mime(filename)

    response = client.models.generate_content(
        model="gemini-3.5-flash",
        contents=[
            OCR_EXTRACTION_PROMPT,
            types.Part.from_bytes(
                data=file_bytes,
                mime_type=mime,
            ),
        ],
    )

    if not response.text:
        print("Gemini Response:")
        print(response.text)
        raise RuntimeError("Gemini returned an empty response.")


    return _extract_json_from_text(response.text)



def _ocr_with_openai(file_bytes: bytes, filename: str) -> dict:
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    mime = _guess_mime(filename)
    content: List[Dict[str, Any]] = [{"type": "text", "text": OCR_EXTRACTION_PROMPT}]

    if mime == "application/pdf":
        images_b64 = _pdf_to_images_base64(file_bytes)
        if not images_b64:
            raise RuntimeError("PDF page rendering unavailable (install pdf2image + poppler) for OpenAI vision path")
        for img_b64 in images_b64:
            content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}})
    else:
        b64 = _image_to_base64(file_bytes, filename)
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})

    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": content}],
        temperature=0.0,
        max_tokens=4096,
    )
    return _extract_json_from_text(resp.choices[0].message.content)


def _ocr_with_local(file_bytes: bytes, filename: str) -> dict:
    """Local fallback: pdfplumber for PDFs, pytesseract for images, then regex-based line parsing."""
    raw_text = ""
    mime = _guess_mime(filename)
    if mime == "application/pdf":
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                raw_text += (page.extract_text() or "") + "\n"
        if not raw_text.strip():
            images_b64 = _pdf_to_images_base64(file_bytes)
            if images_b64:
                import pytesseract
                from PIL import Image as PILImage
                for img_b64 in images_b64:
                    img = PILImage.open(io.BytesIO(base64.b64decode(img_b64)))
                    raw_text += pytesseract.image_to_string(img) + "\n"
    else:
        import pytesseract
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(file_bytes))
        raw_text = pytesseract.image_to_string(img)

    return _parse_raw_text_to_accounts(raw_text)


LINE_ITEM_RE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z0-9&,\.\'/\-\s]{2,80}?)\s{1,}\$?\(?(?P<amount>-?[\d,]+(?:\.\d{1,2})?)\)?\s*$"
)

ASSET_HINTS = ["cash", "receivable", "inventory", "prepaid", "asset", "equipment", "property", "investment", "goodwill", "intangible"]
LIABILITY_HINTS = ["payable", "liability", "liabilities", "loan", "debt", "accrued", "deferred revenue", "unearned", "note payable", "lease liability"]
EQUITY_HINTS = ["equity", "retained earnings", "common stock", "paid-in capital", "treasury stock", "owner's capital", "shareholders"]
REVENUE_HINTS = ["revenue", "sales", "income from", "service income", "fees earned"]
COGS_HINTS = ["cost of goods", "cogs", "cost of sales", "cost of revenue"]
EXPENSE_HINTS = ["expense", "salaries", "wages", "rent", "utilities", "depreciation", "amortization", "interest expense", "tax expense", "advertising", "insurance", "supplies"]


def _classify_account(name: str) -> str:
    n = name.lower()
    if any(h in n for h in COGS_HINTS): return "cogs"
    if any(h in n for h in REVENUE_HINTS): return "revenue"
    if any(h in n for h in LIABILITY_HINTS): return "liability"
    if any(h in n for h in EQUITY_HINTS): return "equity"
    if any(h in n for h in EXPENSE_HINTS): return "expense"
    if any(h in n for h in ASSET_HINTS): return "asset"
    return "other"


def _parse_raw_text_to_accounts(raw_text: str) -> dict:
    accounts = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line or len(line) < 4:
            continue
        m = LINE_ITEM_RE.match(line)
        if not m:
            continue
        name = re.sub(r"\s{2,}", " ", m.group("name")).strip(" .:-")
        amount_str = m.group("amount").replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        if "(" in line and ")" in line and amount > 0:
            amount = -amount
        if not name or name.lower() in {"total", "subtotal"}:
            continue
        accounts.append({
            "account_name": name,
            "category": _classify_account(name),
            "amount": amount,
            "subcategory": None,
        })
    period_match = re.search(r"(?:year ended|period ended|as of|fiscal year)\s+([A-Za-z0-9,\s]+\d{4})", raw_text, re.IGNORECASE)
    return {
        "company_name": None,
        "period_label": period_match.group(1).strip() if period_match else None,
        "currency": "USD",
        "accounts": accounts,
        "notes": "Extracted via local OCR (pdfplumber/pytesseract) — lower fidelity than AI Vision OCR. Review figures carefully." if accounts else "No line items could be confidently parsed from this document.",
    }


def extract_document(file_bytes: bytes, filename: str) -> dict:
    """
    OCR dispatcher.
    Tries the selected provider first, then falls back to the others.
    """

    available = []

    if ANTHROPIC_API_KEY:
        available.append(("claude", _ocr_with_claude))

    if GEMINI_API_KEY:
        available.append(("gemini", _ocr_with_gemini))

    if OPENAI_API_KEY:
        available.append(("openai", _ocr_with_openai))

    # Put the configured provider first
    provider_chain = sorted(
        available,
        key=lambda x: 0 if x[0] == OCR_PROVIDER else 1
    )

    # Local OCR is always the final fallback
    provider_chain.append(("local", _ocr_with_local))

    last_error = None

    for provider_name, fn in provider_chain:

        print(f"\n===== Trying OCR Provider: {provider_name} =====")

        try:
            result = fn(file_bytes, filename)
            result["_ocr_provider_used"] = provider_name

            print(f"SUCCESS using {provider_name}")

            return result

        except Exception as e:

            print("=" * 60)
            print(f"OCR Provider Failed: {provider_name}")
            print(f"Error Type: {type(e).__name__}")
            print(f"Error: {e}")
            print("=" * 60)

            last_error = e

    raise HTTPException(
        status_code=422,
        detail=f"All OCR providers failed to extract this document. Last error: {last_error}"
    )


def get_ocr_status() -> dict:
    labels = {
        "claude": "Claude Sonnet 5 Vision",
        "gemini": "Gemini 1.5 Pro Vision",
        "openai": "OpenAI GPT-4o Vision",
    }
    keys = {"claude": ANTHROPIC_API_KEY, "gemini": GEMINI_API_KEY, "openai": OPENAI_API_KEY}
    if OCR_PROVIDER in labels and keys[OCR_PROVIDER]:
        return {"provider": OCR_PROVIDER, "label": labels[OCR_PROVIDER], "note": "AI Vision OCR active"}
    for provider, key in keys.items():
        if key:
            return {"provider": provider, "label": labels[provider], "note": "AI Vision OCR active (auto-selected)"}
    return {"provider": "local", "label": "Local OCR (pdfplumber/Tesseract)", "note": "No API key set — using local fallback"}


# =====================================================================================
# SECTION 2: ACCOUNT NORMALIZATION & FINANCIAL CALCULATIONS
# =====================================================================================

def _sum_by_category(accounts: List[Dict[str, Any]], category: str) -> float:
    return sum(a["amount"] for a in accounts if a.get("category") == category)


def _find_account(accounts: List[Dict[str, Any]], *keywords: str) -> float:
    """Sum amounts of accounts whose name contains any of the given (lowercase) keywords."""
    total = 0.0
    for a in accounts:
        name = a.get("account_name", "").lower()
        if any(kw in name for kw in keywords):
            total += a.get("amount", 0.0)
    return total


def compute_core_financials(accounts: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Derive core statement totals (Revenue, COGS, Expenses, Assets, Liabilities, Equity)
    from a flat list of classified accounts. This is the foundation all ratio/KPI math builds on."""
    revenue = _sum_by_category(accounts, "revenue")
    cogs = abs(_sum_by_category(accounts, "cogs"))
    opex = abs(_sum_by_category(accounts, "expense"))
    total_assets = _sum_by_category(accounts, "asset")
    total_liabilities = abs(_sum_by_category(accounts, "liability"))
    total_equity = _sum_by_category(accounts, "equity")

    # If equity wasn't classified/extracted but balance sheet should balance, derive it.
    if total_equity == 0 and total_assets > 0 and total_liabilities > 0:
        total_equity = total_assets - total_liabilities

    gross_profit = revenue - cogs
    depreciation_amort = _find_account(accounts, "depreciation", "amortization")
    interest_expense = abs(_find_account(accounts, "interest expense"))
    tax_expense = abs(_find_account(accounts, "tax expense", "income tax"))

    operating_income = gross_profit - opex
    ebit = operating_income
    ebitda = ebit + abs(depreciation_amort)
    net_income = operating_income - interest_expense - tax_expense

    cash = _find_account(accounts, "cash", "bank")
    accounts_receivable = _find_account(accounts, "accounts receivable", "receivable")
    inventory = _find_account(accounts, "inventory")
    accounts_payable = abs(_find_account(accounts, "accounts payable", "payable"))

    current_assets = cash + accounts_receivable + inventory
    if current_assets == 0:
        current_assets = total_assets * 0.5  # heuristic fallback when subcategory tagging absent
    current_liabilities = accounts_payable
    if current_liabilities == 0:
        current_liabilities = total_liabilities * 0.5

    return {
        "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
        "operating_expenses": opex, "operating_income": operating_income,
        "ebit": ebit, "ebitda": ebitda, "net_income": net_income,
        "interest_expense": interest_expense, "tax_expense": tax_expense,
        "depreciation_amortization": abs(depreciation_amort),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,
        "cash": cash, "accounts_receivable": accounts_receivable, "inventory": inventory,
        "accounts_payable": accounts_payable,
        "current_assets": current_assets, "current_liabilities": current_liabilities,
    }


def _safe_div(numerator: float, denominator: float) -> Optional[float]:
    if denominator == 0 or denominator is None:
        return None
    return numerator / denominator


def compute_full_analytics(accounts: List[Dict[str, Any]], prior_accounts: Optional[List[Dict[str, Any]]] = None,
                            share_price: Optional[float] = None) -> Dict[str, Any]:
    """Computes the full ratio suite shown in the frontend's Overview + Ratio Analysis tabs."""
    f = compute_core_financials(accounts)

    revenue, cogs, gross_profit = f["revenue"], f["cogs"], f["gross_profit"]
    opex, op_income, ebit, ebitda = f["operating_expenses"], f["operating_income"], f["ebit"], f["ebitda"]
    net_income = f["net_income"]
    total_assets, total_liabilities, total_equity = f["total_assets"], f["total_liabilities"], f["total_equity"]
    cash, ar, inv, ap = f["cash"], f["accounts_receivable"], f["inventory"], f["accounts_payable"]
    current_assets, current_liabilities = f["current_assets"], f["current_liabilities"]
    interest_expense = f["interest_expense"]

    quick_assets = cash + ar
    working_capital = current_assets - current_liabilities

    # Approximate operating cash flow (no full CF statement available from a single TB).
    operating_cf = net_income + f["depreciation_amortization"]

    analytics: Dict[str, Any] = {
        "revenue": revenue, "net_income": net_income, "gross_profit": gross_profit,
        "gross_margin": _safe_div(gross_profit, revenue),
        "net_margin": _safe_div(net_income, revenue),
        "operating_margin": _safe_div(op_income, revenue),
        "ebitda": ebitda, "ebitda_margin": _safe_div(ebitda, revenue),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,

        # Liquidity
        "current_ratio": _safe_div(current_assets, current_liabilities),
        "quick_ratio": _safe_div(quick_assets, current_liabilities),
        "cash_ratio": _safe_div(cash, current_liabilities),
        "working_capital": working_capital,
        "working_capital_ratio": _safe_div(current_assets, current_liabilities),

        # Profitability
        "roa": _safe_div(net_income, total_assets),
        "roe": _safe_div(net_income, total_equity),
        "roic": _safe_div(ebit * 0.79, (total_equity + total_liabilities)) if (total_equity + total_liabilities) else None,

        # Efficiency
        "asset_turnover": _safe_div(revenue, total_assets),
        "inventory_turnover": _safe_div(cogs, inv) if inv else None,
        "ar_turnover": _safe_div(revenue, ar) if ar else None,
        "ap_turnover": _safe_div(cogs, ap) if ap else None,
        "wc_turnover": _safe_div(revenue, working_capital) if working_capital else None,
        "fixed_asset_turnover": _safe_div(revenue, total_assets - current_assets) if (total_assets - current_assets) > 0 else None,

        # Leverage / Solvency
        "debt_to_equity": _safe_div(total_liabilities, total_equity),
        "debt_ratio": _safe_div(total_liabilities, total_assets),
        "interest_coverage": _safe_div(ebit, interest_expense) if interest_expense else None,
        "dscr": _safe_div(operating_cf, interest_expense) if interest_expense else None,
        "equity_ratio": _safe_div(total_equity, total_assets),

        # Cash flow (approximated)
        "ocf_ratio": _safe_div(operating_cf, current_liabilities),
        "cf_coverage": _safe_div(operating_cf, total_liabilities),
        "fcf_ratio": _safe_div(operating_cf, revenue),
        "cash_conversion": _safe_div(operating_cf, net_income) if net_income else None,

        # Growth (requires prior period)
        "rev_growth": None, "gp_growth": None, "ebitda_growth": None, "ni_growth": None, "cf_growth": None,

        # Valuation (requires share price / market data — not available from financial statements alone)
        "pe_ratio": None, "ps_ratio": None, "ev_ebitda": None, "pb_ratio": None,

        # SaaS metrics (not derivable from a generic TB without subscription data)
        "mrr": None, "arr": None,
    }

    if prior_accounts:
        pf = compute_core_financials(prior_accounts)
        analytics["rev_growth"] = _safe_div(revenue - pf["revenue"], pf["revenue"]) if pf["revenue"] else None
        analytics["gp_growth"] = _safe_div(gross_profit - pf["gross_profit"], pf["gross_profit"]) if pf["gross_profit"] else None
        prior_ebitda = pf["ebit"] + pf["depreciation_amortization"]
        analytics["ebitda_growth"] = _safe_div(ebitda - prior_ebitda, prior_ebitda) if prior_ebitda else None
        analytics["ni_growth"] = _safe_div(net_income - pf["net_income"], pf["net_income"]) if pf["net_income"] else None
        prior_ocf = pf["net_income"] + pf["depreciation_amortization"]
        analytics["cf_growth"] = _safe_div(operating_cf - prior_ocf, prior_ocf) if prior_ocf else None

    if share_price is not None and net_income and total_equity:
        shares_outstanding = None  # Not derivable without cap table data; left as None deliberately.

    return analytics


def generate_ai_insights(analytics: Dict[str, Any], company: str, period: str, statement_type: str) -> Any:
    """Calls the local Ollama Llama 3.1 model to produce narrative commentary on the analytics.
    Falls back to a deterministic rules-based summary if Ollama is unavailable."""
    try:
        import ollama
        prompt = f"""You are a senior financial analyst writing commentary for {company}'s {period} {STATEMENTS.get(statement_type, {}).get('label', 'financial statement')}.

Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None}, indent=2)}

Quanto is not responsible for financial decisions — keep that spirit in mind (informative, not advisory).

Respond ONLY with valid JSON in this exact shape, no markdown fences:
{{
  "financial_analysis": "2-4 sentences on overall financial position and performance",
  "management_insights": ["3-5 short bullet observations a CFO would care about"],
  "risks_and_opportunities": ["3-5 short bullet items, mixing risks and opportunities"]
}}"""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.3})
        content = resp["message"]["content"]
        return _extract_json_from_text(content)
    except Exception:
        return _fallback_insights(analytics)


def _fallback_insights(a: Dict[str, Any]) -> Dict[str, Any]:
    """Deterministic, rules-based insights used when Ollama is not running."""
    obs = []
    risks = []
    gm = a.get("gross_margin")
    nm = a.get("net_margin")
    cr = a.get("current_ratio")
    de = a.get("debt_to_equity")

    if gm is not None:
        obs.append(f"Gross margin stands at {gm*100:.1f}%, {'a healthy level' if gm > 0.3 else 'on the thinner side for most industries'}.")
    if nm is not None:
        obs.append(f"Net margin of {nm*100:.1f}% reflects {'solid' if nm > 0.1 else 'modest'} bottom-line conversion.")
    if cr is not None:
        if cr < 1.0:
            risks.append(f"Current ratio of {cr:.2f}x is below 1.0 — short-term obligations may exceed liquid assets.")
        else:
            obs.append(f"Current ratio of {cr:.2f}x indicates the company can cover short-term liabilities.")
    if de is not None and de > 2.0:
        risks.append(f"Debt-to-equity of {de:.2f}x signals elevated leverage relative to the equity base.")
    if not obs:
        obs.append("Key margin and liquidity figures were derived from the available data; review the ratio tables for full detail.")
    if not risks:
        risks.append("No major red flags surfaced from the ratios computed; continue monitoring trends period-over-period.")

    return {
        "financial_analysis": " ".join(obs[:2]) if obs else "Financial figures have been computed from the uploaded statement.",
        "management_insights": obs,
        "risks_and_opportunities": risks,
    }


def chat_about_financials(message: str, history: List[Dict[str, str]], analytics: Dict[str, Any],
                           company: str, period: str) -> str:
    """Powers the 'Ask Quanto' chat tab using local Ollama, with a deterministic fallback."""
    try:
        import ollama
        system_msg = f"""You are Quanto AI, a financial analyst assistant embedded in the Quanto platform.
You are discussing {company}'s financials for {period}.
Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None})}
Be concise, factual, and grounded only in the figures provided. Remind the user that Quanto is not responsible
for financial decisions if they ask for advice on what action to take. Do not invent figures not present above."""
        messages = [{"role": "system", "content": system_msg}]
        for h in history[-10:]:
            messages.append({"role": h.get("role", "user"), "content": h.get("content", "")})
        messages.append({"role": "user", "content": message})
        resp = ollama.chat(model=OLLAMA_MODEL, messages=messages, options={"temperature": 0.4})
        return resp["message"]["content"]
    except Exception:
        return ("I can't reach the local Ollama AI engine right now, so I can only point you to the numbers "
                "directly: check the Overview and Ratio Analysis tabs for the figures relevant to your question. "
                "Make sure Ollama is running (`ollama serve`) with the llama3.1 model pulled to enable full chat.")


# =====================================================================================
# SECTION 3: EXCEL WORKBOOK GENERATION (openpyxl, commercial-grade formatting)
# =====================================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XL_NAVY = "1F2D4A"
XL_ACCENT = "4F8EF7"
XL_LIGHT = "EAF1FE"
XL_GREEN = "2E8B57"
XL_RED = "C0392B"
XL_GREY = "7B8DB0"

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=False, color="FFFFFF", italic=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=XL_NAVY)
LABEL_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=XL_NAVY)
NUMBER_FONT = Font(name="Calibri", size=10, color="333333")
DISCLAIMER_FONT = Font(name="Calibri", size=8, italic=True, color=XL_GREY)

HEADER_FILL = PatternFill(start_color=XL_NAVY, end_color=XL_NAVY, fill_type="solid")
SECTION_FILL = PatternFill(start_color=XL_LIGHT, end_color=XL_LIGHT, fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E4F5", end_color="D9E4F5", fill_type="solid")

THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
TOTAL_BORDER = Border(top=Side(style="thin", color=XL_NAVY), bottom=Side(style="double", color=XL_NAVY))

CURRENCY_FMT = '#,##0;[Red](#,##0)'
PCT_FMT = '0.0%'
X_FMT = '0.00"x"'


def _new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_statement_header(ws: Worksheet, company: str, statement_label: str, period: str, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    c = ws.cell(row=1, column=start_col, value=company or "Company Name")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    c = ws.cell(row=2, column=start_col, value=statement_label)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
    c = ws.cell(row=3, column=start_col, value=period or "")
    c.font = SUBHEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 20 if r != 1 else 24

    return 5  # next free row


def _write_disclaimer_footer(ws: Worksheet, row: int, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value="Quanto is not responsible for financial decisions. Generated by Quanto Financial Intelligence Platform.")
    c.font = DISCLAIMER_FONT
    c.alignment = Alignment(horizontal="center")


def _autosize_columns(ws: Worksheet, widths: Dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_section_title(ws: Worksheet, row: int, title: str, span: int = 4, start_col: int = 1):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=title)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_line_item(ws: Worksheet, row: int, label: str, value: Optional[float], indent: int = 1,
                      bold: bool = False, currency: bool = True, label_col: int = 1, value_col: int = 4):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = BOLD_LABEL_FONT if bold else LABEL_FONT
    lc.alignment = Alignment(indent=indent)
    lc.border = THIN_BORDER
    for col in range(label_col + 1, value_col):
        ws.cell(row=row, column=col).border = THIN_BORDER
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT if bold else NUMBER_FONT
    if currency and isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    vc.border = THIN_BORDER
    return row + 1


def _write_total_row(ws: Worksheet, row: int, label: str, value: Optional[float], label_col: int = 1, value_col: int = 4):
    for col in range(label_col, value_col + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = TOTAL_BORDER
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = TOTAL_FONT
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT
    if isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    return row + 2


# ---- Individual statement-tab builders -------------------------------------------------

def _build_income_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Income Statement")
    row = _write_statement_header(ws, company, "Income Statement", period)
    row = _write_section_title(ws, row, "Revenue")
    for a in accounts:
        if a["category"] == "revenue":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Revenue", f["revenue"])

    row = _write_section_title(ws, row, "Cost of Goods Sold")
    for a in accounts:
        if a["category"] == "cogs":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total COGS", f["cogs"])
    row = _write_total_row(ws, row, "Gross Profit", f["gross_profit"])

    row = _write_section_title(ws, row, "Operating Expenses")
    for a in accounts:
        if a["category"] == "expense":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Operating Expenses", f["operating_expenses"])
    row = _write_total_row(ws, row, "Operating Income (EBIT)", f["ebit"])

    row = _write_line_item(ws, row, "Interest Expense", f["interest_expense"], bold=False)
    row = _write_line_item(ws, row, "Income Tax Expense", f["tax_expense"], bold=False)
    row = _write_total_row(ws, row, "NET INCOME", f["net_income"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_balance_sheet_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Balance Sheet")
    row = _write_statement_header(ws, company, "Balance Sheet", period)
    row = _write_section_title(ws, row, "Assets")
    for a in accounts:
        if a["category"] == "asset":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Assets", f["total_assets"])

    row = _write_section_title(ws, row, "Liabilities")
    for a in accounts:
        if a["category"] == "liability":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Liabilities", f["total_liabilities"])

    row = _write_section_title(ws, row, "Equity")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Equity", f["total_equity"])
    row = _write_total_row(ws, row, "Total Liabilities & Equity", f["total_liabilities"] + f["total_equity"])

    balance_check = abs(f["total_assets"] - (f["total_liabilities"] + f["total_equity"]))
    note_row = row
    c = ws.cell(row=note_row, column=1, value=("✓ Balance sheet balances." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f} — review extracted figures."))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, note_row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_retained_earnings_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Retained Earnings")
    row = _write_statement_header(ws, company, "Statement of Retained Earnings", period)
    opening_re = _find_account(accounts, "retained earnings") - f["net_income"]
    dividends = abs(_find_account(accounts, "dividend"))
    row = _write_line_item(ws, row, "Retained Earnings — Beginning of Period", opening_re)
    row = _write_line_item(ws, row, "Add: Net Income", f["net_income"])
    row = _write_line_item(ws, row, "Less: Dividends Declared", -dividends if dividends else 0)
    row = _write_total_row(ws, row, "Retained Earnings — End of Period", opening_re + f["net_income"] - dividends)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


def _build_equity_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Equity Statement")
    row = _write_statement_header(ws, company, "Statement of Shareholders' Equity", period)
    row = _write_section_title(ws, row, "Equity Components")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Shareholders' Equity", f["total_equity"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_trial_balance_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Trial Balance")
    row = _write_statement_header(ws, company, "Trial Balance", period, span=5)
    headers = ["Account", "Category", "Debit", "Credit"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    row += 1
    total_debit = total_credit = 0.0
    debit_categories = {"asset", "cogs", "expense"}
    for a in accounts:
        is_debit = a["category"] in debit_categories
        amt = abs(a["amount"])
        ws.cell(row=row, column=1, value=a["account_name"]).font = LABEL_FONT
        ws.cell(row=row, column=2, value=a["category"].title()).font = LABEL_FONT
        debit_cell = ws.cell(row=row, column=3, value=amt if is_debit else None)
        credit_cell = ws.cell(row=row, column=4, value=None if is_debit else amt)
        debit_cell.number_format = CURRENCY_FMT
        credit_cell.number_format = CURRENCY_FMT
        if is_debit:
            total_debit += amt
        else:
            total_credit += amt
        row += 1
    row = _write_total_row(ws, row, "TOTAL", None, value_col=4)
    ws.cell(row=row - 2, column=3, value=total_debit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=3).font = TOTAL_FONT
    ws.cell(row=row - 2, column=4, value=total_credit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=4).font = TOTAL_FONT
    balance_check = abs(total_debit - total_credit)
    c = ws.cell(row=row, column=1, value=("✓ Trial balance is in balance." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f}"))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, row + 1, span=4)
    _autosize_columns(ws, {1: 38, 2: 16, 3: 16, 4: 16})
    return ws


def _build_ratio_tab(wb: Workbook, analytics: Dict[str, Any], company, period, sheet_name="Ratio Analysis"):
    ws = wb.create_sheet(sheet_name)
    row = _write_statement_header(ws, company, sheet_name, period)
    sections = [
        ("Liquidity Ratios", [
            ("Current Ratio", analytics.get("current_ratio"), X_FMT),
            ("Quick Ratio", analytics.get("quick_ratio"), X_FMT),
            ("Cash Ratio", analytics.get("cash_ratio"), X_FMT),
            ("Working Capital", analytics.get("working_capital"), CURRENCY_FMT),
        ]),
        ("Profitability Ratios", [
            ("Gross Margin", analytics.get("gross_margin"), PCT_FMT),
            ("Operating Margin", analytics.get("operating_margin"), PCT_FMT),
            ("Net Margin", analytics.get("net_margin"), PCT_FMT),
            ("EBITDA Margin", analytics.get("ebitda_margin"), PCT_FMT),
            ("Return on Assets (ROA)", analytics.get("roa"), PCT_FMT),
            ("Return on Equity (ROE)", analytics.get("roe"), PCT_FMT),
        ]),
        ("Efficiency Ratios", [
            ("Asset Turnover", analytics.get("asset_turnover"), X_FMT),
            ("Inventory Turnover", analytics.get("inventory_turnover"), X_FMT),
            ("AR Turnover", analytics.get("ar_turnover"), X_FMT),
            ("AP Turnover", analytics.get("ap_turnover"), X_FMT),
        ]),
        ("Leverage / Solvency Ratios", [
            ("Debt-to-Equity", analytics.get("debt_to_equity"), X_FMT),
            ("Debt Ratio", analytics.get("debt_ratio"), X_FMT),
            ("Interest Coverage", analytics.get("interest_coverage"), X_FMT),
            ("Equity Ratio", analytics.get("equity_ratio"), PCT_FMT),
        ]),
    ]
    for title, rows_data in sections:
        row = _write_section_title(ws, row, title)
        for label, value, fmt in rows_data:
            lc = ws.cell(row=row, column=1, value=label); lc.font = LABEL_FONT; lc.alignment = Alignment(indent=1)
            vc = ws.cell(row=row, column=4, value=value if value is not None else "N/A")
            vc.font = NUMBER_FONT
            if isinstance(value, (int, float)):
                vc.number_format = fmt
            vc.alignment = Alignment(horizontal="right")
            row += 1
        row += 1
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 32, 2: 10, 3: 10, 4: 16})
    return ws


def _build_generic_schedule_tab(wb: Workbook, statement_key: str, statement_label: str,
                                 extracted_docs: Dict[str, dict], company: str, period: str):
    """Generic builder for the 25+ multi-source schedules (AR aging, fixed asset schedule, debt schedule, etc.)
    Lists all extracted line items per source document in clearly labeled sections, since each schedule type
    has bespoke business logic that depends on real-world source documents Quanto can refine further per type."""
    ws = wb.create_sheet(statement_label[:31])
    row = _write_statement_header(ws, company, statement_label, period)
    info = STATEMENTS.get(statement_key, {})
    for source_key in info.get("sources", []):
        doc = extracted_docs.get(source_key)
        src_title = SOURCE_LABELS.get(source_key, (source_key, ""))[0]
        row = _write_section_title(ws, row, f"Source: {src_title}")
        if not doc or not doc.get("accounts"):
            row = _write_line_item(ws, row, "(No line items extracted from this document)", None, currency=False)
            continue
        section_total = 0.0
        for a in doc["accounts"]:
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
            section_total += a["amount"]
        row = _write_total_row(ws, row, f"Subtotal — {src_title}", section_total)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


# =====================================================================================
# SECTION 4: STATEMENT GENERATION ORCHESTRATION
# =====================================================================================

def build_statement_workbook(statement_type: str, extracted_docs: Dict[str, dict],
                              company: str, period: str) -> Tuple[Workbook, Dict[str, Any], int]:
    """Builds the full output workbook for any of the 40 statement types and returns
    (workbook, analytics_dict_or_None, total_accounts_found)."""
    wb = _new_workbook()
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    is_tb_only = statement_type in TB_ONLY_STATEMENTS
    analytics = None
    total_accounts = 0

    if is_tb_only:
        tb_doc = extracted_docs.get("trial_balance")
        if not tb_doc:
            raise HTTPException(status_code=400, detail="Trial balance document is required but was not found.")
        accounts = tb_doc.get("accounts", [])
        total_accounts = len(accounts)
        f = compute_core_financials(accounts)
        analytics = compute_full_analytics(accounts)

        if statement_type == "income_statement":
            _build_income_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "balance_sheet":
            _build_balance_sheet_tab(wb, accounts, company, period, f)
        elif statement_type == "retained_earnings":
            _build_retained_earnings_tab(wb, accounts, company, period, f)
        elif statement_type == "equity_statement":
            _build_equity_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "trial_balance":
            _build_trial_balance_tab(wb, accounts, company, period, f)
        elif statement_type in ("ratio_analysis", "liquidity_report", "solvency_report",
                                 "profitability_report", "working_capital"):
            label_map = {
                "ratio_analysis": "Financial Ratio Analysis",
                "liquidity_report": "Liquidity Report",
                "solvency_report": "Solvency Report",
                "profitability_report": "Profitability Report",
                "working_capital": "Working Capital Report",
            }
            _build_ratio_tab(wb, analytics, company, period, sheet_name=label_map[statement_type])
        # Always include the source trial balance for traceability
        _build_trial_balance_tab(wb, accounts, company, period, f) if statement_type != "trial_balance" else None

    else:
        # Multi-source schedules: 25+ specialized statements built generically from their
        # required source documents, organized into clearly labeled sections per source.
        for doc in extracted_docs.values():
            total_accounts += len(doc.get("accounts", []))
        _build_generic_schedule_tab(wb, statement_type, info["label"], extracted_docs, company, period)

        # If an income statement and/or balance sheet were among the sources, compute analytics too
        combined_accounts = []
        for doc in extracted_docs.values():
            combined_accounts.extend(doc.get("accounts", []))
        if combined_accounts:
            analytics = compute_full_analytics(combined_accounts)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Statement")
        _write_statement_header(ws, company, info["label"], period)

    return wb, analytics, total_accounts


def determine_company_and_period(extracted_docs: Dict[str, dict]) -> Tuple[str, str]:
    company = None
    period = None
    for doc in extracted_docs.values():
        if not company and doc.get("company_name"):
            company = doc["company_name"]
        if not period and doc.get("period_label"):
            period = doc["period_label"]
    return company or "Unnamed Company", period or datetime.now().strftime("FY%Y")


def save_workbook_and_get_filename(wb: Workbook, prefix: str) -> str:
    filename = f"{prefix}_{uuid.uuid4().hex[:10]}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return filename


# =====================================================================================
# SECTION 5: FORECASTING ENGINE — PHASES 1-8 (Historical Analysis & Projection)
# =====================================================================================

FORECAST_YEARS_OUT = 5


def fc_phase1_2_validate(yearly_docs: List[dict]) -> Tuple[List[dict], List[str]]:
    """Phase 1-2: Validate and quality-check uploaded trial balances. Returns
    (sorted list of {year, accounts, label}, list of warning strings)."""
    warnings: List[str] = []
    parsed = []
    for i, doc in enumerate(yearly_docs):
        accounts = doc.get("accounts", [])
        if not accounts:
            warnings.append(f"Document {i+1} ({doc.get('period_label') or 'unknown period'}) yielded no extractable line items.")
        label = doc.get("period_label") or f"Year {i+1}"
        year_match = re.search(r"(20\d{2}|19\d{2})", label)
        year = int(year_match.group(1)) if year_match else (2020 + i)
        parsed.append({"year": year, "label": label, "accounts": accounts})

    parsed.sort(key=lambda x: x["year"])
    if len(parsed) < 3:
        warnings.append(f"Only {len(parsed)} fiscal years provided — minimum 3 recommended for reliable trend forecasting.")

    years_seen = [p["year"] for p in parsed]
    if len(set(years_seen)) != len(years_seen):
        warnings.append("Duplicate or indeterminate fiscal years detected — year labels were inferred from document content where possible.")

    return parsed, warnings


def fc_phase3_normalize(parsed_years: List[dict]) -> List[Dict[str, Any]]:
    """Phase 3: Normalize account names/categories across years into a consistent
    per-year core-financials series for trend computation."""
    series = []
    for p in parsed_years:
        f = compute_core_financials(p["accounts"])
        f["year"] = p["year"]
        f["label"] = p["label"]
        series.append(f)
    return series


def fc_phase4_historical_ratios(series: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Phase 4: Compute historical ratios for each year in the series for trend analysis."""
    ratio_history = []
    for yr in series:
        gm = _safe_div(yr["gross_profit"], yr["revenue"])
        nm = _safe_div(yr["net_income"], yr["revenue"])
        cr = _safe_div(yr["current_assets"], yr["current_liabilities"])
        de = _safe_div(yr["total_liabilities"], yr["total_equity"])
        ratio_history.append({"year": yr["year"], "gross_margin": gm, "net_margin": nm,
                               "current_ratio": cr, "debt_to_equity": de})
    return ratio_history


def _cagr(start: float, end: float, periods: int) -> Optional[float]:
    if start is None or end is None or start <= 0 or periods <= 0:
        return None
    try:
        return (end / start) ** (1 / periods) - 1
    except (ValueError, ZeroDivisionError):
        return None


def fc_phase5_6_drivers_and_growth(series: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 5-6: Identify revenue/expense drivers and compute growth forecasts via
    CAGR, weighted-average growth, and simple linear trend — the three methods referenced
    in the 'Growth Rate Forecast' statement type."""
    revenues = [s["revenue"] for s in series]
    n_periods = len(series) - 1

    cagr = _cagr(revenues[0], revenues[-1], n_periods) if n_periods > 0 else None

    yoy_growth_rates = []
    for i in range(1, len(revenues)):
        g = _safe_div(revenues[i] - revenues[i-1], revenues[i-1])
        if g is not None:
            yoy_growth_rates.append(g)
    weighted_growth = None
    if yoy_growth_rates:
        weights = list(range(1, len(yoy_growth_rates) + 1))
        weighted_growth = sum(g * w for g, w in zip(yoy_growth_rates, weights)) / sum(weights)

    # Simple linear trend (least squares slope) expressed as an implied growth rate off the latest year
    trend_growth = None
    if len(revenues) >= 2:
        x = list(range(len(revenues)))
        x_mean = statistics.mean(x)
        y_mean = statistics.mean(revenues)
        denom = sum((xi - x_mean) ** 2 for xi in x)
        if denom > 0:
            slope = sum((xi - x_mean) * (yi - y_mean) for xi, yi in zip(x, revenues)) / denom
            trend_growth = _safe_div(slope, revenues[-1])

    expense_ratios = [_safe_div(s["operating_expenses"], s["revenue"]) for s in series]
    expense_ratios = [e for e in expense_ratios if e is not None]
    avg_expense_ratio = statistics.mean(expense_ratios) if expense_ratios else 0.3

    cogs_ratios = [_safe_div(s["cogs"], s["revenue"]) for s in series]
    cogs_ratios = [c for c in cogs_ratios if c is not None]
    avg_cogs_ratio = statistics.mean(cogs_ratios) if cogs_ratios else 0.4

    chosen_growth = weighted_growth if weighted_growth is not None else (cagr if cagr is not None else 0.05)

    return {
        "cagr": cagr, "weighted_growth": weighted_growth, "trend_growth": trend_growth,
        "chosen_growth_rate": chosen_growth,
        "avg_expense_ratio": avg_expense_ratio, "avg_cogs_ratio": avg_cogs_ratio,
        "yoy_growth_rates": yoy_growth_rates,
    }


def fc_phase7_three_statement_model(series: List[Dict[str, Any]], drivers: Dict[str, Any],
                                     years_out: int = FORECAST_YEARS_OUT) -> List[Dict[str, Any]]:
    """Phase 7: Build a linked 3-statement (simplified) forecast for N years forward,
    driven by the chosen revenue growth rate and historical expense/COGS ratios."""
    last = series[-1]
    growth = drivers["chosen_growth_rate"]
    cogs_ratio = drivers["avg_cogs_ratio"]
    expense_ratio = drivers["avg_expense_ratio"]

    projections = []
    prev_revenue = last["revenue"]
    prev_assets = last["total_assets"]
    prev_liabilities = last["total_liabilities"]
    prev_equity = last["total_equity"]
    prev_cash = last["cash"]

    for i in range(1, years_out + 1):
        revenue = prev_revenue * (1 + growth)
        cogs = revenue * cogs_ratio
        gross_profit = revenue - cogs
        opex = revenue * expense_ratio
        operating_income = gross_profit - opex
        tax_rate = _safe_div(last["tax_expense"], (operating_income if operating_income else 1)) or 0.21
        tax_rate = min(max(tax_rate, 0.0), 0.40)
        interest_expense = last["interest_expense"]
        net_income = (operating_income - interest_expense) * (1 - tax_rate)

        assets = prev_assets * (1 + growth * 0.6)
        liabilities = prev_liabilities * (1 + growth * 0.5)
        equity = prev_equity + net_income
        cash = prev_cash + net_income * 0.7

        projections.append({
            "year_offset": i, "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
            "operating_expenses": opex, "operating_income": operating_income,
            "interest_expense": interest_expense, "net_income": net_income,
            "total_assets": assets, "total_liabilities": liabilities, "total_equity": equity,
            "cash": cash,
        })
        prev_revenue, prev_assets, prev_liabilities, prev_equity, prev_cash = revenue, assets, liabilities, equity, cash

    return projections


def fc_phase8_cashflow_forecast(projections: List[Dict[str, Any]], current_cash: float) -> List[Dict[str, Any]]:
    """Phase 8: Project operating/investing/financing cash flows and resulting cash runway."""
    cf_rows = []
    running_cash = current_cash
    for p in projections:
        operating_cf = p["net_income"] * 0.85
        investing_cf = -p["revenue"] * 0.03
        financing_cf = 0.0
        net_change = operating_cf + investing_cf + financing_cf
        running_cash += net_change
        monthly_burn = abs(net_change) / 12 if net_change < 0 else 0
        runway_months = (running_cash / monthly_burn) if monthly_burn > 0 else None
        cf_rows.append({
            "year_offset": p["year_offset"], "operating_cf": operating_cf, "investing_cf": investing_cf,
            "financing_cf": financing_cf, "net_change_in_cash": net_change, "ending_cash": running_cash,
            "runway_months": runway_months,
        })
    return cf_rows


# =====================================================================================
# SECTION 6: FORECASTING ENGINE — PHASES 9-15 (Scenarios, Valuation, Risk, Narrative)
# =====================================================================================

def fc_phase9_scenarios(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """Phase 9: Run Base / Best / Worst case scenario analysis by flexing the growth rate
    and expense ratio assumptions, each tagged with an illustrative probability weight."""
    base_growth = drivers["chosen_growth_rate"]
    scenarios = {}
    scenario_defs = [
        ("base", base_growth, drivers["avg_expense_ratio"], 0.50),
        ("best", base_growth + 0.08, drivers["avg_expense_ratio"] * 0.92, 0.25),
        ("worst", max(base_growth - 0.12, -0.10), drivers["avg_expense_ratio"] * 1.10, 0.25),
    ]
    for name, growth, expense_ratio, probability in scenario_defs:
        flexed_drivers = dict(drivers)
        flexed_drivers["chosen_growth_rate"] = growth
        flexed_drivers["avg_expense_ratio"] = expense_ratio
        projections = fc_phase7_three_statement_model(series, flexed_drivers, years_out=FORECAST_YEARS_OUT)
        scenarios[name] = {"projections": projections, "growth_assumed": growth,
                            "expense_ratio_assumed": expense_ratio, "probability": probability}
    return scenarios


def fc_phase10_stakeholder_analysis(series: List[Dict[str, Any]], projections: List[Dict[str, Any]],
                                     analytics_latest: Dict[str, Any]) -> Dict[str, str]:
    """Phase 10: Generate stakeholder-specific summaries (Owner, CFO, Investor, Bank, Auditor)."""
    last = series[-1]
    y1, y5 = projections[0], projections[-1]
    rev_cagr_fwd = _cagr(last["revenue"], y5["revenue"], len(projections))

    owner = (f"Revenue is projected to grow from {last['revenue']:,.0f} to {y5['revenue']:,.0f} "
             f"over {len(projections)} years (~{(rev_cagr_fwd or 0)*100:.1f}% CAGR), with net income reaching "
             f"{y5['net_income']:,.0f} by year {len(projections)}.")
    cfo = (f"Year 1 net income of {y1['net_income']:,.0f} assumes a {(analytics_latest.get('gross_margin') or 0)*100:.1f}% "
           f"gross margin held roughly flat; working capital and capex assumptions are simplified and should be "
           f"refined with a full budget once available.")
    investor = (f"Implied forward growth of ~{(rev_cagr_fwd or 0)*100:.1f}% CAGR with equity growing from "
                f"{last['total_equity']:,.0f} to {y5['total_equity']:,.0f}; see the Valuation tab for DCF-based "
                f"enterprise and equity value estimates.")
    bank = (f"Debt service capacity depends on interest coverage; current leverage shows debt-to-equity of "
            f"{analytics_latest.get('debt_to_equity') or 0:.2f}x. Cash flow forecast tab details projected runway "
            f"and coverage ratios.")
    auditor = ("This forecast is a model-based projection derived from historical trial balances using simplified "
               "linear/CAGR-based assumptions; it does not constitute audited financial statements and carries "
               "material estimation uncertainty around working capital, capex, and financing assumptions.")
    return {"owner": owner, "cfo": cfo, "investor": investor, "bank": bank, "auditor": auditor}


def fc_phase11_dcf_valuation(projections: List[Dict[str, Any]], discount_rate: float = 0.12,
                              terminal_growth: float = 0.025) -> Dict[str, Any]:
    """Phase 11: DCF valuation — discounts a simplified unlevered FCF proxy (NI + back of envelope
    add-backs) and computes terminal value via the Gordon Growth method."""
    fcf_series = [p["net_income"] * 0.85 - p["revenue"] * 0.03 for p in projections]
    pv_sum = 0.0
    pv_detail = []
    for i, fcf in enumerate(fcf_series, start=1):
        pv = fcf / ((1 + discount_rate) ** i)
        pv_sum += pv
        pv_detail.append({"year_offset": i, "fcf": fcf, "present_value": pv})

    terminal_fcf = fcf_series[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (discount_rate - terminal_growth) if discount_rate > terminal_growth else None
    pv_terminal = terminal_value / ((1 + discount_rate) ** len(fcf_series)) if terminal_value else None

    enterprise_value = pv_sum + (pv_terminal or 0)
    net_debt = projections[0]["total_liabilities"] - projections[0]["cash"]
    equity_value = enterprise_value - net_debt

    return {
        "discount_rate": discount_rate, "terminal_growth": terminal_growth,
        "pv_detail": pv_detail, "pv_of_explicit_fcf": pv_sum,
        "terminal_value": terminal_value, "pv_of_terminal_value": pv_terminal,
        "enterprise_value": enterprise_value, "net_debt_estimate": net_debt,
        "equity_value": equity_value,
    }


def fc_phase12_risk_scoring(series: List[Dict[str, Any]], analytics_latest: Dict[str, Any],
                             cashflow_forecast: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 12: Score risk across 6 dimensions (0-100, higher = riskier) and produce an overall score."""
    def clamp(v, lo=0, hi=100): return max(lo, min(hi, v))

    revenues = [s["revenue"] for s in series]
    rev_volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if len(revenues) > 1 and statistics.mean(revenues) else 0
    concentration_risk = clamp(rev_volatility * 100)

    cr = analytics_latest.get("current_ratio") or 1.0
    liquidity_risk = clamp((1.5 - cr) * 50) if cr < 1.5 else clamp((1.5 - cr) * 20)

    de = analytics_latest.get("debt_to_equity") or 0.5
    debt_risk = clamp(de * 30)

    negative_cf_years = sum(1 for cf in cashflow_forecast if cf["net_change_in_cash"] < 0)
    burn_risk = clamp((negative_cf_years / max(len(cashflow_forecast), 1)) * 100)

    nm = analytics_latest.get("net_margin")
    profitability_risk = clamp((0.1 - nm) * 300) if nm is not None and nm < 0.1 else 0

    n_years = len(series)
    data_quality_risk = clamp((3 - n_years) * 20) if n_years < 3 else 5

    dimensions = {
        "revenue_concentration_volatility": round(concentration_risk, 1),
        "liquidity_risk": round(liquidity_risk, 1),
        "debt_leverage_risk": round(debt_risk, 1),
        "cash_burn_risk": round(burn_risk, 1),
        "profitability_risk": round(profitability_risk, 1),
        "data_quality_risk": round(data_quality_risk, 1),
    }
    overall = round(statistics.mean(dimensions.values()), 1)
    if overall < 25: band = "Low Risk"
    elif overall < 50: band = "Moderate Risk"
    elif overall < 75: band = "Elevated Risk"
    else: band = "High Risk"

    return {"dimensions": dimensions, "overall_score": overall, "risk_band": band}


def fc_phase13_narrative(series, projections, drivers, risk: Dict[str, Any], company: str) -> str:
    """Phase 13: AI narrative commentary via local Ollama (Llama 3.1), with deterministic fallback."""
    try:
        import ollama
        prompt = f"""You are a financial forecasting analyst. Write a concise narrative (4-6 sentences) summarizing
this forecast for {company}.

Historical revenue: {[round(s['revenue']) for s in series]}
Forecast growth rate assumed: {drivers['chosen_growth_rate']*100:.1f}%
Year 1 forecast revenue: {projections[0]['revenue']:,.0f}, Year {len(projections)} forecast revenue: {projections[-1]['revenue']:,.0f}
Overall risk score: {risk['overall_score']}/100 ({risk['risk_band']})

Quanto is not responsible for financial decisions — write as an informative analyst, not as advice to act on.
Respond with plain text only, no markdown, no JSON."""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.4})
        return resp["message"]["content"].strip()
    except Exception:
        return (f"{company}'s historical revenue trend implies a forward growth assumption of "
                f"{drivers['chosen_growth_rate']*100:.1f}% annually, projecting revenue from "
                f"{projections[0]['revenue']:,.0f} in year 1 to {projections[-1]['revenue']:,.0f} by year "
                f"{len(projections)}. The model's overall risk score of {risk['overall_score']}/100 places this "
                f"forecast in the '{risk['risk_band']}' band, driven primarily by the dimensions with the highest "
                f"individual scores. As with any model-based projection, actual results will depend on factors "
                f"such as market conditions, execution, and financing decisions not captured in the historical "
                f"trial balances alone. (Narrative generated via fallback — start Ollama with llama3.1 for richer AI commentary.)")


def fc_phase14_confidence_scores(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, str]:
    """Phase 14: Qualitative confidence scoring for revenue and margin assumptions based on
    data sufficiency and historical volatility."""
    n_years = len(series)
    revenues = [s["revenue"] for s in series]
    volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if n_years > 1 and statistics.mean(revenues) else 1

    def score(n_years_local, volatility_local):
        if n_years_local >= 5 and volatility_local < 0.15: return "High"
        if n_years_local >= 3 and volatility_local < 0.30: return "Medium"
        return "Low"

    revenue_confidence = score(n_years, volatility)
    margin_volatility = statistics.pstdev([_safe_div(s["gross_profit"], s["revenue"]) or 0 for s in series]) if n_years > 1 else 0.2
    margins_confidence = score(n_years, margin_volatility * 2)

    return {"revenue": revenue_confidence, "margins": margins_confidence}


def chat_about_financials_unused_placeholder():
    pass


def fc_phase15_build_workbook(company: str, period: str, series: List[Dict[str, Any]],
                               drivers: Dict[str, Any], projections: List[Dict[str, Any]],
                               cashflow_forecast: List[Dict[str, Any]], scenarios: Dict[str, Any],
                               stakeholder: Dict[str, str], dcf: Dict[str, Any], risk: Dict[str, Any],
                               narrative: str, confidence: Dict[str, str]) -> Workbook:
    """Phase 15: Assemble the full 13-tab forecast Excel workbook."""
    wb = _new_workbook()

    # Tab 1: Historical Summary
    ws = wb.create_sheet("Historical Summary")
    row = _write_statement_header(ws, company, "Historical Financial Summary", period)
    headers = ["Metric"] + [str(s["year"]) for s in series]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    metrics = ["revenue", "cogs", "gross_profit", "operating_expenses", "ebit", "net_income",
               "total_assets", "total_liabilities", "total_equity", "cash"]
    for m in metrics:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, s in enumerate(series):
            cell = ws.cell(row=row, column=2 + j, value=s.get(m))
            cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(series))}})

    # Tab 2: Growth & Driver Assumptions
    ws = wb.create_sheet("Growth Assumptions")
    row = _write_statement_header(ws, company, "Growth Rate & Driver Assumptions", period)
    rows_data = [
        ("CAGR (Historical)", drivers["cagr"], PCT_FMT),
        ("Weighted-Average Growth", drivers["weighted_growth"], PCT_FMT),
        ("Linear Trend-Implied Growth", drivers["trend_growth"], PCT_FMT),
        ("Chosen Forward Growth Rate", drivers["chosen_growth_rate"], PCT_FMT),
        ("Average COGS / Revenue Ratio", drivers["avg_cogs_ratio"], PCT_FMT),
        ("Average Opex / Revenue Ratio", drivers["avg_expense_ratio"], PCT_FMT),
    ]
    for label, value, fmt in rows_data:
        row = _write_line_item(ws, row, label, value, currency=False)
        ws.cell(row=row - 1, column=4).number_format = fmt
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 3: Three-Statement Forecast
    ws = wb.create_sheet("3-Statement Forecast")
    row = _write_statement_header(ws, company, "Three-Statement Forecast Model", period, span=len(projections) + 1)
    headers = ["Metric"] + [f"Year +{p['year_offset']}" for p in projections]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["revenue", "cogs", "gross_profit", "operating_expenses", "operating_income",
              "net_income", "total_assets", "total_liabilities", "total_equity", "cash"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, p in enumerate(projections):
            cell = ws.cell(row=row, column=2 + j, value=p.get(m)); cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(projections))}})

    # Tab 4: Cash Flow Forecast
    ws = wb.create_sheet("Cash Flow Forecast")
    row = _write_statement_header(ws, company, "Cash Flow Forecast", period, span=len(cashflow_forecast) + 1)
    headers = ["Metric"] + [f"Year +{c['year_offset']}" for c in cashflow_forecast]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["operating_cf", "investing_cf", "financing_cf", "net_change_in_cash", "ending_cash", "runway_months"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, cf in enumerate(cashflow_forecast):
            val = cf.get(m)
            cell = ws.cell(row=row, column=2 + j, value=val)
            if m != "runway_months":
                cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(cashflow_forecast))}})

    # Tab 5: Scenario Analysis
    ws = wb.create_sheet("Scenario Analysis")
    row = _write_statement_header(ws, company, "Scenario Analysis (Base / Best / Worst)", period, span=4)
    headers = ["Scenario", "Growth Assumed", "Probability", f"Revenue Yr+{FORECAST_YEARS_OUT}"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for name, sc in scenarios.items():
        ws.cell(row=row, column=1, value=name.title()).font = BOLD_LABEL_FONT
        gc = ws.cell(row=row, column=2, value=sc["growth_assumed"]); gc.number_format = PCT_FMT
        pc = ws.cell(row=row, column=3, value=sc["probability"]); pc.number_format = PCT_FMT
        rc = ws.cell(row=row, column=4, value=sc["projections"][-1]["revenue"]); rc.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 16, 2: 16, 3: 14, 4: 18})

    # Tab 6: Valuation (DCF)
    ws = wb.create_sheet("Valuation (DCF)")
    row = _write_statement_header(ws, company, "DCF Valuation Model", period)
    row = _write_line_item(ws, row, "Discount Rate (WACC proxy)", dcf["discount_rate"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "Terminal Growth Rate", dcf["terminal_growth"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "PV of Explicit-Period FCF", dcf["pv_of_explicit_fcf"])
    row = _write_line_item(ws, row, "Terminal Value", dcf["terminal_value"])
    row = _write_line_item(ws, row, "PV of Terminal Value", dcf["pv_of_terminal_value"])
    row = _write_total_row(ws, row, "Enterprise Value", dcf["enterprise_value"])
    row = _write_line_item(ws, row, "Less: Net Debt (Estimate)", -dcf["net_debt_estimate"] if dcf["net_debt_estimate"] else 0)
    row = _write_total_row(ws, row, "Equity Value", dcf["equity_value"])
    _autosize_columns(ws, {1: 32, 4: 18})

    # Tab 7: Risk Analysis
    ws = wb.create_sheet("Risk Analysis")
    row = _write_statement_header(ws, company, "Risk Analysis Report", period)
    for dim, score in risk["dimensions"].items():
        row = _write_line_item(ws, row, dim.replace("_", " ").title(), score, currency=False)
        ws.cell(row=row - 1, column=4).number_format = '0.0'
    row = _write_total_row(ws, row, "Overall Risk Score (0-100)", risk["overall_score"])
    c = ws.cell(row=row, column=1, value=f"Risk Band: {risk['risk_band']}")
    c.font = Font(bold=True, color=(XL_GREEN if risk["overall_score"] < 25 else XL_RED if risk["overall_score"] >= 75 else "B8860B"))
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 8: Stakeholder Analysis
    ws = wb.create_sheet("Stakeholder Analysis")
    row = _write_statement_header(ws, company, "Stakeholder Analysis", period)
    for role, text in stakeholder.items():
        row = _write_section_title(ws, row, role.title())
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=text)
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 45
        row += 2
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 9: AI Narrative Insights
    ws = wb.create_sheet("AI Narrative")
    row = _write_statement_header(ws, company, "AI Narrative Insights", period)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=narrative)
    c.font = LABEL_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 120
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 10: Confidence Scores
    ws = wb.create_sheet("Confidence Scores")
    row = _write_statement_header(ws, company, "Forecast Confidence Scores", period)
    row = _write_line_item(ws, row, "Revenue Forecast Confidence", confidence["revenue"], currency=False)
    row = _write_line_item(ws, row, "Margin Forecast Confidence", confidence["margins"], currency=False)
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 11: Historical Ratios Trend
    ws = wb.create_sheet("Historical Ratio Trend")
    row = _write_statement_header(ws, company, "Historical Ratio Trend", period, span=len(series) + 1)
    ratio_hist = fc_phase4_historical_ratios(series)
    headers = ["Ratio"] + [str(r["year"]) for r in ratio_hist]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m, fmt in [("gross_margin", PCT_FMT), ("net_margin", PCT_FMT), ("current_ratio", X_FMT), ("debt_to_equity", X_FMT)]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, r in enumerate(ratio_hist):
            cell = ws.cell(row=row, column=2 + j, value=r.get(m))
            if r.get(m) is not None: cell.number_format = fmt
        row += 1
    _autosize_columns(ws, {1: 22, **{i: 12 for i in range(2, 2 + len(series))}})

    # Tab 12: Methodology & Assumptions Notes
    ws = wb.create_sheet("Methodology Notes")
    row = _write_statement_header(ws, company, "Methodology & Assumptions", period)
    notes = [
        "Revenue forecast uses a weighted-average year-over-year growth rate (more weight on recent years), "
        "falling back to historical CAGR if insufficient growth history exists.",
        "COGS and operating expenses are forecast as a constant percentage of revenue based on the historical average ratio.",
        "Balance sheet items (assets, liabilities) are scaled at a fraction of the revenue growth rate as a simplifying assumption.",
        "Cash flow forecast approximates operating cash flow from net income and assumes capex of 3% of revenue with no financing activity.",
        "DCF valuation discounts an unlevered free-cash-flow proxy at the specified discount rate, with Gordon Growth terminal value.",
        "This model is for planning and directional insight only — it is not a substitute for a full FP&A build-out or professional valuation.",
    ]
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"• {note}")
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 13: Cover / Disclaimer
    ws = wb.create_sheet("Cover", 0)
    ws.merge_cells("A1:D3")
    c = ws.cell(row=1, column=1, value=f"{company}\nFull Forecasting Package\n{period}")
    c.font = Font(size=16, bold=True, color=XL_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A5:D7")
    c2 = ws.cell(row=5, column=1, value="Generated by Quanto Financial Intelligence Platform — 15-Phase Forecasting Engine.\n\n"
                                         "Quanto is not responsible for financial decisions. This document is a model-based "
                                         "forecast built on historical trial balances and simplified assumptions; actual "
                                         "results will vary.")
    c2.font = Font(size=10, italic=True, color=XL_GREY)
    c2.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    _autosize_columns(ws, {1: 24, 2: 24, 3: 24, 4: 24})

    return wb


# =====================================================================================
# SECTION 7: FASTAPI ROUTES
# =====================================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    return HTMLResponse(content=HTML)


@app.get("/api/ocr-status")
async def api_ocr_status():
    return JSONResponse(get_ocr_status())


@app.get("/api/statements")
async def api_statements():
    return JSONResponse(STATEMENTS)


@app.post("/api/generate")
async def api_generate(statement_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    _enforce_free_plan_limit("financial_statements_generated", FREE_STATEMENT_LIMIT, STATEMENT_LIMIT_MESSAGE)

    extracted_docs: Dict[str, dict] = {}
    is_tb_only = statement_type in TB_ONLY_STATEMENTS

    if is_tb_only:
        if not files:
            raise HTTPException(status_code=400, detail="A trial balance file is required.")
        file = files[0]
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        extracted_docs["trial_balance"] = extract_document(content, file.filename or "upload")
    else:
        required_sources = set(info.get("sources", []))
        for file in files:
            raw_name = file.filename or ""
            if "::" in raw_name:
                source_key, original_name = raw_name.split("::", 1)
            else:
                source_key, original_name = (required_sources.pop() if required_sources else "unknown"), raw_name
            content = await file.read()
            if not content:
                continue
            extracted_docs[source_key] = extract_document(content, original_name or raw_name)

        missing = [s for s in info.get("sources", []) if s not in extracted_docs]
        if missing:
            missing_labels = [SOURCE_LABELS.get(m, (m, ""))[0] for m in missing]
            raise HTTPException(status_code=400, detail=f"Missing required source document(s): {', '.join(missing_labels)}")

    company, period = determine_company_and_period(extracted_docs)
    wb, analytics, accounts_found = build_statement_workbook(statement_type, extracted_docs, company, period)
    filename = save_workbook_and_get_filename(wb, prefix=statement_type)

    ocr_provider_used = next(iter(extracted_docs.values()), {}).get("_ocr_provider_used", get_ocr_status()["provider"])

    response_payload = {
        "statement_type": statement_type,
        "statement_label": info["label"],
        "company": company,
        "period": period,
        "accounts_found": accounts_found,
        "ocr_provider": ocr_provider_used,
        "filename": filename,
        "analytics": analytics,
    }

    if QUANTO_PLAN != "paid":
        _increment_usage("financial_statements_generated")

    return JSONResponse(response_payload)


@app.post("/api/insights")
async def api_insights(payload: Dict[str, Any]):
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    statement_type = payload.get("statement_type", "")
    insights = generate_ai_insights(analytics, company, period, statement_type)
    return JSONResponse({"insights": insights})


@app.post("/api/chat")
async def api_chat(payload: Dict[str, Any]):
    message = payload.get("message", "")
    history = payload.get("history", [])
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    if not message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty.")
    reply = chat_about_financials(message, history, analytics, company, period)
    return JSONResponse({"reply": reply})


@app.post("/api/forecast")
async def api_forecast(forecast_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(forecast_type)
    if not info or info.get("category") != "forecast":
        raise HTTPException(status_code=400, detail=f"Unknown forecast type: {forecast_type}")
    if len(files) < 3:
        raise HTTPException(status_code=400, detail="At least 3 fiscal years of trial balances are required.")

    _enforce_free_plan_limit("forecasts_generated", FREE_FORECAST_LIMIT, FORECAST_LIMIT_MESSAGE)

    yearly_docs = []
    for file in files:
        content = await file.read()
        if not content:
            continue
        yearly_docs.append(extract_document(content, file.filename or "upload"))

    if len(yearly_docs) < 3:
        raise HTTPException(status_code=400, detail="At least 3 valid (non-empty) trial balance files are required.")

    parsed_years, warnings = fc_phase1_2_validate(yearly_docs)
    series = fc_phase3_normalize(parsed_years)
    drivers = fc_phase5_6_drivers_and_growth(series)
    projections = fc_phase7_three_statement_model(series, drivers, years_out=FORECAST_YEARS_OUT)
    cashflow_forecast = fc_phase8_cashflow_forecast(projections, series[-1]["cash"])
    scenarios = fc_phase9_scenarios(series, drivers)

    latest_accounts = parsed_years[-1]["accounts"]
    analytics_latest = compute_full_analytics(latest_accounts)

    stakeholder = fc_phase10_stakeholder_analysis(series, projections, analytics_latest)
    dcf = fc_phase11_dcf_valuation(projections)
    risk = fc_phase12_risk_scoring(series, analytics_latest, cashflow_forecast)
    company, _ = determine_company_and_period({"_": yearly_docs[-1]})
    narrative = fc_phase13_narrative(series, projections, drivers, risk, company)
    confidence = fc_phase14_confidence_scores(series, drivers)

    period_label = f"{series[0]['year']}–{series[-1]['year']} Historical · {series[-1]['year']+1}–{series[-1]['year']+FORECAST_YEARS_OUT} Forecast"

    wb = fc_phase15_build_workbook(company, period_label, series, drivers, projections, cashflow_forecast,
                                    scenarios, stakeholder, dcf, risk, narrative, confidence)
    filename = save_workbook_and_get_filename(wb, prefix=forecast_type)

    response_payload = {
        "forecast_type": forecast_type,
        "forecast_label": info["label"],
        "company": company,
        "period": period_label,
        "years_analyzed": len(series),
        "phases_run": 15,
        "confidence_scores": confidence,
        "validation_warnings": warnings,
        "filename": filename,
    }

    if QUANTO_PLAN != "paid":
        _increment_usage("forecasts_generated")

    return JSONResponse(response_payload)


# =====================================================================================
# SECTION 8: INTERACTIVE FINANCIAL LEDGER MODE (STATEFUL)
# =====================================================================================
# A dynamic, stateful accounting ledger session. Unlike the one-shot Statement Generator,
# this maintains an in-memory accounting state per session that persists across turns:
# every accepted adjustment (add/remove/increase/decrease/replace/reclassify/merge), plus
# undo/reset, updates the ledger and triggers a full deterministic recalculation of the
# trial balance and every ratio. No figure is ever estimated by an LLM — all math is plain
# Python arithmetic built on the same compute_core_financials / compute_full_analytics
# functions used by the Statement Generator, so the two stay consistent.

import copy

LEDGER_SESSIONS: Dict[str, Dict[str, Any]] = {}

LEDGER_DEBIT_CATEGORIES = {"asset", "cogs", "expense"}


def _ledger_new_account(name: str, category: str, amount: float, source: str) -> Dict[str, Any]:
    return {
        "account_name": name.strip(),
        "account_code": None,
        "category": category,
        "amount": round(float(amount), 2),
        "source_document": source,
        "last_modified": datetime.now().isoformat(timespec="seconds"),
        "modification_history": ["Created"],
    }


def _ledger_parse_number(raw: str) -> float:
    return float(raw.replace(",", "").replace("$", "").strip())


def _ledger_find_account_idx(accounts: List[Dict[str, Any]], query: str) -> Optional[int]:
    import difflib
    q = query.strip().lower().strip(" .")
    if not q:
        return None
    # 1) exact match
    for i, a in enumerate(accounts):
        if a["account_name"].strip().lower() == q:
            return i
    # 2) substring match either direction
    candidates = [i for i, a in enumerate(accounts)
                  if q in a["account_name"].lower() or a["account_name"].lower() in q]
    if candidates:
        return candidates[0]
    # 3) stem-stripped word-overlap match (handles "salaries" vs "Salary Expense")
    strip_words = {"expense", "expenses", "account", "cost", "costs"}
    def stem_tokens(s):
        return {w[:5] for w in re.findall(r"[a-z]+", s.lower()) if w not in strip_words}
    q_tokens = stem_tokens(q)
    if q_tokens:
        for i, a in enumerate(accounts):
            a_tokens = stem_tokens(a["account_name"])
            if q_tokens & a_tokens:
                return i
    # 4) fuzzy closest-match fallback
    names = [a["account_name"] for a in accounts]
    close = difflib.get_close_matches(query.strip(), names, n=1, cutoff=0.72)
    if close:
        return next(i for i, a in enumerate(accounts) if a["account_name"] == close[0])
    return None


def _ledger_infer_category(name: str) -> str:
    n = name.lower()
    if "owner" in n and any(k in n for k in ("investment", "capital", "contribution")):
        return "equity"
    return _classify_account(name)


def _ledger_apply_command(session: Dict[str, Any], raw_command: str) -> str:
    """Parses a plain-English accounting command, applies it deterministically to the
    session's ledger state, and returns a human-readable description of what changed.
    Raises ValueError with a clarification message if the command cannot be understood
    or references an account that cannot be found — per spec, Quanto never guesses."""
    cmd = raw_command.strip()
    if not cmd:
        raise ValueError("Please enter an accounting command (e.g. 'increase rent to 15,000').")
    cmd_l = cmd.lower()
    accounts = session["accounts"]
    pre_snapshot = copy.deepcopy(accounts)

    def commit(description: str) -> str:
        session["history"].append({"description": description, "snapshot": pre_snapshot,
                                    "timestamp": datetime.now().isoformat(timespec="seconds")})
        return description

    # --- reset ---
    if re.match(r'^reset\b', cmd_l):
        session["accounts"] = copy.deepcopy(session["original_accounts"])
        session["history"] = []
        return "Reset to original trial balance."

    # --- undo / reverse ---
    if re.match(r'^(undo|reverse)\b', cmd_l):
        if not session["history"]:
            raise ValueError("There is no prior adjustment to undo.")
        last = session["history"].pop()
        session["accounts"] = last["snapshot"]
        return f"Reversed: {last['description']}"

    # --- replace X with Y ---
    m = re.match(r'^replace\s+(.+?)\s+with\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        old_name, new_name = m.group(1).strip(), m.group(2).strip()
        idx = _ledger_find_account_idx(accounts, old_name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{old_name}'. Please check the account name.")
        accounts[idx]["modification_history"].append(f"Renamed from '{accounts[idx]['account_name']}' to '{new_name}'")
        accounts[idx]["account_name"] = new_name
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"Replaced '{old_name}' with '{new_name}'")

    # --- merge X into Y ---
    m = re.match(r'^merge\s+(.+?)\s+into\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        src_name, dst_name = m.group(1).strip(), m.group(2).strip()
        src_idx = _ledger_find_account_idx(accounts, src_name)
        if src_idx is None:
            raise ValueError(f"I couldn't find an account matching '{src_name}'.")
        dst_idx = _ledger_find_account_idx(accounts, dst_name)
        if dst_idx is None:
            accounts.append(_ledger_new_account(dst_name, accounts[src_idx]["category"], 0.0, "user adjustment"))
            dst_idx = len(accounts) - 1
        moved_name, moved_amount = accounts[src_idx]["account_name"], accounts[src_idx]["amount"]
        accounts[dst_idx]["amount"] = round(accounts[dst_idx]["amount"] + moved_amount, 2)
        accounts[dst_idx]["modification_history"].append(f"Absorbed {moved_name} ({moved_amount:,.2f})")
        accounts.pop(src_idx)
        return commit(f"Merged '{moved_name}' into '{accounts[dst_idx]['account_name']}'")

    # --- remove / delete X ---
    m = re.match(r'^(?:remove|delete)\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'. Please check the account name.")
        removed = accounts.pop(idx)
        return commit(f"Removed {removed['account_name']} ({removed['amount']:,.2f})")

    # --- reclassify X as <category> ---
    m = re.match(r'^reclassify\s+(.+?)\s+as\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        name, new_cat_raw = m.group(1).strip(), m.group(2).strip().lower()
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'.")
        valid_cats = ["asset", "liability", "equity", "revenue", "cogs", "expense"]
        new_cat = next((c for c in valid_cats if c in new_cat_raw), None)
        if not new_cat:
            raise ValueError(f"I couldn't determine a valid category from '{new_cat_raw}'. "
                              f"Use one of: {', '.join(valid_cats)}.")
        old_cat = accounts[idx]["category"]
        accounts[idx]["category"] = new_cat
        accounts[idx]["modification_history"].append(f"Reclassified from {old_cat} to {new_cat}")
        return commit(f"Reclassified {accounts[idx]['account_name']} from {old_cat} to {new_cat}")

    # --- change/set/increase/decrease X to Y (absolute set) ---
    m = re.match(r'^(?:change|set|increase|decrease)\s+(.+?)\s+to\s+([\d,\.]+)$', cmd, re.IGNORECASE)
    if m:
        name, amount_str = m.group(1).strip(), m.group(2)
        new_amount = round(_ledger_parse_number(amount_str), 2)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            accounts.append(_ledger_new_account(name, _ledger_infer_category(name), new_amount, "user adjustment"))
            return commit(f"Added {name} at {new_amount:,.2f}")
        old_amount = accounts[idx]["amount"]
        accounts[idx]["amount"] = new_amount
        accounts[idx]["modification_history"].append(f"Changed from {old_amount:,.2f} to {new_amount:,.2f}")
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"Changed {accounts[idx]['account_name']} from {old_amount:,.2f} to {new_amount:,.2f}")

    # --- increase/decrease X by Y (amount or %) ---
    m = re.match(r'^(increase|decrease)\s+(.+?)\s+by\s+([\d,\.]+)\s*(%|percent)?$', cmd, re.IGNORECASE)
    if m:
        direction, name, amount_str, pct_flag = m.group(1).lower(), m.group(2).strip(), m.group(3), m.group(4)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'. Please check the account name.")
        old_amount = accounts[idx]["amount"]
        delta_value = _ledger_parse_number(amount_str)
        delta = old_amount * (delta_value / 100.0) if pct_flag else delta_value
        new_amount = old_amount + delta if direction == "increase" else old_amount - delta
        new_amount = round(max(new_amount, 0.0), 2)
        accounts[idx]["amount"] = new_amount
        accounts[idx]["modification_history"].append(f"{direction.title()}d from {old_amount:,.2f} to {new_amount:,.2f}")
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"{direction.title()}d {accounts[idx]['account_name']} from {old_amount:,.2f} to {new_amount:,.2f}")

    # --- add X of Y / record X of Y ---
    m = re.match(r'^(?:add|record)\s+(.+?)\s+of\s+([\d,\.]+)$', cmd, re.IGNORECASE)
    if m:
        name, amount_str = m.group(1).strip(), m.group(2)
        amount = round(_ledger_parse_number(amount_str), 2)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is not None:
            old_amount = accounts[idx]["amount"]
            accounts[idx]["amount"] = round(old_amount + amount, 2)
            accounts[idx]["modification_history"].append(f"Added {amount:,.2f} (was {old_amount:,.2f})")
            return commit(f"Added {amount:,.2f} to existing account {accounts[idx]['account_name']}")
        category = _ledger_infer_category(name)
        accounts.append(_ledger_new_account(name, category, amount, "user adjustment"))
        return commit(f"Added new account '{name}' of {amount:,.2f} ({category})")

    raise ValueError(
        "I couldn't parse that as an accounting command. Try formats like: 'remove rent expense', "
        "'increase salaries to 50,000', 'decrease advertising by 10%', 'add fuel expense of 2,500', "
        "'replace utilities with electricity expense', 'record depreciation of 7,000', "
        "'reclassify loan as liability', 'merge travel into operating expenses', 'undo', or 'reset'."
    )


def _ledger_format_state(session: Dict[str, Any]) -> str:
    """Renders the full required output format: Ledger State, Trial Balance, Ratios, and
    Modification History — all computed fresh from the current session accounts."""
    accounts = session["accounts"]
    f = compute_core_financials(accounts)
    analytics = compute_full_analytics(accounts)

    def pct(v): return f"{v*100:.1f}%" if v is not None else "N/A"
    def xr(v): return f"{v:.2f}x" if v is not None else "N/A"

    lines: List[str] = []
    lines.append(f"{session.get('company','Company')} · {session.get('period','')}".strip(" ·"))
    lines.append("")
    lines.append("UPDATED FINANCIAL LEDGER STATE")
    lines.append("-" * 62)
    for a in accounts:
        lines.append(f"  {a['account_name']:<34} {a['amount']:>15,.2f}   [{a['category']}]")
    lines.append("-" * 62)
    lines.append(f"  {'Total Revenue':<34} {f['revenue']:>15,.2f}")
    lines.append(f"  {'Total COGS':<34} {f['cogs']:>15,.2f}")
    lines.append(f"  {'Gross Profit':<34} {f['gross_profit']:>15,.2f}")
    lines.append(f"  {'Total Operating Expenses':<34} {f['operating_expenses']:>15,.2f}")
    lines.append(f"  {'Operating Income (EBIT)':<34} {f['ebit']:>15,.2f}")
    lines.append(f"  {'Net Income':<34} {f['net_income']:>15,.2f}")
    lines.append("")

    lines.append("CURRENT TRIAL BALANCE")
    lines.append("-" * 62)
    total_debit = total_credit = 0.0
    for a in accounts:
        is_debit = a["category"] in LEDGER_DEBIT_CATEGORIES
        amt = abs(a["amount"])
        if is_debit:
            total_debit += amt
            lines.append(f"  {a['account_name']:<34} DR {amt:>13,.2f}")
        else:
            total_credit += amt
            lines.append(f"  {a['account_name']:<34}    CR {amt:>13,.2f}")
    balance_diff = abs(total_debit - total_credit)
    lines.append("-" * 62)
    lines.append(f"  {'TOTALS':<34} DR {total_debit:>13,.2f} / CR {total_credit:>13,.2f}")
    lines.append("  ✓ Trial balance is in balance." if balance_diff < 0.01
                  else f"  ⚠ Out of balance by {balance_diff:,.2f} — review the accounts above.")
    lines.append("")

    lines.append("PERFORMANCE RATIOS")
    lines.append("-" * 62)
    lines.append(f"  Net Margin Ratio:              {pct(analytics['net_margin'])}")
    lines.append(f"  Gross Margin Ratio:            {pct(analytics['gross_margin'])}")
    lines.append(f"  Operating Margin:               {pct(analytics['operating_margin'])}")
    lines.append(f"  Current Ratio:                  {xr(analytics['current_ratio'])}")
    lines.append(f"  Quick Ratio:                    {xr(analytics['quick_ratio'])}")
    lines.append(f"  Debt-to-Equity:                 {xr(analytics['debt_to_equity'])}")
    lines.append(f"  Return on Assets (ROA):         {pct(analytics['roa'])}")
    lines.append(f"  Return on Equity (ROE):         {pct(analytics['roe'])}")
    lines.append("")

    lines.append("MODIFICATION HISTORY")
    lines.append("-" * 62)
    if session["history"]:
        for h in session["history"]:
            lines.append(f"  ✓ {h['description']}")
    else:
        lines.append("  (No adjustments yet — this is the original extracted trial balance.)")
    lines.append("")
    lines.append("Ready for additional accounting adjustments. Continue using natural language.")
    lines.append("Quanto is not responsible for financial decisions.")
    return "\n".join(lines)


# ---- Interactive Ledger Mode API routes ----

@app.post("/api/ledger/init")
async def api_ledger_init(file: UploadFile = File(...)):
    """Initializes a new stateful ledger session from an uploaded trial balance / ledger
    document. Extraction reuses the same OCR pipeline as the Statement Generator so figures
    stay consistent across features."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    doc = extract_document(content, file.filename or "upload")
    raw_accounts = doc.get("accounts", [])
    if not raw_accounts:
        raise HTTPException(status_code=422, detail="No accounts could be extracted from this document.")

    accounts: List[Dict[str, Any]] = []
    for a in raw_accounts:
        accounts.append({
            "account_name": a.get("account_name", "Unnamed Account"),
            "account_code": None,
            "category": a.get("category") or _classify_account(a.get("account_name", "")),
            "amount": round(abs(float(a.get("amount", 0.0))), 2),
            "source_document": file.filename or "upload",
            "last_modified": datetime.now().isoformat(timespec="seconds"),
            "modification_history": ["Extracted from source document"],
        })

    session_id = uuid.uuid4().hex[:16]
    company = doc.get("company_name") or "Unnamed Company"
    period = doc.get("period_label") or datetime.now().strftime("FY%Y")

    session = {
        "session_id": session_id,
        "company": company,
        "period": period,
        "accounts": accounts,
        "original_accounts": copy.deepcopy(accounts),
        "history": [],
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    LEDGER_SESSIONS[session_id] = session

    return JSONResponse({
        "session_id": session_id,
        "company": company,
        "period": period,
        "accounts_found": len(accounts),
        "formatted_state": _ledger_format_state(session),
    })


@app.post("/api/ledger/command")
async def api_ledger_command(payload: Dict[str, Any]):
    """Applies a single natural-language accounting command to an existing ledger session,
    recalculates the trial balance and every ratio deterministically, and returns the
    updated full state plus the running modification history."""
    session_id = payload.get("session_id", "")
    command = payload.get("command", "")
    session = LEDGER_SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Ledger session not found or has expired. Please upload a trial balance to start a new session.")
    if not command or not command.strip():
        raise HTTPException(status_code=400, detail="Command cannot be empty.")

    try:
        _ledger_apply_command(session, command)
    except ValueError as e:
        # Clarification needed — do not mutate state, just surface the guidance alongside the current state.
        return JSONResponse({
            "session_id": session_id,
            "applied": False,
            "clarification": str(e),
            "formatted_state": str(e) + "\n\n" + _ledger_format_state(session),
        })

    return JSONResponse({
        "session_id": session_id,
        "applied": True,
        "formatted_state": _ledger_format_state(session),
    })


@app.get("/api/download/{filename}")
async def api_download(filename: str):
    safe_name = Path(filename).name  # prevent path traversal
    filepath = OUTPUT_DIR / safe_name
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found or has expired.")
    return FileResponse(
        path=filepath,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =====================================================================================
# SECTION 9: ENTERPRISE FORECASTING ENGINE ARCHITECTURE
# =====================================================================================
# A modular, extensible architecture of independent "analysis engines" — one per
# financial forecasting model type (Growth Rate, Driver-Based, Cash Flow, Stakeholder,
# Valuation, Risk, Three-Statement, Scenario, Sensitivity, Profitability, Forecast
# Quality) plus a Document Classification Engine that identifies which model types are
# present in the uploaded data. Every engine implements the same lifecycle — detect,
# extract, validate, calculate, generate_kpis, generate_insights, detect_risks,
# generate_recommendations — so new engines can be added by subclassing ForecastEngine
# without touching the orchestrator (EnterpriseReportGenerator) or any other engine.
#
# All financial math here is real, deterministic Python arithmetic. Where an engine
# needs data this application does not currently capture end-to-end (e.g. per-customer
# unit economics or true customer-concentration data), it is explicit about that
# limitation in its confidence score and insights rather than fabricating figures.

from abc import ABC, abstractmethod
from enum import Enum
from pydantic import BaseModel, Field
from fastapi.encoders import jsonable_encoder
import copy as _copy


# ---- Shared data models -------------------------------------------------------------

class RiskSeverity(str, Enum):
    LOW = "low"
    MODERATE = "moderate"
    ELEVATED = "elevated"
    HIGH = "high"
    CRITICAL = "critical"


def _severity_from_score(score_0_100: float) -> RiskSeverity:
    if score_0_100 < 20: return RiskSeverity.LOW
    if score_0_100 < 40: return RiskSeverity.MODERATE
    if score_0_100 < 60: return RiskSeverity.ELEVATED
    if score_0_100 < 80: return RiskSeverity.HIGH
    return RiskSeverity.CRITICAL


_RISK_MITIGATION_HINTS = {
    "revenue_concentration_volatility": "Diversify revenue sources and smooth seasonal/customer concentration where possible.",
    "liquidity_risk": "Build a cash buffer and monitor near-term cash flow closely.",
    "debt_leverage_risk": "Consider a deleveraging plan or renegotiating debt terms.",
    "cash_burn_risk": "Reduce burn rate or secure additional runway via financing.",
    "profitability_risk": "Review pricing and cost structure to improve margins.",
    "data_quality_risk": "Provide additional historical periods to improve forecast reliability.",
}


def _risk_mitigation_hint(dim: str) -> str:
    return _RISK_MITIGATION_HINTS.get(dim, "Review this risk dimension with finance leadership.")


class EngineKPI(BaseModel):
    name: str
    value: Optional[float] = None
    unit: str = ""
    trend: Optional[str] = None


class RiskItem(BaseModel):
    category: str
    description: str
    probability: float
    impact: float
    severity: RiskSeverity
    mitigation: str


class EngineReport(BaseModel):
    engine_name: str
    applicable: bool
    confidence: float
    summary: str
    kpis: List[EngineKPI] = Field(default_factory=list)
    insights: List[str] = Field(default_factory=list)
    risks: List[RiskItem] = Field(default_factory=list)
    recommendations: List[str] = Field(default_factory=list)
    data: Dict[str, Any] = Field(default_factory=dict)


class DetectionResult(BaseModel):
    document_type: str
    confidence: float
    reasoning: str
    detected_sections: List[str]
    detected_tables: List[str]
    detected_financial_statements: List[str]


class FinalEnterpriseReport(BaseModel):
    company: str
    period: str
    generated_at: str
    executive_summary: str
    document_classification: List[DetectionResult]
    engine_reports: List[EngineReport]
    top_risks: List[RiskItem]
    strategic_recommendations: List[str]
    board_level_conclusion: str
    validation_warnings: List[str] = Field(default_factory=list)


# ---- Shared forecast context (computed once, reused by every engine) ----------------

class ForecastContext:
    """Holds the shared, pre-computed forecasting data every engine reads from — built
    once per request so no engine has to redundantly re-run the core projection math.
    This is the primary mechanism for avoiding duplicated logic across engines."""

    def __init__(self, company: str, period_label: str, parsed_years: List[dict], series: List[Dict[str, Any]],
                 drivers: Dict[str, Any], projections: List[Dict[str, Any]], cashflow_forecast: List[Dict[str, Any]],
                 scenarios: Dict[str, Any], stakeholder: Dict[str, str], dcf: Dict[str, Any], risk: Dict[str, Any],
                 narrative: str, confidence: Dict[str, str], analytics_latest: Dict[str, Any],
                 operational_inputs: Optional[List[Dict[str, Any]]], discount_rate: float, terminal_growth: float,
                 comparable_multiples: Dict[str, float], validation_warnings: List[str]):
        self.company = company
        self.period_label = period_label
        self.parsed_years = parsed_years
        self.series = series
        self.drivers = drivers
        self.projections = projections
        self.cashflow_forecast = cashflow_forecast
        self.scenarios = scenarios
        self.stakeholder = stakeholder
        self.dcf = dcf
        self.risk = risk
        self.narrative = narrative
        self.confidence = confidence
        self.analytics_latest = analytics_latest
        self.operational_inputs = operational_inputs
        self.discount_rate = discount_rate
        self.terminal_growth = terminal_growth
        self.comparable_multiples = comparable_multiples
        self.validation_warnings = validation_warnings


def build_enterprise_context(yearly_docs: List[dict], operational_inputs: Optional[List[Dict[str, Any]]] = None,
                              discount_rate: float = 0.12, terminal_growth: float = 0.025,
                              comparable_multiples: Optional[Dict[str, float]] = None) -> ForecastContext:
    parsed_years, warnings = fc_phase1_2_validate(yearly_docs)
    series = fc_phase3_normalize(parsed_years)
    drivers = fc_phase5_6_drivers_and_growth(series)
    projections = fc_phase7_three_statement_model(series, drivers, years_out=FORECAST_YEARS_OUT)
    cashflow_forecast = fc_phase8_cashflow_forecast(projections, series[-1]["cash"])
    scenarios = fc_phase9_scenarios(series, drivers)
    latest_accounts = parsed_years[-1]["accounts"]
    analytics_latest = compute_full_analytics(latest_accounts)
    stakeholder = fc_phase10_stakeholder_analysis(series, projections, analytics_latest)
    dcf = fc_phase11_dcf_valuation(projections, discount_rate=discount_rate, terminal_growth=terminal_growth)
    risk = fc_phase12_risk_scoring(series, analytics_latest, cashflow_forecast)
    company, period = determine_company_and_period({"_": yearly_docs[-1]})
    narrative = fc_phase13_narrative(series, projections, drivers, risk, company)
    confidence = fc_phase14_confidence_scores(series, drivers)
    multiples = comparable_multiples or {"revenue_multiple": 3.0, "ebitda_multiple": 8.0}

    return ForecastContext(
        company=company, period_label=period, parsed_years=parsed_years, series=series, drivers=drivers,
        projections=projections, cashflow_forecast=cashflow_forecast, scenarios=scenarios, stakeholder=stakeholder,
        dcf=dcf, risk=risk, narrative=narrative, confidence=confidence, analytics_latest=analytics_latest,
        operational_inputs=operational_inputs, discount_rate=discount_rate, terminal_growth=terminal_growth,
        comparable_multiples=multiples, validation_warnings=warnings,
    )


# ---- Base engine class ---------------------------------------------------------------

class ForecastEngine(ABC):
    """Base class every analysis engine implements. run() is the fixed orchestration
    pipeline: Detection Logic -> Data Extraction -> Validation Rules -> Financial
    Calculations -> KPI Generation -> Executive Insights -> Risk Detection ->
    Recommendations -> Final Report Generator (EngineReport)."""

    name: str = "BaseEngine"

    @abstractmethod
    def detect(self, ctx: ForecastContext) -> Tuple[bool, float, str]:
        """Returns (applicable, confidence 0-1, reasoning)."""
        ...

    @abstractmethod
    def extract(self, ctx: ForecastContext) -> Dict[str, Any]: ...

    @abstractmethod
    def validate(self, extracted: Dict[str, Any]) -> List[str]:
        """Returns a list of validation warning strings."""
        ...

    @abstractmethod
    def calculate(self, extracted: Dict[str, Any]) -> Dict[str, Any]: ...

    @abstractmethod
    def generate_kpis(self, calc: Dict[str, Any]) -> List[EngineKPI]: ...

    @abstractmethod
    def generate_insights(self, calc: Dict[str, Any], kpis: List[EngineKPI]) -> List[str]: ...

    @abstractmethod
    def detect_risks(self, calc: Dict[str, Any]) -> List[RiskItem]: ...

    @abstractmethod
    def generate_recommendations(self, calc: Dict[str, Any], risks: List[RiskItem]) -> List[str]: ...

    def build_summary(self, calc: Dict[str, Any], kpis: List[EngineKPI]) -> str:
        return f"{self.name} analysis complete."

    def run(self, ctx: ForecastContext) -> EngineReport:
        applicable, confidence, reason = self.detect(ctx)
        if not applicable:
            return EngineReport(engine_name=self.name, applicable=False, confidence=confidence, summary=reason)
        extracted = self.extract(ctx)
        warnings = self.validate(extracted)
        calc = self.calculate(extracted)
        calc["_validation_warnings"] = warnings
        kpis = self.generate_kpis(calc)
        insights = self.generate_insights(calc, kpis)
        risks = self.detect_risks(calc)
        recommendations = self.generate_recommendations(calc, risks)
        summary = self.build_summary(calc, kpis)
        return EngineReport(engine_name=self.name, applicable=True, confidence=confidence, summary=summary,
                             kpis=kpis, insights=insights, risks=risks, recommendations=recommendations, data=calc)


# ---- Engine 1: Growth Rate Analysis --------------------------------------------------

class GrowthRateAnalysisEngine(ForecastEngine):
    name = "Growth Rate Analysis"

    def detect(self, ctx):
        n = len(ctx.series)
        if n < 2:
            return False, 0.0, "Fewer than 2 fiscal years available — growth trend analysis requires at least 2 periods."
        return True, (0.95 if n >= 3 else 0.6), f"{n} fiscal years of revenue history detected — sufficient for growth trend analysis."

    def extract(self, ctx):
        return {"years": [s["year"] for s in ctx.series], "revenues": [s["revenue"] for s in ctx.series]}

    def validate(self, extracted):
        warnings = []
        if any(r <= 0 for r in extracted["revenues"]):
            warnings.append("One or more periods show zero or negative revenue — growth rates for those periods are not meaningful.")
        return warnings

    def calculate(self, extracted):
        revs, years = extracted["revenues"], extracted["years"]
        yoy = [_safe_div(revs[i] - revs[i - 1], revs[i - 1]) for i in range(1, len(revs))]
        cagr = _cagr(revs[0], revs[-1], len(revs) - 1)
        valid_yoy = [g for g in yoy if g is not None]
        avg_growth = statistics.mean(valid_yoy) if valid_yoy else None
        volatility = statistics.pstdev(valid_yoy) if len(valid_yoy) > 1 else 0.0
        accel = [yoy[i] - yoy[i - 1] for i in range(1, len(yoy)) if yoy[i] is not None and yoy[i - 1] is not None]
        trend = "insufficient data"
        if accel:
            trend = "accelerating" if accel[-1] > 0.01 else ("decelerating" if accel[-1] < -0.01 else "stable")
        return {"years": years, "revenues": revs, "yoy_growth": yoy, "cagr": cagr, "avg_growth": avg_growth,
                "volatility": volatility, "acceleration": accel, "trend": trend}

    def generate_kpis(self, calc):
        return [
            EngineKPI(name="CAGR", value=calc["cagr"], unit="%"),
            EngineKPI(name="Average YoY Growth", value=calc["avg_growth"], unit="%"),
            EngineKPI(name="Latest YoY Growth", value=(calc["yoy_growth"][-1] if calc["yoy_growth"] else None), unit="%"),
            EngineKPI(name="Growth Volatility (stdev)", value=calc["volatility"], unit="%"),
        ]

    def generate_insights(self, calc, kpis):
        insights = []
        if calc["cagr"] is not None:
            insights.append(f"Revenue grew at a {calc['cagr']*100:.1f}% CAGR across the analyzed periods.")
        insights.append(f"The growth trend is currently classified as {calc['trend']}.")
        if calc["volatility"] and calc["volatility"] > 0.25:
            insights.append("Growth has been volatile year-over-year, making single-point forecasts less reliable.")
        return insights

    def detect_risks(self, calc):
        risks = []
        if calc["trend"] == "decelerating":
            risks.append(RiskItem(category="Growth Risk", description="Revenue growth is decelerating year-over-year.",
                                   probability=0.6, impact=0.5, severity=RiskSeverity.MODERATE,
                                   mitigation="Investigate demand-side drivers (pricing, churn, market saturation) behind the slowdown."))
        if calc["volatility"] and calc["volatility"] > 0.35:
            risks.append(RiskItem(category="Growth Risk", description="High growth volatility reduces forecast reliability.",
                                   probability=0.5, impact=0.4, severity=RiskSeverity.MODERATE,
                                   mitigation="Use scenario ranges rather than point estimates when planning around this revenue line."))
        return risks

    def generate_recommendations(self, calc, risks):
        recs = ["Track growth drivers (pricing, volume, retention) explicitly rather than relying on top-line trend alone."]
        if calc["trend"] == "decelerating":
            recs.append("Reassess growth assumptions used elsewhere in the forecast (cash flow, valuation) given the decelerating trend.")
        return recs

    def build_summary(self, calc, kpis):
        cagr_txt = f"{calc['cagr']*100:.1f}% CAGR" if calc["cagr"] is not None else "an undetermined CAGR"
        return f"Revenue grew at {cagr_txt} over {len(calc['years'])} periods; growth trend is {calc['trend']}."


# ---- Engine 2: Driver-Based Forecast --------------------------------------------------

class DriverBasedForecastEngine(ForecastEngine):
    name = "Driver-Based Forecast"

    def detect(self, ctx):
        has_ops = bool(ctx.operational_inputs)
        reason = ("Operational driver data (customers/units/price) supplied — full driver decomposition available."
                   if has_ops else
                   "No operational driver data supplied — falling back to cost-structure driver ratios (COGS %, "
                   "Opex % of revenue). Supply per-year customers/units/price via operational_inputs for a full "
                   "revenue-driver decomposition.")
        return True, (0.9 if has_ops else 0.45), reason

    def extract(self, ctx):
        return {"series": ctx.series, "operational_inputs": ctx.operational_inputs or []}

    def validate(self, extracted):
        return [f"Year {op.get('year')} reports zero customers — ARPU is undefined for that period."
                for op in extracted["operational_inputs"] if op.get("customers") == 0]

    def calculate(self, extracted):
        series, ops = extracted["series"], extracted["operational_inputs"]
        result: Dict[str, Any] = {"has_operational_data": bool(ops)}
        if ops:
            ops_sorted = sorted(ops, key=lambda o: o.get("year", 0))
            arpu, customer_growth = [], []
            for i, o in enumerate(ops_sorted):
                customers = o.get("customers")
                rev = next((s["revenue"] for s in series if s["year"] == o.get("year")), None)
                if rev is not None and customers:
                    arpu.append({"year": o.get("year"), "arpu": _safe_div(rev, customers)})
                if i > 0 and ops_sorted[i - 1].get("customers"):
                    customer_growth.append(_safe_div(customers - ops_sorted[i - 1].get("customers"), ops_sorted[i - 1].get("customers")))
            result.update({"arpu_series": arpu, "customer_growth": customer_growth})
        else:
            result.update({
                "cogs_ratio_trend": [_safe_div(s["cogs"], s["revenue"]) for s in series],
                "opex_ratio_trend": [_safe_div(s["operating_expenses"], s["revenue"]) for s in series],
            })
        return result

    def generate_kpis(self, calc):
        kpis = []
        if calc["has_operational_data"]:
            arpu_vals = [a["arpu"] for a in calc["arpu_series"] if a["arpu"] is not None]
            if arpu_vals:
                kpis.append(EngineKPI(name="Latest ARPU", value=arpu_vals[-1], unit="$"))
            if calc["customer_growth"]:
                kpis.append(EngineKPI(name="Latest Customer Growth", value=calc["customer_growth"][-1], unit="%"))
        else:
            valid_cogs = [c for c in calc["cogs_ratio_trend"] if c is not None]
            valid_opex = [o for o in calc["opex_ratio_trend"] if o is not None]
            if valid_cogs:
                kpis.append(EngineKPI(name="Latest COGS % of Revenue", value=valid_cogs[-1], unit="%"))
            if valid_opex:
                kpis.append(EngineKPI(name="Latest Opex % of Revenue", value=valid_opex[-1], unit="%"))
        return kpis

    def generate_insights(self, calc, kpis):
        if calc["has_operational_data"]:
            return ["Revenue growth can be decomposed into customer growth and ARPU trends using the supplied operational data."]
        return ["No operational driver data was supplied; this analysis substitutes cost-structure ratios as a proxy "
                "driver view. Provide customers/units/price per year for a full ARPU/volume-based decomposition."]

    def detect_risks(self, calc):
        if calc["has_operational_data"]:
            return []
        return [RiskItem(category="Data Risk", description="Revenue cannot be traced to operational drivers without customer/unit/price data.",
                          probability=0.7, impact=0.3, severity=RiskSeverity.MODERATE,
                          mitigation="Capture customers, units sold, and pricing per period to enable true driver-based forecasting.")]

    def generate_recommendations(self, calc, risks):
        if calc["has_operational_data"]:
            return ["Continue tracking ARPU and customer growth separately — a slowdown in either should trigger a forecast re-check."]
        return ["Supply per-period customer/unit/price data to unlock full driver-based revenue forecasting."]

    def build_summary(self, calc, kpis):
        return ("Driver-based forecast built from supplied operational data." if calc["has_operational_data"]
                else "Driver-based forecast approximated via cost-structure ratios (no operational driver data supplied).")


# ---- Engine 3: Cash Flow Forecast ------------------------------------------------------

class CashFlowForecastEngine(ForecastEngine):
    name = "Cash Flow Forecast"

    def detect(self, ctx):
        return True, 0.9, "Cash balances and projected cash flows are available from the trial balance series."

    def extract(self, ctx):
        return {"cashflow_forecast": ctx.cashflow_forecast, "current_cash": ctx.series[-1]["cash"]}

    def validate(self, extracted):
        return ["Latest reported cash balance is zero or negative."] if extracted["current_cash"] <= 0 else []

    def calculate(self, extracted):
        cf = extracted["cashflow_forecast"]
        negative_years = [c["year_offset"] for c in cf if c["net_change_in_cash"] < 0]
        runways = [c["runway_months"] for c in cf if c["runway_months"] is not None]
        return {"cashflow_forecast": cf, "negative_years": negative_years,
                "min_runway_months": min(runways) if runways else None,
                "ending_cash_final": cf[-1]["ending_cash"] if cf else None, "current_cash": extracted["current_cash"]}

    def generate_kpis(self, calc):
        kpis = [EngineKPI(name="Current Cash", value=calc["current_cash"], unit="$"),
                EngineKPI(name=f"Projected Cash (Year +{FORECAST_YEARS_OUT})", value=calc["ending_cash_final"], unit="$")]
        if calc["min_runway_months"] is not None:
            kpis.append(EngineKPI(name="Minimum Projected Runway", value=calc["min_runway_months"], unit="months"))
        return kpis

    def generate_insights(self, calc, kpis):
        if calc["negative_years"]:
            return [f"Cash is projected to decline in year(s) {calc['negative_years']} of the forecast."]
        return ["Cash is projected to grow in every forecast year under base-case assumptions."]

    def detect_risks(self, calc):
        risks = []
        if calc["min_runway_months"] is not None and calc["min_runway_months"] < 12:
            risks.append(RiskItem(category="Liquidity Risk",
                                   description=f"Projected runway falls below 12 months ({calc['min_runway_months']:.1f} months).",
                                   probability=0.6, impact=0.8, severity=RiskSeverity.HIGH,
                                   mitigation="Plan a funding round or cost reduction well ahead of the projected cash shortfall."))
        if calc["current_cash"] <= 0:
            risks.append(RiskItem(category="Liquidity Risk", description="Current cash balance is zero or negative.",
                                   probability=0.9, impact=0.9, severity=RiskSeverity.CRITICAL,
                                   mitigation="Immediate liquidity action required — secure bridge financing or reduce burn."))
        return risks

    def generate_recommendations(self, calc, risks):
        recs = []
        if any(r.severity in (RiskSeverity.HIGH, RiskSeverity.CRITICAL) for r in risks):
            recs.append("Begin fundraising or cost-reduction planning now, given the projected liquidity risk.")
        recs.append("Track actual vs. forecast cash monthly; this model uses simplified capex/working-capital assumptions.")
        return recs

    def build_summary(self, calc, kpis):
        if calc["ending_cash_final"] is None:
            return "Cash flow forecast could not be computed."
        return f"Cash is projected to move from {calc['current_cash']:,.0f} to {calc['ending_cash_final']:,.0f} over {FORECAST_YEARS_OUT} years."


# ---- Engine 4: Stakeholder Analysis -----------------------------------------------------

class StakeholderAnalysisEngine(ForecastEngine):
    name = "Stakeholder Analysis"

    def detect(self, ctx):
        return True, 0.85, "Sufficient forecast data to generate stakeholder-specific commentary."

    def extract(self, ctx):
        return {"stakeholder": ctx.stakeholder, "analytics_latest": ctx.analytics_latest,
                "series": ctx.series, "projections": ctx.projections}

    def validate(self, extracted):
        return []

    def calculate(self, extracted):
        last, y5 = extracted["series"][-1], extracted["projections"][-1]
        liquidity_impact = "positive" if y5["cash"] > last["cash"] else "negative"
        nm_last = extracted["analytics_latest"].get("net_margin")
        profitability_impact = "positive" if (nm_last or 0) > 0 else "negative"
        de = extracted["analytics_latest"].get("debt_to_equity")
        funding_impact = "elevated" if (de or 0) > 2.0 else "manageable"
        return {"narratives": extracted["stakeholder"], "liquidity_impact": liquidity_impact,
                "profitability_impact": profitability_impact, "funding_impact": funding_impact}

    def generate_kpis(self, calc):
        return []

    def generate_insights(self, calc, kpis):
        insights = [f"Liquidity impact on stakeholders is {calc['liquidity_impact']}.",
                    f"Profitability impact is {calc['profitability_impact']}.",
                    f"Funding-related risk exposure is {calc['funding_impact']}."]
        insights.extend(calc["narratives"].values())
        return insights

    def detect_risks(self, calc):
        if calc["funding_impact"] == "elevated":
            return [RiskItem(category="Stakeholder Risk", description="Elevated leverage increases risk exposure for lenders and investors.",
                              probability=0.5, impact=0.5, severity=RiskSeverity.MODERATE,
                              mitigation="Communicate a clear deleveraging or equity-funding plan to bank/investor stakeholders.")]
        return []

    def generate_recommendations(self, calc, risks):
        return ["Share role-specific summaries (owner/CFO/investor/bank/auditor) rather than a single generic report."]

    def build_summary(self, calc, kpis):
        return (f"Liquidity impact: {calc['liquidity_impact']} · Profitability impact: {calc['profitability_impact']} "
                f"· Funding risk: {calc['funding_impact']}.")

# ---- Engine 5: Valuation Model -----------------------------------------------------------

class ValuationModelEngine(ForecastEngine):
    name = "Valuation Model"

    def detect(self, ctx):
        return True, 0.8, "DCF and comparable-multiple valuation computed from projected financials."

    def extract(self, ctx):
        return {"dcf": ctx.dcf, "series": ctx.series, "multiples": ctx.comparable_multiples}

    def validate(self, extracted):
        if extracted["dcf"]["equity_value"] is not None and extracted["dcf"]["equity_value"] < 0:
            return ["DCF-implied equity value is negative — net debt exceeds enterprise value."]
        return []

    def calculate(self, extracted):
        last = extracted["series"][-1]
        revenue_multiple_val = last["revenue"] * extracted["multiples"].get("revenue_multiple", 3.0)
        ebitda = last["ebit"] + last["depreciation_amortization"]
        ebitda_multiple_val = ebitda * extracted["multiples"].get("ebitda_multiple", 8.0) if ebitda > 0 else None
        candidates = [v for v in [extracted["dcf"]["enterprise_value"], revenue_multiple_val, ebitda_multiple_val] if v is not None]
        val_range = (min(candidates), max(candidates)) if candidates else (None, None)
        return {"dcf": extracted["dcf"], "revenue_multiple_valuation": revenue_multiple_val,
                "ebitda_multiple_valuation": ebitda_multiple_val, "valuation_range": val_range,
                "multiples_used": extracted["multiples"]}

    def generate_kpis(self, calc):
        kpis = [EngineKPI(name="DCF Enterprise Value", value=calc["dcf"]["enterprise_value"], unit="$"),
                EngineKPI(name="DCF Equity Value", value=calc["dcf"]["equity_value"], unit="$"),
                EngineKPI(name="Revenue-Multiple Valuation", value=calc["revenue_multiple_valuation"], unit="$")]
        if calc["ebitda_multiple_valuation"] is not None:
            kpis.append(EngineKPI(name="EBITDA-Multiple Valuation", value=calc["ebitda_multiple_valuation"], unit="$"))
        return kpis

    def generate_insights(self, calc, kpis):
        lo, hi = calc["valuation_range"]
        insights = []
        if lo is not None:
            insights.append(f"Valuation estimates across methods range from {lo:,.0f} to {hi:,.0f}.")
        insights.append(f"Comparable multiples used: {calc['multiples_used'].get('revenue_multiple')}x revenue, "
                         f"{calc['multiples_used'].get('ebitda_multiple')}x EBITDA (illustrative defaults unless overridden).")
        return insights

    def detect_risks(self, calc):
        if calc["dcf"]["equity_value"] is not None and calc["dcf"]["equity_value"] < 0:
            return [RiskItem(category="Valuation Risk", description="DCF-implied equity value is negative.",
                              probability=0.4, impact=0.7, severity=RiskSeverity.ELEVATED,
                              mitigation="Revisit leverage and growth assumptions; negative equity value under DCF often "
                                         "signals excess debt or overly conservative growth.")]
        return []

    def generate_recommendations(self, calc, risks):
        return ["Treat comparable multiples as illustrative until replaced with real, sector-specific comparable transactions.",
                "Present a valuation range rather than a single point estimate to investors."]

    def build_summary(self, calc, kpis):
        lo, hi = calc["valuation_range"]
        if lo is None:
            return "Valuation could not be computed."
        return f"DCF enterprise value: {calc['dcf']['enterprise_value']:,.0f}; valuation range {lo:,.0f}–{hi:,.0f}."


# ---- Engine 6: Risk Analysis --------------------------------------------------------------

class RiskAnalysisEngine(ForecastEngine):
    name = "Risk Analysis"

    def detect(self, ctx):
        return True, 0.9, "Composite risk scoring computed across liquidity, leverage, growth, and data-quality dimensions."

    def extract(self, ctx):
        return {"risk": ctx.risk}

    def validate(self, extracted):
        return []

    def calculate(self, extracted):
        return {"risk": extracted["risk"]}

    def generate_kpis(self, calc):
        kpis = [EngineKPI(name=k.replace("_", " ").title(), value=v, unit="score/100")
                for k, v in calc["risk"]["dimensions"].items()]
        kpis.append(EngineKPI(name="Overall Risk Score", value=calc["risk"]["overall_score"], unit="score/100"))
        return kpis

    def generate_insights(self, calc, kpis):
        return [f"Overall risk band: {calc['risk']['risk_band']}."]

    def detect_risks(self, calc):
        risks = []
        for dim, score in calc["risk"]["dimensions"].items():
            if score >= 40:
                risks.append(RiskItem(category=dim.replace("_", " ").title(),
                                       description=f"{dim.replace('_', ' ').title()} scored {score}/100.",
                                       probability=min(score / 100, 0.95), impact=min(score / 100, 0.95),
                                       severity=_severity_from_score(score), mitigation=_risk_mitigation_hint(dim)))
        return risks

    def generate_recommendations(self, calc, risks):
        if not risks:
            return ["No elevated risk dimensions detected; continue standard periodic monitoring."]
        return [f"Prioritize mitigation for: {', '.join(sorted({r.category for r in risks}))}."]

    def build_summary(self, calc, kpis):
        return f"Overall risk score {calc['risk']['overall_score']}/100 ({calc['risk']['risk_band']})."


# ---- Engine 7: Three-Statement Forecast ----------------------------------------------------

class ThreeStatementForecastEngine(ForecastEngine):
    name = "Three Statement Forecast"

    def detect(self, ctx):
        return True, 0.85, "Linked income statement, balance sheet, and cash flow projections available."

    def extract(self, ctx):
        return {"projections": ctx.projections, "cashflow_forecast": ctx.cashflow_forecast}

    def validate(self, extracted):
        warnings = []
        for p in extracted["projections"]:
            implied_bs = p["total_assets"]
            implied_le = p["total_liabilities"] + p["total_equity"]
            diff = abs(implied_bs - implied_le)
            if diff > max(implied_bs, 1) * 0.01:
                warnings.append(f"Year +{p['year_offset']}: projected balance sheet does not balance within 1% (diff {diff:,.0f}).")
        return warnings

    def calculate(self, extracted):
        return extracted

    def generate_kpis(self, calc):
        p5 = calc["projections"][-1]
        return [EngineKPI(name="Year+5 Revenue", value=p5["revenue"], unit="$"),
                EngineKPI(name="Year+5 Net Income", value=p5["net_income"], unit="$"),
                EngineKPI(name="Year+5 Total Assets", value=p5["total_assets"], unit="$")]

    def generate_insights(self, calc, kpis):
        return ["The three-statement model links revenue growth assumptions through to a simplified balance sheet and "
                "cash flow projection; balance sheet items scale off revenue growth as a modeling simplification "
                "rather than a full working-capital build."]

    def detect_risks(self, calc):
        return [RiskItem(category="Accounting Consistency", description=w, probability=0.5, impact=0.4,
                          severity=RiskSeverity.MODERATE,
                          mitigation="Review the simplified balance sheet scaling assumptions for that period.")
                for w in calc.get("_validation_warnings", [])]

    def generate_recommendations(self, calc, risks):
        return ["Replace the simplified asset/liability scaling assumption with a full working-capital and capex "
                "schedule for investor-grade accuracy."]

    def build_summary(self, calc, kpis):
        n_issues = len(calc.get("_validation_warnings", []))
        return (f"Three-statement forecast generated; {n_issues} balance-sheet tie-out issue(s) flagged." if n_issues
                else "Three-statement forecast generated; all projected years balance within tolerance.")


# ---- Engine 8: Scenario Analysis -----------------------------------------------------------

class ScenarioAnalysisEngine(ForecastEngine):
    name = "Scenario Analysis"

    def detect(self, ctx):
        return True, 0.9, "Base/Best/Worst case scenarios computed by flexing growth and expense assumptions."

    def extract(self, ctx):
        return {"scenarios": ctx.scenarios}

    def validate(self, extracted):
        return []

    def calculate(self, extracted):
        scen = extracted["scenarios"]
        base_rev = scen["base"]["projections"][-1]["revenue"]
        variance = {name: _safe_div(s["projections"][-1]["revenue"] - base_rev, base_rev) for name, s in scen.items()}
        return {"scenarios": scen, "variance_vs_base": variance}

    def generate_kpis(self, calc):
        return [EngineKPI(name=f"{name.title()} Case Year+{FORECAST_YEARS_OUT} Revenue",
                           value=s["projections"][-1]["revenue"], unit="$") for name, s in calc["scenarios"].items()]

    def generate_insights(self, calc, kpis):
        v = calc["variance_vs_base"]
        return [f"Best case revenue is {v['best']*100:+.1f}% vs base; worst case is {v['worst']*100:+.1f}% vs base."]

    def detect_risks(self, calc):
        v = calc["variance_vs_base"]
        if v["worst"] < -0.25:
            return [RiskItem(category="Scenario Risk", description=f"Worst-case scenario is {v['worst']*100:.1f}% below base case.",
                              probability=0.25, impact=0.7, severity=RiskSeverity.ELEVATED,
                              mitigation="Stress-test operating plans and covenants against the worst-case scenario.")]
        return []

    def generate_recommendations(self, calc, risks):
        return ["Use the worst-case scenario, not the base case, when setting minimum cash covenants or credit facility sizing."]

    def build_summary(self, calc, kpis):
        return "Base/Best/Worst scenario set generated with revenue variance quantified vs. base case."


# ---- Engine 9: Sensitivity Analysis ---------------------------------------------------------

class SensitivityAnalysisEngine(ForecastEngine):
    name = "Sensitivity Analysis"

    def detect(self, ctx):
        return True, 0.85, "Tornado sensitivity computed by flexing key assumptions ±10% (±500bps for discount rate)."

    def extract(self, ctx):
        return {"series": ctx.series, "drivers": ctx.drivers, "discount_rate": ctx.discount_rate,
                "terminal_growth": ctx.terminal_growth}

    def validate(self, extracted):
        return []

    def calculate(self, extracted):
        series, base_drivers = extracted["series"], extracted["drivers"]
        base_projections = fc_phase7_three_statement_model(series, base_drivers, years_out=FORECAST_YEARS_OUT)
        base_ni = base_projections[-1]["net_income"]
        base_dcf = fc_phase11_dcf_valuation(base_projections, extracted["discount_rate"], extracted["terminal_growth"])
        base_ev = base_dcf["enterprise_value"]

        def flex_and_measure(driver_key, delta, rate_delta=0.0, term_delta=0.0):
            flexed = dict(base_drivers)
            if driver_key:
                flexed[driver_key] = flexed[driver_key] + delta
            proj = fc_phase7_three_statement_model(series, flexed, years_out=FORECAST_YEARS_OUT)
            ni = proj[-1]["net_income"]
            dcf = fc_phase11_dcf_valuation(proj, extracted["discount_rate"] + rate_delta, extracted["terminal_growth"] + term_delta)
            return ni, dcf["enterprise_value"]

        variables = []
        for label, key, delta in [
            ("Revenue Growth Rate +10%", "chosen_growth_rate", base_drivers["chosen_growth_rate"] * 0.10),
            ("Revenue Growth Rate -10%", "chosen_growth_rate", -base_drivers["chosen_growth_rate"] * 0.10),
            ("COGS Ratio +10%", "avg_cogs_ratio", base_drivers["avg_cogs_ratio"] * 0.10),
            ("COGS Ratio -10%", "avg_cogs_ratio", -base_drivers["avg_cogs_ratio"] * 0.10),
            ("Opex Ratio +10%", "avg_expense_ratio", base_drivers["avg_expense_ratio"] * 0.10),
            ("Opex Ratio -10%", "avg_expense_ratio", -base_drivers["avg_expense_ratio"] * 0.10),
        ]:
            ni, ev = flex_and_measure(key, delta)
            variables.append({"variable": label, "net_income_delta": ni - base_ni,
                               "enterprise_value_delta": (ev - base_ev) if (ev is not None and base_ev is not None) else None})

        for label, rate_delta in [("Discount Rate +500bps", 0.05), ("Discount Rate -500bps", -0.05)]:
            ni, ev = flex_and_measure(None, 0.0, rate_delta=rate_delta)
            variables.append({"variable": label, "net_income_delta": 0.0,
                               "enterprise_value_delta": (ev - base_ev) if (ev is not None and base_ev is not None) else None})

        variables.sort(key=lambda v: abs(v["enterprise_value_delta"] or v["net_income_delta"] or 0), reverse=True)
        return {"base_net_income": base_ni, "base_enterprise_value": base_ev, "tornado": variables}

    def generate_kpis(self, calc):
        return [EngineKPI(name="Base Net Income (Yr+5)", value=calc["base_net_income"], unit="$"),
                EngineKPI(name="Base Enterprise Value", value=calc["base_enterprise_value"], unit="$")]

    def generate_insights(self, calc, kpis):
        insights = []
        for v in calc["tornado"][:3]:
            if v["enterprise_value_delta"] is not None:
                insights.append(f"{v['variable']} has a large impact, moving enterprise value by {v['enterprise_value_delta']:,.0f}.")
            else:
                insights.append(f"{v['variable']} moves Year+5 net income by {v['net_income_delta']:,.0f}.")
        return insights

    def detect_risks(self, calc):
        return []

    def generate_recommendations(self, calc, risks):
        if calc["tornado"]:
            return [f"Focus assumption validation effort on '{calc['tornado'][0]['variable']}' — it has the largest valuation impact."]
        return []

    def build_summary(self, calc, kpis):
        return f"Tornado sensitivity computed across {len(calc['tornado'])} variables."


# ---- Engine 10: Profitability Analysis --------------------------------------------------------

class ProfitabilityAnalysisEngine(ForecastEngine):
    name = "Profitability Analysis"

    def detect(self, ctx):
        return True, 0.95, "Margin and return ratios computed from the latest period's financials."

    def extract(self, ctx):
        return {"analytics_latest": ctx.analytics_latest, "series": ctx.series}

    def validate(self, extracted):
        return []

    def calculate(self, extracted):
        a = extracted["analytics_latest"]
        margins_trend = [_safe_div(s["net_income"], s["revenue"]) for s in extracted["series"]]
        return {"analytics": a, "net_margin_trend": margins_trend}

    def generate_kpis(self, calc):
        a = calc["analytics"]
        return [EngineKPI(name="Gross Margin", value=a.get("gross_margin"), unit="%"),
                EngineKPI(name="Operating Margin", value=a.get("operating_margin"), unit="%"),
                EngineKPI(name="Net Margin", value=a.get("net_margin"), unit="%"),
                EngineKPI(name="ROE", value=a.get("roe"), unit="%"),
                EngineKPI(name="ROA", value=a.get("roa"), unit="%")]

    def generate_insights(self, calc, kpis):
        valid = [t for t in calc["net_margin_trend"] if t is not None]
        if len(valid) >= 2:
            direction = "improving" if valid[-1] > valid[0] else "declining"
            return [f"Net margin has been {direction} across the analyzed periods."]
        return []

    def detect_risks(self, calc):
        nm = calc["analytics"].get("net_margin")
        if nm is not None and nm < 0:
            return [RiskItem(category="Profitability Risk", description="Net margin is negative in the latest period.",
                              probability=0.7, impact=0.6, severity=RiskSeverity.ELEVATED,
                              mitigation="Identify the largest cost drivers relative to revenue and target margin-improvement actions.")]
        return []

    def generate_recommendations(self, calc, risks):
        return ["Benchmark margins against sector peers to contextualize whether current profitability is competitive."]

    def build_summary(self, calc, kpis):
        a = calc["analytics"]
        if a.get("net_margin") is not None and a.get("gross_margin") is not None:
            return f"Net margin {a['net_margin']*100:.1f}%, gross margin {a['gross_margin']*100:.1f}%."
        return "Profitability metrics computed."


# ---- Engine 11: Forecast Quality Engine -----------------------------------------------------

class ForecastQualityEngine(ForecastEngine):
    name = "Forecast Quality Engine"

    def detect(self, ctx):
        return True, 1.0, "Automated data-quality and forecast-integrity checks run against the uploaded model."

    def extract(self, ctx):
        return {"series": ctx.series, "confidence": ctx.confidence, "validation_warnings": ctx.validation_warnings,
                "cashflow_forecast": ctx.cashflow_forecast, "risk": ctx.risk}

    def validate(self, extracted):
        return list(extracted["validation_warnings"])

    def calculate(self, extracted):
        n_years = len(extracted["series"])
        completeness_score = min(100, n_years / 5 * 100)
        negative_cash_years = sum(1 for c in extracted["cashflow_forecast"] if c["ending_cash"] < 0)
        accuracy_score = max(0, 100 - negative_cash_years * 20 - len(extracted["validation_warnings"]) * 10)
        revs = [s["revenue"] for s in extracted["series"]]
        volatility = statistics.pstdev(revs) / statistics.mean(revs) if len(revs) > 1 and statistics.mean(revs) else 0
        reliability_score = max(0, 100 - volatility * 100)
        investor_readiness_score = round(statistics.mean(
            [completeness_score, accuracy_score, reliability_score, 100 - extracted["risk"]["overall_score"]]), 1)
        return {"completeness_score": round(completeness_score, 1), "accuracy_score": round(accuracy_score, 1),
                "reliability_score": round(reliability_score, 1), "investor_readiness_score": investor_readiness_score,
                "confidence": extracted["confidence"], "negative_cash_years": negative_cash_years}

    def generate_kpis(self, calc):
        return [EngineKPI(name="Completeness Score", value=calc["completeness_score"], unit="score/100"),
                EngineKPI(name="Accuracy Score", value=calc["accuracy_score"], unit="score/100"),
                EngineKPI(name="Reliability Score", value=calc["reliability_score"], unit="score/100"),
                EngineKPI(name="Investor Readiness Score", value=calc["investor_readiness_score"], unit="score/100")]

    def generate_insights(self, calc, kpis):
        return [f"Revenue forecast confidence: {calc['confidence']['revenue']}; "
                f"margin forecast confidence: {calc['confidence']['margins']}."]

    def detect_risks(self, calc):
        risks = []
        if calc["completeness_score"] < 60:
            risks.append(RiskItem(category="Data Quality Risk", description="Fewer than 3 fiscal years of history limits forecast reliability.",
                                   probability=0.8, impact=0.5, severity=RiskSeverity.MODERATE,
                                   mitigation="Provide additional historical trial balances to improve statistical confidence."))
        if calc["negative_cash_years"] > 0:
            risks.append(RiskItem(category="Forecast Integrity Risk",
                                   description=f"{calc['negative_cash_years']} forecast year(s) show negative ending cash.",
                                   probability=0.6, impact=0.7, severity=RiskSeverity.HIGH,
                                   mitigation="Revisit growth/expense assumptions or plan financing before the projected shortfall."))
        return risks

    def generate_recommendations(self, calc, risks):
        if calc["investor_readiness_score"] < 60:
            return ["Address the data-quality and integrity issues above before presenting this model to investors."]
        return ["Model quality is investor-ready; keep historical data current each period to maintain confidence."]

    def build_summary(self, calc, kpis):
        return f"Investor readiness score: {calc['investor_readiness_score']}/100."


# ---- Document Classification Engine ----------------------------------------------------------

MODEL_TYPE_KEYWORDS: Dict[str, List[str]] = {
    "Revenue Forecast": ["revenue", "sales forecast"],
    "Growth Rate Analysis": ["growth", "cagr", "yoy"],
    "Driver-Based Forecast": ["customer", "unit", "price", "arpu", "churn"],
    "Cash Flow Forecast": ["cash flow", "cash balance", "runway", "burn"],
    "Operating Expense Forecast": ["operating expense", "opex", "payroll", "salary"],
    "Budget Forecast": ["budget"],
    "Working Capital Forecast": ["working capital", "receivable", "payable"],
    "Three Statement Forecast": ["balance sheet", "income statement", "cash flow statement"],
    "Income Statement Forecast": ["income statement", "revenue", "expense"],
    "Balance Sheet Forecast": ["balance sheet", "asset", "liability", "equity"],
    "Startup Financial Model": ["burn", "runway", "seed", "series a"],
    "DCF Valuation": ["discounted cash flow", "dcf", "terminal value", "discount rate"],
    "Comparable Company Analysis": ["comparable", "multiple", "ebitda multiple"],
    "Enterprise Valuation": ["enterprise value", "equity value"],
    "Risk Analysis": ["risk"],
    "Scenario Analysis": ["scenario", "best case", "worst case", "base case"],
    "Sensitivity Analysis": ["sensitivity", "tornado"],
    "Burn Rate / Runway Analysis": ["burn rate", "runway"],
    "Stakeholder Analysis": ["stakeholder", "investor", "board"],
    "Profitability Analysis": ["margin", "profitability", "roe", "roa"],
    "SaaS Financial Model": ["mrr", "arr", "churn", "saas"],
}


class DocumentClassificationEngine:
    """Identifies every applicable financial model type present in the uploaded documents,
    using both keyword signals from account names/period labels and structural evidence
    (which financial-statement sections were actually extracted)."""

    def classify(self, ctx: ForecastContext) -> List[DetectionResult]:
        corpus_parts: List[str] = []
        for py in ctx.parsed_years:
            corpus_parts.append(py.get("label", ""))
            for a in py["accounts"]:
                corpus_parts.append(a.get("account_name", ""))
        corpus = " ".join(corpus_parts).lower()

        detected_sections = []
        if any(s["revenue"] for s in ctx.series): detected_sections.append("Revenue Section")
        if any(s["operating_expenses"] for s in ctx.series): detected_sections.append("Operating Expenses Section")
        if any(s["total_assets"] for s in ctx.series): detected_sections.append("Assets Section")
        if any(s["total_liabilities"] for s in ctx.series): detected_sections.append("Liabilities Section")
        if any(s["total_equity"] for s in ctx.series): detected_sections.append("Equity Section")

        detected_statements = []
        if "Revenue Section" in detected_sections and "Operating Expenses Section" in detected_sections:
            detected_statements.append("Income Statement")
        if "Assets Section" in detected_sections and "Liabilities Section" in detected_sections:
            detected_statements.append("Balance Sheet")
        if len(ctx.series) >= 2:
            detected_statements.append("Cash Flow Statement (derivable from multi-period balances)")

        detected_tables = [py.get("label") or f"Year {py['year']}" for py in ctx.parsed_years]

        results = []
        for model_type, keywords in MODEL_TYPE_KEYWORDS.items():
            matches = sum(1 for kw in keywords if kw in corpus)
            score = matches / len(keywords)
            if model_type == "Three Statement Forecast" and {"Income Statement", "Balance Sheet"} <= set(detected_statements):
                score = max(score, 0.8)
            if model_type in ("Growth Rate Analysis", "Cash Flow Forecast") and len(ctx.series) >= 2:
                score = max(score, 0.5)
            if score > 0:
                reasoning = f"Matched {matches}/{len(keywords)} keyword signal(s)" + (
                    " plus structural evidence from detected financial statements." if score >= 0.5 and matches < len(keywords) else "."
                )
                results.append(DetectionResult(document_type=model_type, confidence=round(min(score, 1.0), 2),
                                                 reasoning=reasoning, detected_sections=detected_sections,
                                                 detected_tables=detected_tables, detected_financial_statements=detected_statements))
        results.sort(key=lambda r: r.confidence, reverse=True)
        return results


# ---- Enterprise Report Generator (orchestrator) -----------------------------------------------

class EnterpriseReportGenerator:
    """Orchestrates the full pipeline: build shared context -> classify document -> run every
    registered engine -> aggregate risks/recommendations -> assemble the final report and,
    optionally, a multi-tab Excel workbook. New engines are added by appending to self.engines."""

    def __init__(self):
        self.engines: List[ForecastEngine] = [
            GrowthRateAnalysisEngine(), DriverBasedForecastEngine(), CashFlowForecastEngine(),
            StakeholderAnalysisEngine(), ValuationModelEngine(), RiskAnalysisEngine(),
            ThreeStatementForecastEngine(), ScenarioAnalysisEngine(), SensitivityAnalysisEngine(),
            ProfitabilityAnalysisEngine(), ForecastQualityEngine(),
        ]
        self.classifier = DocumentClassificationEngine()

    def generate(self, yearly_docs: List[dict], operational_inputs: Optional[List[Dict[str, Any]]] = None,
                 discount_rate: float = 0.12, terminal_growth: float = 0.025,
                 comparable_multiples: Optional[Dict[str, float]] = None) -> FinalEnterpriseReport:
        ctx = build_enterprise_context(yearly_docs, operational_inputs, discount_rate, terminal_growth, comparable_multiples)
        classification = self.classifier.classify(ctx)
        engine_reports = [engine.run(ctx) for engine in self.engines]

        all_risks = [r for er in engine_reports for r in er.risks]
        severity_rank = {RiskSeverity.CRITICAL: 4, RiskSeverity.HIGH: 3, RiskSeverity.ELEVATED: 2,
                          RiskSeverity.MODERATE: 1, RiskSeverity.LOW: 0}
        all_risks_sorted = sorted(all_risks, key=lambda r: severity_rank[r.severity], reverse=True)
        # Cap at 2 risks per category so N identical yearly tie-out warnings (or similar repeats)
        # don't crowd out other risk types in the board-level top-risks view.
        category_counts: Dict[str, int] = {}
        top_risks: List[RiskItem] = []
        for r in all_risks_sorted:
            if len(top_risks) >= 8:
                break
            if category_counts.get(r.category, 0) >= 2:
                continue
            category_counts[r.category] = category_counts.get(r.category, 0) + 1
            top_risks.append(r)

        seen, strategic_recs = set(), []
        for rec in [rec for er in engine_reports for rec in er.recommendations]:
            if rec not in seen:
                seen.add(rec)
                strategic_recs.append(rec)

        return FinalEnterpriseReport(
            company=ctx.company, period=ctx.period_label, generated_at=datetime.now().isoformat(timespec="seconds"),
            executive_summary=self._build_executive_summary(ctx, engine_reports, top_risks),
            document_classification=classification, engine_reports=engine_reports, top_risks=top_risks,
            strategic_recommendations=strategic_recs[:10],
            board_level_conclusion=self._build_board_conclusion(top_risks),
            validation_warnings=ctx.validation_warnings,
        )

    def _build_executive_summary(self, ctx: ForecastContext, engine_reports: List[EngineReport],
                                  top_risks: List[RiskItem]) -> str:
        parts = [f"{ctx.company}'s financial model spans {len(ctx.series)} historical fiscal year(s) "
                 f"({ctx.series[0]['year']}–{ctx.series[-1]['year']})."]
        for engine_name in ("Growth Rate Analysis", "Cash Flow Forecast", "Forecast Quality Engine"):
            er = next((e for e in engine_reports if e.engine_name == engine_name), None)
            if er and er.applicable:
                parts.append(er.summary)
        if top_risks:
            parts.append(f"The most significant identified risk is: {top_risks[0].description}")
        return " ".join(parts)

    def _build_board_conclusion(self, top_risks: List[RiskItem]) -> str:
        critical = [r for r in top_risks if r.severity in (RiskSeverity.CRITICAL, RiskSeverity.HIGH)]
        if critical:
            return ("The board should treat this forecast as directionally useful but action-gated: "
                     f"{len(critical)} high/critical-severity risk(s) require mitigation before the plan can be "
                     "considered fundable or bankable as presented.")
        return ("The board can treat this forecast as a reasonable planning baseline, subject to the standard "
                 "caveats of model-based projections; no critical risks were identified in this review.")

    def build_workbook(self, report: FinalEnterpriseReport) -> Workbook:
        wb = _new_workbook()

        ws = wb.create_sheet("Cover", 0)
        ws.merge_cells("A1:D3")
        c = ws.cell(row=1, column=1, value=f"{report.company}\nEnterprise Forecast Report\n{report.period}")
        c.font = Font(size=16, bold=True, color=XL_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells("A5:D9")
        c2 = ws.cell(row=5, column=1, value=report.executive_summary)
        c2.font = Font(size=10, color=XL_GREY)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        _autosize_columns(ws, {1: 26, 2: 26, 3: 26, 4: 26})

        ws = wb.create_sheet("Classification")
        row = _write_statement_header(ws, report.company, "Document Classification", report.period)
        for i, h in enumerate(["Document Type", "Confidence", "Reasoning"]):
            cell = ws.cell(row=row, column=1 + i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
        row += 1
        for d in report.document_classification:
            ws.cell(row=row, column=1, value=d.document_type).font = LABEL_FONT
            ws.cell(row=row, column=2, value=d.confidence).font = LABEL_FONT
            ws.cell(row=row, column=3, value=d.reasoning).font = LABEL_FONT
            row += 1
        _autosize_columns(ws, {1: 30, 2: 12, 3: 60})

        for er in report.engine_reports:
            ws = wb.create_sheet(er.engine_name[:31])
            row = _write_statement_header(ws, report.company, er.engine_name, report.period)
            if not er.applicable:
                ws.cell(row=row, column=1, value=f"Not applicable: {er.summary}").font = LABEL_FONT
                continue
            ws.cell(row=row, column=1, value=er.summary).font = BOLD_LABEL_FONT
            row += 2
            row = _write_section_title(ws, row, "KPIs")
            for k in er.kpis:
                ws.cell(row=row, column=1, value=k.name).font = LABEL_FONT
                ws.cell(row=row, column=4, value=(k.value if k.value is not None else "N/A")).font = NUMBER_FONT
                row += 1
            row += 1
            row = _write_section_title(ws, row, "Insights")
            for ins in er.insights:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = ws.cell(row=row, column=1, value=f"• {ins}")
                cell.font = LABEL_FONT
                cell.alignment = Alignment(wrap_text=True)
                row += 1
            row += 1
            row = _write_section_title(ws, row, "Risks")
            for r in er.risks:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = ws.cell(row=row, column=1, value=f"[{r.severity.value.upper()}] {r.description} — {r.mitigation}")
                cell.font = LABEL_FONT
                cell.alignment = Alignment(wrap_text=True)
                row += 1
            row += 1
            row = _write_section_title(ws, row, "Recommendations")
            for rec in er.recommendations:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = ws.cell(row=row, column=1, value=f"• {rec}")
                cell.font = LABEL_FONT
                cell.alignment = Alignment(wrap_text=True)
                row += 1
            _autosize_columns(ws, {1: 40, 2: 14, 3: 14, 4: 18})

        ws = wb.create_sheet("Board Conclusion")
        row = _write_statement_header(ws, report.company, "Board-Level Conclusion", report.period)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=4)
        cell = ws.cell(row=row, column=1, value=report.board_level_conclusion)
        cell.font = LABEL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

        return wb


# ---- API route: full enterprise forecast report ---------------------------------------------

@app.post("/api/enterprise-forecast")
async def api_enterprise_forecast(files: List[UploadFile] = File(...),
                                   operational_inputs: Optional[str] = Form(None),
                                   discount_rate: float = Form(0.12),
                                   terminal_growth: float = Form(0.025)):
    """Runs the full 11-engine enterprise forecasting suite plus document classification
    against 2+ years of uploaded trial balances, returning a structured JSON report and a
    downloadable multi-tab Excel workbook. `operational_inputs` is an optional JSON string:
    a list of {"year": int, "customers": number, "price": number, ...} objects that unlocks
    the full Driver-Based Forecast decomposition."""
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="At least 2 fiscal years of trial balances are required for enterprise forecast analysis.")

    yearly_docs = []
    for file in files:
        content = await file.read()
        if not content:
            continue
        yearly_docs.append(extract_document(content, file.filename or "upload"))
    if len(yearly_docs) < 2:
        raise HTTPException(status_code=400, detail="At least 2 valid (non-empty) trial balance files are required.")

    parsed_ops = None
    if operational_inputs:
        try:
            parsed_ops = json.loads(operational_inputs)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="operational_inputs must be valid JSON (a list of per-year driver objects).")

    generator = EnterpriseReportGenerator()
    report = generator.generate(yearly_docs, operational_inputs=parsed_ops,
                                 discount_rate=discount_rate, terminal_growth=terminal_growth)
    wb = generator.build_workbook(report)
    filename = save_workbook_and_get_filename(wb, prefix="enterprise_forecast")

    return JSONResponse({"filename": filename, "report": jsonable_encoder(report)})
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("core:app", host="0.0.0.0", port=8000, reload=True)

# =====================================================================================
# SECTION 8: INTERACTIVE FINANCIAL LEDGER MODE (STATEFUL)
# =====================================================================================
# A dynamic, stateful accounting ledger session. Unlike the one-shot Statement Generator,
# this maintains an in-memory accounting state per session that persists across turns:
# every accepted adjustment (add/remove/increase/decrease/replace/reclassify/merge), plus
# undo/reset, updates the ledger and triggers a full deterministic recalculation of the
# trial balance and every ratio. No figure is ever estimated by an LLM — all math is plain
# Python arithmetic built on the same compute_core_financials / compute_full_analytics
# functions used by the Statement Generator, so the two stay consistent.

import copy

LEDGER_SESSIONS: Dict[str, Dict[str, Any]] = {}

LEDGER_DEBIT_CATEGORIES = {"asset", "cogs", "expense"}


def _ledger_new_account(name: str, category: str, amount: float, source: str) -> Dict[str, Any]:
    return {
        "account_name": name.strip(),
        "account_code": None,
        "category": category,
        "amount": round(float(amount), 2),
        "source_document": source,
        "last_modified": datetime.now().isoformat(timespec="seconds"),
        "modification_history": ["Created"],
    }


def _ledger_parse_number(raw: str) -> float:
    return float(raw.replace(",", "").replace("$", "").strip())


def _ledger_find_account_idx(accounts: List[Dict[str, Any]], query: str) -> Optional[int]:
    import difflib
    q = query.strip().lower().strip(" .")
    if not q:
        return None
    # 1) exact match
    for i, a in enumerate(accounts):
        if a["account_name"].strip().lower() == q:
            return i
    # 2) substring match either direction
    candidates = [i for i, a in enumerate(accounts)
                  if q in a["account_name"].lower() or a["account_name"].lower() in q]
    if candidates:
        return candidates[0]
    # 3) stem-stripped word-overlap match (handles "salaries" vs "Salary Expense")
    strip_words = {"expense", "expenses", "account", "cost", "costs"}
    def stem_tokens(s):
        return {w[:5] for w in re.findall(r"[a-z]+", s.lower()) if w not in strip_words}
    q_tokens = stem_tokens(q)
    if q_tokens:
        for i, a in enumerate(accounts):
            a_tokens = stem_tokens(a["account_name"])
            if q_tokens & a_tokens:
                return i
    # 4) fuzzy closest-match fallback
    names = [a["account_name"] for a in accounts]
    close = difflib.get_close_matches(query.strip(), names, n=1, cutoff=0.72)
    if close:
        return next(i for i, a in enumerate(accounts) if a["account_name"] == close[0])
    return None


def _ledger_infer_category(name: str) -> str:
    n = name.lower()
    if "owner" in n and any(k in n for k in ("investment", "capital", "contribution")):
        return "equity"
    return _classify_account(name)


def _ledger_apply_command(session: Dict[str, Any], raw_command: str) -> str:
    """Parses a plain-English accounting command, applies it deterministically to the
    session's ledger state, and returns a human-readable description of what changed.
    Raises ValueError with a clarification message if the command cannot be understood
    or references an account that cannot be found — per spec, Quanto never guesses."""
    cmd = raw_command.strip()
    if not cmd:
        raise ValueError("Please enter an accounting command (e.g. 'increase rent to 15,000').")
    cmd_l = cmd.lower()
    accounts = session["accounts"]
    pre_snapshot = copy.deepcopy(accounts)

    def commit(description: str) -> str:
        session["history"].append({"description": description, "snapshot": pre_snapshot,
                                    "timestamp": datetime.now().isoformat(timespec="seconds")})
        return description

    # --- reset ---
    if re.match(r'^reset\b', cmd_l):
        session["accounts"] = copy.deepcopy(session["original_accounts"])
        session["history"] = []
        return "Reset to original trial balance."

    # --- undo / reverse ---
    if re.match(r'^(undo|reverse)\b', cmd_l):
        if not session["history"]:
            raise ValueError("There is no prior adjustment to undo.")
        last = session["history"].pop()
        session["accounts"] = last["snapshot"]
        return f"Reversed: {last['description']}"

    # --- replace X with Y ---
    m = re.match(r'^replace\s+(.+?)\s+with\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        old_name, new_name = m.group(1).strip(), m.group(2).strip()
        idx = _ledger_find_account_idx(accounts, old_name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{old_name}'. Please check the account name.")
        accounts[idx]["modification_history"].append(f"Renamed from '{accounts[idx]['account_name']}' to '{new_name}'")
        accounts[idx]["account_name"] = new_name
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"Replaced '{old_name}' with '{new_name}'")

    # --- merge X into Y ---
    m = re.match(r'^merge\s+(.+?)\s+into\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        src_name, dst_name = m.group(1).strip(), m.group(2).strip()
        src_idx = _ledger_find_account_idx(accounts, src_name)
        if src_idx is None:
            raise ValueError(f"I couldn't find an account matching '{src_name}'.")
        dst_idx = _ledger_find_account_idx(accounts, dst_name)
        if dst_idx is None:
            accounts.append(_ledger_new_account(dst_name, accounts[src_idx]["category"], 0.0, "user adjustment"))
            dst_idx = len(accounts) - 1
        moved_name, moved_amount = accounts[src_idx]["account_name"], accounts[src_idx]["amount"]
        accounts[dst_idx]["amount"] = round(accounts[dst_idx]["amount"] + moved_amount, 2)
        accounts[dst_idx]["modification_history"].append(f"Absorbed {moved_name} ({moved_amount:,.2f})")
        accounts.pop(src_idx)
        return commit(f"Merged '{moved_name}' into '{accounts[dst_idx]['account_name']}'")

    # --- remove / delete X ---
    m = re.match(r'^(?:remove|delete)\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'. Please check the account name.")
        removed = accounts.pop(idx)
        return commit(f"Removed {removed['account_name']} ({removed['amount']:,.2f})")

    # --- reclassify X as <category> ---
    m = re.match(r'^reclassify\s+(.+?)\s+as\s+(.+)$', cmd, re.IGNORECASE)
    if m:
        name, new_cat_raw = m.group(1).strip(), m.group(2).strip().lower()
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'.")
        valid_cats = ["asset", "liability", "equity", "revenue", "cogs", "expense"]
        new_cat = next((c for c in valid_cats if c in new_cat_raw), None)
        if not new_cat:
            raise ValueError(f"I couldn't determine a valid category from '{new_cat_raw}'. "
                              f"Use one of: {', '.join(valid_cats)}.")
        old_cat = accounts[idx]["category"]
        accounts[idx]["category"] = new_cat
        accounts[idx]["modification_history"].append(f"Reclassified from {old_cat} to {new_cat}")
        return commit(f"Reclassified {accounts[idx]['account_name']} from {old_cat} to {new_cat}")

    # --- change/set/increase/decrease X to Y (absolute set) ---
    m = re.match(r'^(?:change|set|increase|decrease)\s+(.+?)\s+to\s+([\d,\.]+)$', cmd, re.IGNORECASE)
    if m:
        name, amount_str = m.group(1).strip(), m.group(2)
        new_amount = round(_ledger_parse_number(amount_str), 2)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            accounts.append(_ledger_new_account(name, _ledger_infer_category(name), new_amount, "user adjustment"))
            return commit(f"Added {name} at {new_amount:,.2f}")
        old_amount = accounts[idx]["amount"]
        accounts[idx]["amount"] = new_amount
        accounts[idx]["modification_history"].append(f"Changed from {old_amount:,.2f} to {new_amount:,.2f}")
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"Changed {accounts[idx]['account_name']} from {old_amount:,.2f} to {new_amount:,.2f}")

    # --- increase/decrease X by Y (amount or %) ---
    m = re.match(r'^(increase|decrease)\s+(.+?)\s+by\s+([\d,\.]+)\s*(%|percent)?$', cmd, re.IGNORECASE)
    if m:
        direction, name, amount_str, pct_flag = m.group(1).lower(), m.group(2).strip(), m.group(3), m.group(4)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is None:
            raise ValueError(f"I couldn't find an account matching '{name}'. Please check the account name.")
        old_amount = accounts[idx]["amount"]
        delta_value = _ledger_parse_number(amount_str)
        delta = old_amount * (delta_value / 100.0) if pct_flag else delta_value
        new_amount = old_amount + delta if direction == "increase" else old_amount - delta
        new_amount = round(max(new_amount, 0.0), 2)
        accounts[idx]["amount"] = new_amount
        accounts[idx]["modification_history"].append(f"{direction.title()}d from {old_amount:,.2f} to {new_amount:,.2f}")
        accounts[idx]["last_modified"] = datetime.now().isoformat(timespec="seconds")
        return commit(f"{direction.title()}d {accounts[idx]['account_name']} from {old_amount:,.2f} to {new_amount:,.2f}")

    # --- add X of Y / record X of Y ---
    m = re.match(r'^(?:add|record)\s+(.+?)\s+of\s+([\d,\.]+)$', cmd, re.IGNORECASE)
    if m:
        name, amount_str = m.group(1).strip(), m.group(2)
        amount = round(_ledger_parse_number(amount_str), 2)
        idx = _ledger_find_account_idx(accounts, name)
        if idx is not None:
            old_amount = accounts[idx]["amount"]
            accounts[idx]["amount"] = round(old_amount + amount, 2)
            accounts[idx]["modification_history"].append(f"Added {amount:,.2f} (was {old_amount:,.2f})")
            return commit(f"Added {amount:,.2f} to existing account {accounts[idx]['account_name']}")
        category = _ledger_infer_category(name)
        accounts.append(_ledger_new_account(name, category, amount, "user adjustment"))
        return commit(f"Added new account '{name}' of {amount:,.2f} ({category})")

    raise ValueError(
        "I couldn't parse that as an accounting command. Try formats like: 'remove rent expense', "
        "'increase salaries to 50,000', 'decrease advertising by 10%', 'add fuel expense of 2,500', "
        "'replace utilities with electricity expense', 'record depreciation of 7,000', "
        "'reclassify loan as liability', 'merge travel into operating expenses', 'undo', or 'reset'."
    )


def _ledger_format_state(session: Dict[str, Any]) -> str:
    """Renders the full required output format: Ledger State, Trial Balance, Ratios, and
    Modification History — all computed fresh from the current session accounts."""
    accounts = session["accounts"]
    f = compute_core_financials(accounts)
    analytics = compute_full_analytics(accounts)

    def pct(v): return f"{v*100:.1f}%" if v is not None else "N/A"
    def xr(v): return f"{v:.2f}x" if v is not None else "N/A"

    lines: List[str] = []
    lines.append(f"{session.get('company','Company')} · {session.get('period','')}".strip(" ·"))
    lines.append("")
    lines.append("UPDATED FINANCIAL LEDGER STATE")
    lines.append("-" * 62)
    for a in accounts:
        lines.append(f"  {a['account_name']:<34} {a['amount']:>15,.2f}   [{a['category']}]")
    lines.append("-" * 62)
    lines.append(f"  {'Total Revenue':<34} {f['revenue']:>15,.2f}")
    lines.append(f"  {'Total COGS':<34} {f['cogs']:>15,.2f}")
    lines.append(f"  {'Gross Profit':<34} {f['gross_profit']:>15,.2f}")
    lines.append(f"  {'Total Operating Expenses':<34} {f['operating_expenses']:>15,.2f}")
    lines.append(f"  {'Operating Income (EBIT)':<34} {f['ebit']:>15,.2f}")
    lines.append(f"  {'Net Income':<34} {f['net_income']:>15,.2f}")
    lines.append("")

    lines.append("CURRENT TRIAL BALANCE")
    lines.append("-" * 62)
    total_debit = total_credit = 0.0
    for a in accounts:
        is_debit = a["category"] in LEDGER_DEBIT_CATEGORIES
        amt = abs(a["amount"])
        if is_debit:
            total_debit += amt
            lines.append(f"  {a['account_name']:<34} DR {amt:>13,.2f}")
        else:
            total_credit += amt
            lines.append(f"  {a['account_name']:<34}    CR {amt:>13,.2f}")
    balance_diff = abs(total_debit - total_credit)
    lines.append("-" * 62)
    lines.append(f"  {'TOTALS':<34} DR {total_debit:>13,.2f} / CR {total_credit:>13,.2f}")
    lines.append("  ✓ Trial balance is in balance." if balance_diff < 0.01
                  else f"  ⚠ Out of balance by {balance_diff:,.2f} — review the accounts above.")
    lines.append("")

    lines.append("PERFORMANCE RATIOS")
    lines.append("-" * 62)
    lines.append(f"  Net Margin Ratio:              {pct(analytics['net_margin'])}")
    lines.append(f"  Gross Margin Ratio:            {pct(analytics['gross_margin'])}")
    lines.append(f"  Operating Margin:               {pct(analytics['operating_margin'])}")
    lines.append(f"  Current Ratio:                  {xr(analytics['current_ratio'])}")
    lines.append(f"  Quick Ratio:                    {xr(analytics['quick_ratio'])}")
    lines.append(f"  Debt-to-Equity:                 {xr(analytics['debt_to_equity'])}")
    lines.append(f"  Return on Assets (ROA):         {pct(analytics['roa'])}")
    lines.append(f"  Return on Equity (ROE):         {pct(analytics['roe'])}")
    lines.append("")

    lines.append("MODIFICATION HISTORY")
    lines.append("-" * 62)
    if session["history"]:
        for h in session["history"]:
            lines.append(f"  ✓ {h['description']}")
    else:
        lines.append("  (No adjustments yet — this is the original extracted trial balance.)")
    lines.append("")
    lines.append("Ready for additional accounting adjustments. Continue using natural language.")
    lines.append("Quanto is not responsible for financial decisions.")
    return "\n".join(lines)


# ---- Interactive Ledger Mode API routes ----

@app.post("/api/ledger/init")
async def api_ledger_init(file: UploadFile = File(...)):
    """Initializes a new stateful ledger session from an uploaded trial balance / ledger
    document. Extraction reuses the same OCR pipeline as the Statement Generator so figures
    stay consistent across features."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    doc = extract_document(content, file.filename or "upload")
    raw_accounts = doc.get("accounts", [])
    if not raw_accounts:
        raise HTTPException(status_code=422, detail="No accounts could be extracted from this document.")

    accounts: List[Dict[str, Any]] = []
    for a in raw_accounts:
        accounts.append({
            "account_name": a.get("account_name", "Unnamed Account"),
            "account_code": None,
            "category": a.get("category") or _classify_account(a.get("account_name", "")),
            "amount": round(abs(float(a.get("amount", 0.0))), 2),
            "source_document": file.filename or "upload",
            "last_modified": datetime.now().isoformat(timespec="seconds"),
            "modification_history": ["Extracted from source document"],
        })

    session_id = uuid.uuid4().hex[:16]
    company = doc.get("company_name") or "Unnamed Company"
    period = doc.get("period_label") or datetime.now().strftime("FY%Y")

    session = {
        "session_id": session_id,
        "company": company,
        "period": period,
        "accounts": accounts,
        "original_accounts": copy.deepcopy(accounts),
        "history": [],
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    LEDGER_SESSIONS[session_id] = session

    return JSONResponse({
        "session_id": session_id,
        "company": company,
        "period": period,
        "accounts_found": len(accounts),
        "formatted_state": _ledger_format_state(session),
    })


@app.post("/api/ledger/command")
async def api_ledger_command(payload: Dict[str, Any]):
    """Applies a single natural-language accounting command to an existing ledger session,
    recalculates the trial balance and every ratio deterministically, and returns the
    updated full state plus the running modification history."""
    session_id = payload.get("session_id", "")
    command = payload.get("command", "")
    session = LEDGER_SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Ledger session not found or has expired. Please upload a trial balance to start a new session.")
    if not command or not command.strip():
        raise HTTPException(status_code=400, detail="Command cannot be empty.")

    try:
        _ledger_apply_command(session, command)
    except ValueError as e:
        # Clarification needed — do not mutate state, just surface the guidance alongside the current state.
        return JSONResponse({
            "session_id": session_id,
            "applied": False,
            "clarification": str(e),
            "formatted_state": str(e) + "\n\n" + _ledger_format_state(session),
        })

    return JSONResponse({
        "session_id": session_id,
        "applied": True,
        "formatted_state": _ledger_format_state(session),
    })


@app.get("/api/download/{filename}")
async def api_download(filename: str):
    safe_name = Path(filename).name  # prevent path traversal
    filepath = OUTPUT_DIR / safe_name
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found or has expired.")
    return FileResponse(
        path=filepath,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("core:app", host="0.0.0.0", port=8000, reload=True)
